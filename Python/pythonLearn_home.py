>>> from openpyxl import Workbook
>>> wb = Workbook()
>>> ws = wb.get_active_sheet()
>>> ws1 = wb.create_sheet() # insert at the end (default)
# or
>>> ws2 = wb.create_sheet(0) # insert at first position
ws.title = "New Title"
>>> ws3 = wb.get_sheet_by_name("New Title")
>>> ws is ws3
>>> print wb.get_sheet_names() #['Sheet2', 'New Title', 'Sheet1']

Warning Because of this feature, scrolling through cells instead of accessing them directly will create them all in memory, even if you don��t assign them a value.
Something like
>>> for i in xrange(0,100):
...             for j in xrange(0,100):
...                     ws.cell(row = i, column = j)
will create 100x100 cells in memory, for nothing.
However, there is a way to clean all those unwanted cells, we��ll see that later.

>>> ws.range('A1:C2')
((<Cell Sheet1.A1>, <Cell Sheet1.B1>, <Cell Sheet1.C1>),
(<Cell Sheet1.A2>, <Cell Sheet1.B2>, <Cell Sheet1.C2>))

>>> for row in ws.range('A1:C2'):
...             for cell in row:
...                     print cell
<Cell Sheet1.A1>
<Cell Sheet1.B1>
<Cell Sheet1.C1>
<Cell Sheet1.A2>
<Cell Sheet1.B2>
<Cell Sheet1.C2>

If you need to iterate through all the rows or columns of a file, you can instead use the openpyxl.worksheet.Worksheet.rows() property:
>>> ws = wb.get_active_sheet()
>>> ws.cell('C9').value = 'hello world'
>>> ws.rows
((<Cell Sheet.A1>, <Cell Sheet.B1>, <Cell Sheet.C1>),
(<Cell Sheet.A2>, <Cell Sheet.B2>, <Cell Sheet.C2>),
(<Cell Sheet.A3>, <Cell Sheet.B3>, <Cell Sheet.C3>),
(<Cell Sheet.A4>, <Cell Sheet.B4>, <Cell Sheet.C4>),
(<Cell Sheet.A5>, <Cell Sheet.B5>, <Cell Sheet.C5>),
(<Cell Sheet.A6>, <Cell Sheet.B6>, <Cell Sheet.C6>),
(<Cell Sheet.A7>, <Cell Sheet.B7>, <Cell Sheet.C7>),
(<Cell Sheet.A8>, <Cell Sheet.B8>, <Cell Sheet.C8>),
(<Cell Sheet.A9>, <Cell Sheet.B9>, <Cell Sheet.C9>))

>>> c.value = 'hello, world'
>>> print c.value
'hello, world'

>>> d.value = 3.14
>>> print d.value
3.14

>>> c.value = '12%'
>>> print c.value
0.12

>>> import datetime
>>> d.value = datetime.datetime.now()
>>> print d.value
datetime.datetime(2010, 9, 10, 22, 25, 18)

>>> c.value = '31.50'
>>> print c.value
31.5

>>> wb = Workbook()
>>> wb.save('balances.xlsx')
>>> from openpyxl import load_workbook
>>> wb2 = load_workbook('test.xlsx')
>>> print wb2.get_sheet_names()
['Sheet2', 'New Title', 'Sheet1']

from openpyxl import load_workbook
wb = load_workbook(filename = r'empty_book.xlsx')
sheet_ranges = wb.get_sheet_by_name(name = 'range names')
print sheet_ranges.cell('D18').value # D18	


#��Ԫ����ʽ����
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

if __name__ == '__main__':
    # https://www.cnblogs.com/BlueSkyyj/p/7571787.html
    # https://blog.csdn.net/weixin_41595432/article/details/79349995
    # https://www.cnblogs.com/Unikfox/p/9124767.html

    wb = Workbook()
    ws = wb.active
    ws.title = "�ҵ�ҳǩ"

    ws.cell(row=1, column=1).value = 99
    ws.cell(row=2, column=1, value=100)
    ws['A3'] = 4  # write
    ws['A4'] = u"�����ǵ�Ԫ��"
    ws.append(['This is A1', 'This is B1', 'This is C1'])
    ws.append({'A': 'This is A1', 'C': 'This is C1'})
    ws.append({1: 'This is A1', 3: 'This is C1'})

    # ���õ�Ԫ������
    font1 = Font(name='Arial',
                 size=20,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
    ws['A3'].font = font1
    ws['A4'].font = font1
    ws['A5'].font = font1
    # row = ws.row_dimensions[6]
    # row.font = font

    # ���õ�Ԫ��߿�
    border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                               color='FF000000'),
                    top=Side(border_style='thick',
                             color='FF000000'),
                    bottom=Side(border_style='thick',
                                color='FF000000'))
    ws['A3'].border = border

    #���ö��뷽ʽ
    alignment1 = Alignment(horizontal='right',
                           vertical='bottom',
                           text_rotation=0,
                           wrap_text=True,
                           shrink_to_fit=False,
                           indent=0)
    ws['B5'].alignment = alignment1

    alignment2 = Alignment(horizontal='left',
                           vertical='center',
                           text_rotation=0,
                           wrap_text=True,
                           shrink_to_fit=False,
                           indent=0)
    ws['C5'].alignment = alignment2

    # �����иߺ��п�
    ws.row_dimensions[6].height = 50
    ws.column_dimensions['C'].width = 100

    wb.save(filename="testExcel.xlsx")
#-----------------------------------------------------------------------------------------
#from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

if __name__ == '__main__':
    #wb1=Workbook()
    wb = load_workbook("db.xlsx")
    ws = wb.get_active_sheet()
    #ws1.title="sheet2"
    #ws2 = wb.get_sheet_by_name("sheet2")
    #print ws2 is ws1 
    #print wb.get_sheet_names()

    cell = ws.cell("B3")
    cell.value = datetime.datetime.now()
    print cell.value
#    print ws.range("A1:B3")
#    for row in ws.range("A1:B3"):
#        for cell in row:
#            print cell
    print ws.rows
    print ws.columns
    wb.save("db.xlsx")
#-----------------------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
wb = Workbook()
dest_filename = r'empty_book.xlsx'
ws = wb.worksheets[0]
ws.title = "range names"
for col_idx in xrange(1, 3):
    col = get_column_letter(col_idx)
    for row in xrange(1, 6):
        ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)
ws = wb.create_sheet()
ws.title = 'Pi'
ws.cell('F5').value = 3.14
wb.save(filename = dest_filename)
#-----------------------------------------------------------------------------------------
import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.worksheets[0]
# set date using a Python datetime
ws.cell('A1').value = datetime.datetime(2010, 7, 21)
print ws.cell('A1').style.number_format.format_code # returns 'yyyy-mm-dd'
# set percentage using a string followed by the percent sign
ws.cell('B1').value = '3.14%'
print ws.cell('B1').value # returns 0.031400000000000004
print ws.cell('B1').style.number_format.format_code # returns '0%'
wb.save("db.xlsx")
#-----------------------------------------------------------------------------------------
#û������ͨ����
from openpyxl import load_workbook
wb = load_workbook(filename = 'db.xlsx', use_iterators = True)
ws = wb.get_sheet_by_name(name = 'sheet') # ws is now an IterableWorksheet
for row in ws.iter_rows(): # it brings a new method: iter_rows()
    for cell in row:
        print cell.internal_value
#-----------------------------------------------------------------------------------------	
from openpyxl import Workbook
wb = Workbook(optimized_write = True)
ws = wb.create_sheet()
# now we'll fill it with 10k rows x 200 columns
for irow in xrange(10):
    ws.append(['%d' % i for i in xrange(10)])
wb.save('db.xlsx') # don't forget to save !
#-----------------------------------------------------------------------------------------
Warning When creating your workbook using optimized_write set to True, you will only be able to call this function once. Subsequents attempts to modify or save the file will raise an openpyxl.shared.exc.WorkbookAlreadySaved exception.

from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
#print wb.get_named_ranges()
ws = wb.create_sheet(title = "test")
ws.cell("A1").value = "A1"
print wb.get_index(ws)
print wb.get_sheet_names()
wb.remove_sheet(ws)
print wb.get_sheet_names()
#wb.add_sheet(ws)
#wb.add_sheet(ws, 1)
wb.save('db.xlsx') # don't forget to save !
#-----------------------------------------------------------------------------------------
append(list_or_dict)[source]
Appends a group of values at the bottom of the current sheet.

If it��s a list: all values are added in order, starting from the first column
If it��s a dict: values are assigned to the columns indicated by the keys (numbers or letters)
Parameters:	
list_or_dict (list/tuple or dict) �C list or dict containing values to append
Usage:

append([��This is A1��, ��This is B1��, ��This is C1��])
or append({��A�� : ��This is A1��, ��C�� : ��This is C1��})
or append({0 : ��This is A1��, 2 : ��This is C1��})
Raise :	TypeError when list_or_dict is neither a list/tuple nor a dict
#-----------------------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
#print wb.get_named_ranges()
ws = wb.create_sheet(title = "test")
ws.cell("A1").value = "A1"
ws.cell("B2").value = "B2"
print ws.calculate_dimension()
rangeStr = ws.calculate_dimension()
print rangeStr.find(":")
wb.save('db.xlsx') # don't forget to save !
#-----------------------------------------------------------------------------------------
cell(coordinate=None, row=None, column=None)[source]
Returns a cell object based on the given coordinates.

Usage: cell(coodinate=��A15��) or cell(row=15, column=1)

If coordinates are not given, then row and column must be given.

Cells are kept in a dictionary which is empty at the worksheet creation. Calling cell creates the cell in memory when they are first accessed, to reduce memory usage.

Parameters:	
coordinate (string) �C coordinates of the cell (e.g. ��B12��)
row (int) �C row index of the cell (e.g. 4)
column (int) �C column index of the cell (e.g. 3)
Raise :	
InsufficientCoordinatesException when coordinate or (row and column) are not given
Return type:	
openpyxl.cell.Cell
#-----------------------------------------------------------------------------------------
TYPE_STRING = 's'
TYPE_FORMULA = 'f'
TYPE_NUMERIC = 'n'
TYPE_BOOL = 'b'
TYPE_NULL = 's'
TYPE_INLINE = 'inlineStr'
TYPE_ERROR = 'e'
TYPE_FORMULA_CACHE_STRING = 'str'

from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
#print wb.get_named_ranges()
ws = wb.create_sheet(title = "test")
ws.cell("A1").value = "A1"
ws.cell("B2").value = "B2"
ws.cell("B3").value = "B3"
print ws.get_cell_collection()
cellA1 = ws.get_cell_collection()[0]
print cellA1.value
print ws.get_highest_column()
print ws.get_highest_row()
print ws.range("A1:B2", 1, 2)
cellB2 = ws.range("A1:A1", 1, 1)
print cellB2
print cellB2[0][0].value
print ws.title
ws.title = "hello"
print cellA1.address
print cellA1.get_coordinate()
print cellA1.data_type
print cellA1.bind_value("world")
print cellA1.check_numeric("123")
cellA1.value = cellA1.check_numeric("123")
print cellA1.data_type
cellA1.value = cellA1.check_string("abc")
print cellA1.data_type
#print cellA1.has_style()
print cellA1.hyperlink
print cellA1.is_date()
anotherCell = cellA1.offset(1, 1)
print anotherCell.address
print cellA1.style
cellA1.set_value_explicit(value = "123.4000", data_type = "s")
#cellA1.set_value_explicit(data_type = "n")
wb.save('db.xlsx') # don't forget to save !
#-----------------------------------------------------------------------------------------
list("hello")

    sequence = [None] * 5
    sequence[2] = 3
    sequence[4] = "what"
    print sequence
	
	name = list("perl")
    name[1:] = list("ython")
    print name 
	
	name = list("perl")
    name[1:1] = list("ABC")
    print name 
	
    name = list("perl")
    name[1:3] = []
    print name 

    name = list("helloworld")
    name[1::2] = list("ABCDE")#Ҫ������ֶ�������Ͱ��ղ����ָ�����ͬ
    print name 	
	
    name = list("helloworld")
    print name.count("l")

    a = [1, 2, 3]
    b = [4, 5, 6]
    a.extend(b)
	#a[len(a):] = b #�ɶ��Բ���extend
    print a	
    print a.index(2)
	a.insert(2, "four")#a[2:2] = ["four"]#�ɶ��Բ���insert
	a.pop()
	a.pop(2)#pop������Ψһһ�������޸��б��ַ���Ԫ��ֵ������None�����б���
	a.remove(1)
	a.reverse()#list(reversed(a))
	a.sort()
	
	a = ["a", 1, 3, 2]
    b = a[:]#����Ч�ʵĸ��������б�ķ�����ֻ�Ǽ򵥵İ�a��ֵ��b��û�õ�
    b.sort()
	
	��һ�ֻ�ȡ��������б��������ǣ����������κ����У��ܷ���һ���б�
	b = sorted(a)
	
	b.sort(cmp)#����Ĭ�ϵıȽϺ������õ����ڽ�����cmp
	
	b = ["a", "abc", "de"]
    b.sort()
    print b
    b.sort(key = len)
    print b
	
	b = ["a", "abc", "de"]
    b.sort()
    print b
    b.sort(key = len, reverse = True)
    print b
#-----------------------------------------------------------------------------------------
#pythonѹ���ͽ�ѹ��
import zipfile
    f = zipfile.ZipFile("db.zip", "w", zipfile.ZIP_DEFLATED)
    f.write("db.xlsx")
    f.write("main.py")
    f.close()
	
    f = zipfile.ZipFile("db.zip", "r")
    f.extractall("c:\Users\Ben\Desktop", ["db.xlsx"])#�����ڶ����������ѹ�������ļ�
    f.close()

    newtxt = open("new.txt", "w")
    newtxt.write("this is a new txt\n")
    newtxt.close()
    f = zipfile.ZipFile("db.zip", "a", zipfile.ZIP_DEFLATED)
    f.write("new.txt")
    f.close()

	print zipfile.is_zipfile("db.zip")	
	
    f = zipfile.ZipFile("db.zip", "a", zipfile.ZIP_DEFLATED)
    f.close()
    print f.namelist()	
	print f.printdir()	
	
	
    f = zipfile.ZipFile("db.zip", "w", zipfile.ZIP_DEFLATED)
    f.write("db.xlsx")
    f.write("main.py")
	#f.close()
    #f = zipfile.ZipFile("db.zip", "r")
    txtFile = f.open("main.py")
    lines = txtFile.readlines()
    for line in lines:
        line = line.rstrip()
        print line
    txtFile.close()
    f.close()
	
    f = zipfile.ZipFile("db.zip", "a", zipfile.ZIP_DEFLATED)
    f.write("db.xlsx", "dbdir/db.xlsx")
    f.close()	
if __name__ == '__main__':
    f = zipfile.ZipFile("db.zip", "r", zipfile.ZIP_DEFLATED)
    fileMain = f.read("main.py")#���������fileMain�ǡ�str����������ĳ�fileMain = f.read("db.xlsx")����ʱ�����Ķ�����excel�ļ���ӡ����������
    print fileMain
    f.close()
	
#��zip.exeʵ��ѹ��
#����Ŀ¼������������Ŀ¼��δ����ѹ������
import os
if __name__ == '__main__':
    cmdOrder = r'zip -rJ "d:\Program Files\eclipse workspace\PyLearngingProgram\main\abc.zip" "d:\Program Files\eclipse workspace\PyLearngingProgram\main\dbDir\*.*"' #��ʱ����Ҫ��zip.exe����python����Ŀ¼���߻�����������Ŀ¼��
    os.system(cmdOrder)
	
#���������Ŀ¼����������Ҫ������Ӧ���ļ����У�����Ҫ���ļ����������
# encoding=utf-8
import os
if __name__ == '__main__':
    dstDir = r"d:\Program Files\eclipse workspace\PyLearngingProgram\main\dbDir"
    curDir = os.getcwd()
    os.chdir(dstDir)
    cmdOrder = r'zip -rJ "abc.zip" *.*' #��ʱ��һ��Ҫ��zip.exe���뻷����������Ŀ¼�У���Ϊ��ʱ����ǰĿ¼���Ѿ��л���Ҫ�����Ŀ¼���ˣ�zip.exe������ֱ����������
    os.system(cmdOrder)
    os.chdir(curDir)
	#����shutil.copy2��abc.zip������Ŀ¼�м���

#���zip.exe��ѹ������������Ѿ�����ѹ�����ˣ��ٶԸ�ѹ��������zip��������ֻ����²���������ѹ�����е������ļ�����update��û�е��ļ�����add�������Ὣ���ɵ�ѹ�����滻��������ʱ���ͬ����.zip�ļ����ᱻ�ٴδ򵽰��У��������other.zip���ƽ��д������ղŵ�ѹ�����ᱻ�����°�other.zip�С�
#-----------------------------------------------------------------------------------------
#pythonĿ¼����
import os

if __name__ == '__main__':
    currentPath = os.getcwd()
    print currentPath
    for fileName in os.listdir(currentPath):
        print fileName
		
if __name__ == '__main__':
    currentPath = os.getcwd()
    print currentPath
    os.mkdir("%s/newDir" % currentPath)
	print os.path.isdir("%s/newDir" % currentPath)#True
	print os.path.isfile("%s/newDir" % currentPath)#False
	print os.path.islink("%s/newDir" % currentPath)#False	
	os.rmdir("%s/newDir" % currentPath)#��Ҫ˵�����ǣ�ʹ��os.rmdirɾ����Ŀ¼����Ϊ��Ŀ¼������������
	
	
	time_of_last_access = os.path.getatime(myfile) 
	time_of_last_modification = os.path.getmtime(myfile) 
	size = os.path.getsize(myfile) 	
	
	#�����ʱ������Ϊ��λ�����Ҵ�1970��1��1�տ�ʼ����Ϊ�˻�ȡ����Ϊ��λ�����������ڣ�����ʹ�����д��룺 
	time_of_last_access = os.path.getatime("db.xlsx") 
    age_in_days = (time.time()-time_of_last_access)/(60*60*24) 
    print age_in_days
	print os.access("db.xlsx", os.R_OK|os.W_OK|os.X_OK)
	
	os.remove("db.zip")

import shutil	
shutil.rmtree("newdir")

filelist = glob.glob('*.xlsx') + glob.glob('*.py')
print filelist 
print isinstance(filelist, list)
print isinstance(filelist, str) 
#-----------------------------------------------------------------------------------------
import os
import shutil
import glob
import operator

def remove(files):
    if isinstance(files,str):
        files = [files]
    if not isinstance(files, list):
	#if not operator.isSequenceType(files):#���õĲ�������
        return
    for eachFile in files:
        if os.path.isdir(eachFile):
            shutil.rmtree(eachFile)
        elif os.path.isfile(eachFile):
            os.remove(eachFile)     

if __name__ == '__main__':
    for i in range(2):
        os.mkdir("tmp_%s" % str(i))
        f = open("tmp__%s.txt" % str(i), "w")
        f.close()
    remove("tmp_1")
    remove(glob.glob("tmp_[0-1]") + glob.glob("tmp__[0-1].txt")) 
	
import shutil
import os

if __name__ == '__main__':
	shutil.copy("db.xlsx", "newDb.xlsx")
    shutil.copy2("db.xlsx", "newDb.xlsx")#����������ʱ�������޸�ʱ��
	shutil.copytree("dbdir", "newDbdir", True)
	
	shutil.copy(os.path.join(os.curdir, "db.xlsx"), os.pardir)
	os.rename("dbdir", "dbdir2")#�ļ��л��ļ�
#-----------------------------------------------------------------------------------------
#�������Ҳ����������ͬ���ļ�ϵͳ֮���ƶ��ļ���������ǽ�myfile�ƶ���Ŀ¼d���棺 
os.rename(myfile, os.path.join(d, myfile)) 

#�ڿ��ļ�ϵͳ�ƶ��ļ���ʱ�򣬿�����ʹ��shutil.copy2�������ļ���Ȼ����ɾ��ԭ���ĸ������ɣ����£� 
shutil.copy2(myfile, os.path.join(d, myfile)) 
os.remove(myfile)
���������ƶ��ļ��ķ������ȫ�ġ� 
#-----------------------------------------------------------------------------------------
import os
import shutil
if __name__ == '__main__':
    currentPath = os.getcwd()
    dbXlsx = os.path.join(currentPath, "db.xlsx")
    print dbXlsx
    baseName = os.path.basename(dbXlsx)#�õ��ļ�����Ŀ¼��������չ��
    dirName = os.path.dirname(dbXlsx)#�õ���Ŀ¼��
	#��dirName, baseName = os.path.split(dbXlsx)
    print baseName
    print dirName
    print os.path.dirname(dirName)
	root, extension = os.path.splitext(dbXlsx)
    print root
    print extension 
������£�
D:\Program Files\eclipse workspace\PyLearngingProgram\main\db.xlsx
db.xlsx
D:\Program Files\eclipse workspace\PyLearngingProgram\main
D:\Program Files\eclipse workspace\PyLearngingProgram
D:\Program Files\eclipse workspace\PyLearngingProgram\main\db
.xlsx

������fname�е���չ�����ּ�.py����������extension�������ಿ���򸳸��˱���root�������õ�������ŵ���չ���Ļ���ֻ��ʹ��os.path.splitext(fname)[1][1:]���ɡ�
����һ���ļ���Ϊf������չ�����⣬���뽫����չ����Ϊext������ʹ������Ĵ��룺 
����newfile = os.path.splitext(f)[0] + ext 
#-----------------------------------------------------------------------------------------
if __name__ == '__main__':
    currentPath = os.getcwd()
    print currentPath
    os.chdir(os.pardir)
    print os.getcwd() 

if __name__ == '__main__':	
    currentPath = os.getcwd()
    newDir = os.path.join(currentPath, "newdbdir")
    if not os.path.isdir(newDir):
        os.mkdir(newDir)
    os.chdir(newDir)
    print os.getcwd()	

if __name__ == '__main__':
    currentPath = os.getcwd()
    newDir = os.path.join(currentPath, "a", "b", "c")
    os.makedirs(newDir)	
#-----------------------------------------------------------------------------------------
import os

def showDir(arg, dirName, files):
    print dirName, "has the files", files

if __name__ == '__main__':
    os.path.walk(os.getcwd(), showDir, None)
������£�
D:\Program Files\eclipse workspace\PyLearngingProgram\main has the files ['a', 'db.xlsx', 'dbdir2', 'empty_book.xlsx', 'main.py', 'new.txt', 'newdbdir', '__init__.py']
D:\Program Files\eclipse workspace\PyLearngingProgram\main\a has the files ['b']
D:\Program Files\eclipse workspace\PyLearngingProgram\main\a\b has the files ['c']
D:\Program Files\eclipse workspace\PyLearngingProgram\main\a\b\c has the files []
D:\Program Files\eclipse workspace\PyLearngingProgram\main\dbdir2 has the files ['another.txt']
D:\Program Files\eclipse workspace\PyLearngingProgram\main\newdbdir has the files []

�����У�����arg���Ǳ��裬������os.path.walk����������ȡֵΪNone���ɡ�

import os

def showDir(arg, dirName, files):
    for eachFile in files:
        filePath = os.path.join(dirName, eachFile)
        if os.path.isfile(filePath):
            size = os.path.getsize(filePath)
            #sizeInKB = size / 1000
            arg.append((size, eachFile))

if __name__ == '__main__':
    fileList = []
    os.path.walk(os.getcwd(), showDir, fileList)
������£�
[(5727L, 'db.xlsx'), (5825L, 'empty_book.xlsx'), (457L, 'main.py'), (19L, 'new.txt'), (0L, '__init__.py'), (0L, 'another.txt')]

����dirname�ǵ�ǰ���ڷ��ʵ�Ŀ¼�ľ���·����������files�ڵ��ļ������������dirname�����·�����ڴ��ڼ䣬��ǰ����Ŀ¼��û�иı䣬�Ǿ���˵�ýű���Ȼ���ڽű�����ʱ�����ڵ�Ŀ¼�С������Ϊʲô������Ҫ��filepathŪ�ɴ���dirname��file�ľ���·����ԭ����Ҫ�ı䵱ǰ����Ŀ¼Ϊdirname��ֻҪ�����ÿ��Ŀ¼����os.path.walk�ĺ����е���һ��os.chdir(dirname)��Ȼ���ڸú�����ĩβ���µ���os.chdir(dirname)����ǰ����Ŀ¼�Ļ�ԭֵ���ɣ�������ʾ�� 
def somefunc(arg, dirname, files): 
	origdir = os.getcwd()
	os.chdir(dirname) 
��#...�� 
	os.chdir(origdir)
	
#��һ�ֱ���Ŀ¼�ķ���
# encoding=utf-8
import os

if __name__ == '__main__':
    rootdir = r"d:\Program Files\eclipse workspace\PyLearngingProgram\main"
    for parent, dirnames, filenames in os.walk(rootdir):
        print parent.decode("gbk")
        print dirnames#ûʲô�ô����õ�һ����Ŀ¼�͵������ļ�������ƴ�������ļ���ȫĿ¼
        print filenames
������£�
d:\Program Files\eclipse workspace\PyLearngingProgram\main
['dbDir']
['abc.bmp', 'main.py', 'simple.xls', 'test.sh', 'test.txt', 'test.xls', 'test.xlsx', '__init__.py']
d:\Program Files\eclipse workspace\PyLearngingProgram\main\dbDir
[]
['test.xlsx']		
#-----------------------------------------------------------------------------------------
#�Լ�ʵ��os.path.walk����
import os

def find(func, rootDir, arg = None):
    files = os.listdir(rootDir)
    files.sort(lambda a, b: cmp(a.lower(), b.lower()))
    for eachFile in files:
        fullPath = os.path.join(rootDir, eachFile)
        if os.path.islink(fullPath):
            pass
        elif os.path.isdir(fullPath):
            find(func, fullPath, arg)
        elif os.path.isfile(fullPath):
            func(fullPath, arg)
        else:
            print "find: cannot treat", fullPath

def checkSize(fullPath, fileList):
    size = os.path.getsize(fullPath)
    fileList.append((size, fullPath))

if __name__ == '__main__':
    fileList = []
    currentDir = os.getcwd()
    find(checkSize, currentDir, fileList)
	for fileInfo in fileList:
        print fileInfo
#-----------------------------------------------------------------------------------------
����arg�����˾޴������ԡ����ǿ���ʹ������ͬʱ����������ݺ����ɵ����ݽṹ����һ���������ռ����д���һ���ߴ�Ĵ��й涨��չ�����ļ����ļ����ʹ�С������Ľ�������ļ���С���С�
bigfiles = {'filelist': [], # �ļ����ʹ�С�б� 
'extensions': ('.*ps', '.tiff', '.bmp'), 
'size_limit': 1000000, # 1 Mb 
} 
find(checksize3, os.environ['HOME'], bigfiles) 

import fnmatch # Unix��shell����ͨ���ƥ��
def checksize3(fullpath, arg): 
	treat_file = False 
	ext = os.path.splitext(fullpath)[1] 
	for s in arg['extensions']: 
		if fnmatch.fnmatch(ext, s): 
			treat_file = True # fullpath������ȷ����չ�� 
	size = os.path.getsize(fullpath) 
	if treat_file and size > arg['size_limit']: 
		size = '%.2fMb' % (size/1000000.0) # ��ӡ 
		arg['filelist'].append({'size': size, 'name': fullpath}) 
		
# ���մ�С�����ļ� 
def filesort(a, b): 
	return cmp(float(a['size'][:-2]), float(b['size'][:-2])) 

bigfiles['filelist'].sort(filesort) 
bigfiles['filelist'].reverse() 
for fileinfo in bigfiles['filelist']: 
	print fileinfo['name'], fileinfo['size'] 		
ע��Ϊ�б�����ĺ�����bigfiles['filelist']�����е�ÿ��Ԫ�ؾ���һ���ֵ䣬��size������һ���ַ����������ڽ��бȽ�֮ǰ���Ǳ��뽫��λMb(��������ַ�)ȥ����������ת��Ϊ�������� 	
#-----------------------------------------------------------------------------------------
import copy

class delTest(object):
    def __init__(self):
        print "init"
    
    def __del__(self):
        print "del"    
     

if __name__ == '__main__':
    c = delTest()
    d = copy.deepcopy(c)
    del c
    print "abc"
#-----------------------------------------------------------------------------------------	
class Test(object):
    def __init__(self):
        print "init"
        show()
        Test.show(self)#�������Ա���������������(self)
        
    def show(self):
        print "world"#����ȫ�ֺ���
        show2()#���п��Ե���ȫ�ֺ������޹�λ��
        
def show():
    print "hello"
    
def show2():
    print "nice"   

if __name__ == '__main__':
    c = Test()
    #c.show()
#-----------------------------------------------------------------------------------------
ȫ�ֱ�����ͬģ��ĵ�����ʹ��
#moduleA
class Test(object):
    def __init__(self):
        self.number = None
    
    def set(self, n):
        self.number = n
    
    def show(self):
        print self.number
        
t = Test()      

#moduleB
from moduleA import t

def functionB():
    t.show()

#main
from moduleB import t, functionB

if __name__ == '__main__':
    t.set(10)
    t.show()
    functionB()	
#-----------------------------------------------------------------------------------------
�о�����ͨ�����ַ�ʽ��C++�д��麯��һ�����������
def show():
    pass

if __name__ == '__main__':
    show()
#-----------------------------------------------------------------------------------------
xlrd��ȡexcel
import xlrd

if __name__ == '__main__':
    wb = xlrd.open_workbook("test.xls")
    ws = wb.sheet_by_index(0)
    print wb.nsheets
    print ws.nrows
    print ws.ncols
    print ws.cell_value(0, 0)
    print ws.cell(0, 0)
	
	rowData = ws.row_values(0)
    cellValue = rowData[0]
    print cellValue
    print rowData
    print ws.col_values(2) 
	
	ws = wb.sheet_by_name("Sheet1")
    nameList = wb.sheet_names()
    print ws.cell(0, 0).value
    print nameList
#-----------------------------------------------------------------------------------------
# encoding : utf-8       #���ñ��뷽ʽ  
  
import xlrd                    #����xlrdģ��  
  
#��ָ���ļ�·����excel�ļ�  
  
xlsfile = r'D:\AutoPlan\apisnew.xls'   
book = xlrd.open_workbook(xlsfile)     #���excel��book����  
  
#��ȡsheet���󣬷�����2�֣�  
sheet_name=book.sheet_names()[0]          #���ָ��������sheet����  
print sheet_name  
sheet1=book.sheet_by_name(sheet_name)  #ͨ��sheet��������ȡ����Ȼ�����֪��sheet�����˿���ֱ��ָ��  
sheet0=book.sheet_by_index(0)     #ͨ��sheet�������sheet����  
  
#��ȡ������������  
  
nrows = sheet.nrows    #������  
ncols = sheet.ncols   #������  
  
#���ָ���С��е�ֵ�����ض���Ϊһ��ֵ�б�  
  
row_data = sheet.row_values(0)   #��õ�1�е������б�  
col_data = sheet.col_values(0)  #��õ�һ�е������б�Ȼ��Ϳ��Ե��������������  
  
#ͨ��cell��λ��������ָ��cell��ֵ  
cell_value1 = sheet.cell_value(0,1)  ##ֻ��cell��ֵ���ݣ��磺http://xxx.xxx.xxx.xxx:8850/2/photos/square/  
print cell_value1  
cell_value2 = sheet.cell(0,1) ##����cellֵ�����⻹�и������ԣ��磺text:u'http://xxx.xxx.xxx.xxx:8850/2/photos/square/'  
print cell_value2  
#-----------------------------------------------------------------------------------------
from xlrd import open_workbook

if __name__ == '__main__':
    wb = open_workbook('test.xls')
    for s in wb.sheets():
        print 'Sheet:', s.name
        for row in range(s.nrows):
            values = []
            for col in range(s.ncols):
                values.append(s.cell(row,col).value)
            print ','.join(values)#��,���������ַ���
        print
	
	
from xlrd import open_workbook

if __name__ == '__main__':
    book = open_workbook('test.xls')
    print book.nsheets
    for sheet_index in range(book.nsheets):
        print book.sheet_by_index(sheet_index)
    print book.sheet_names()
    for sheet_name in book.sheet_names():
        print book.sheet_by_name(sheet_name)
    for sheet in book.sheets():
        print sheet	

		
from xlrd import open_workbook

if __name__ == '__main__':
    book = open_workbook('test.xls')
    sheet = book.sheet_by_index(0)
    print sheet.name
    print sheet.nrows
    print sheet.ncols
    for row_index in range(sheet.nrows):
        for col_index in range(sheet.ncols):
            #print cell.name(row_index,col_index),'-',
            print sheet.cell(row_index,col_index).value	

			
from xlrd import open_workbook

if __name__ == '__main__':
    book = open_workbook('test.xls')
    sheet = book.sheet_by_index(0)
    cell = sheet.cell(0,0)
    print cell
    print cell.value
    print cell.ctype==XL_CELL_TEXT
    for i in range(sheet.ncols):
        print sheet.cell_type(1,i),sheet.cell_value(1,i)

from xlrd import open_workbook

if __name__ == '__main__':    
    book = open_workbook('test.xls')
    sheet0 = book.sheet_by_index(0)
    sheet1 = book.sheet_by_index(1)
    print sheet0.row(0)
    print sheet0.col(0)
    print
    print sheet0.row_slice(0,1)
    print sheet0.row_slice(0,1,2)
    print sheet0.row_values(0,1)
    print sheet0.row_values(0,1,2)
    print sheet0.row_types(0,1)
    print sheet0.row_types(0,1,2)
    print
    print sheet1.col_slice(0,1)
    print sheet0.col_slice(0,1,2)
    print sheet1.col_values(0,1)
    print sheet0.col_values(0,1,2)
    print sheet1.col_types(0,1)
    print sheet0.col_types(0,1,2)

from xlrd import cellname, cellnameabs, colname

if __name__ == '__main__':    
    print cellname(0,0),cellname(10,10),cellname(100,100)
    print cellnameabs(3,1),cellnameabs(41,59),cellnameabs(265,358)
    print colname(0),colname(10),colname(100)

It is recommended that flush_row_data is called for every 1000 or so rows of a normal size that are written to an xlwt.Workbook. If the rows are huge, that number should be reduced.	
#-----------------------------------------------------------------------------------------
from xlwt import Workbook

if __name__ == '__main__':    
    book = Workbook()
    sheet1 = book.add_sheet('Sheet 1')
	sheet1.name = "abc"
    book.add_sheet('Sheet 2')
    sheet1.write(0,0,'A1')
    sheet1.write(0,1,'B1')
    row1 = sheet1.row(1)
    row1.write(0,'A2')
    row1.write(1,'B2')
    sheet1.col(0).width = 10000
    sheet2 = book.get_sheet(1)
    sheet2.row(0).write(0,'Sheet 2 A1')
    sheet2.row(0).write(1,'Sheet 2 B1')
    sheet2.flush_row_data()
    sheet2.write(1,0,'Sheet 2 A3')
    sheet2.col(0).width = 5000
    sheet2.col(0).hidden = True
    book.save('simple.xls')
	

from xlwt import Workbook
book = Workbook()
sheet1 = book.add_sheet('Sheet 1',cell_overwrite_ok=True)
sheet1.write(0,0,'original')
sheet = book.get_sheet(0)
sheet.write(0,0,'new')
sheet2 = book.add_sheet('Sheet 2')
sheet2.write(0,0,'original')
sheet2.write(0,0,'new')

from datetime import date,time,datetime
from decimal import Decimal
from xlwt import Workbook,Style
wb = Workbook()
ws = wb.add_sheet('Type examples')
ws.row(0).write(0,u'\xa3')
ws.row(0).write(1,'Text')
ws.row(1).write(0,3.1415)
ws.row(1).write(1,15)
ws.row(1).write(2,265L)
ws.row(1).write(3,Decimal('3.65'))
ws.row(2).set_cell_number(0,3.1415)
ws.row(2).set_cell_number(1,15)
ws.row(2).set_cell_number(2,265L)
ws.row(2).set_cell_number(3,Decimal('3.65'))
ws.row(3).write(0,date(2009,3,18))
ws.row(3).write(1,datetime(2009,3,18,17,0,1))
ws.row(3).write(2,time(17,1))
ws.row(4).set_cell_date(0,date(2009,3,18))
ws.row(4).set_cell_date(1,datetime(2009,3,18,17,0,1))
ws.row(4).set_cell_date(2,time(17,1))
ws.row(5).write(0,False)
ws.row(5).write(1,True)
ws.row(6).set_cell_boolean(0,False)
ws.row(6).set_cell_boolean(1,True)
ws.row(7).set_cell_error(0,0x17)
ws.row(7).set_cell_error(1,'#NULL!')
ws.row(8).write(
    0,'',Style.easyxf('pattern: pattern solid, fore_colour
green;'))
ws.row(8).write(
    1,None,Style.easyxf('pattern: pattern solid, fore_colour
blue;'))
ws.row(9).set_cell_blank(
    0,Style.easyxf('pattern: pattern solid, fore_colour
yellow;'))
ws.row(10).set_cell_mulblanks(
    5,10,Style.easyxf('pattern: pattern solid, fore_colour
red;')
    )
wb.save('types.xls')

from datetime import date
from xlwt import Workbook, XFStyle, Borders, Pattern, Font

fnt = Font()
fnt.name = 'Arial'
borders = Borders()
borders.left = Borders.THICK
borders.right = Borders.THICK
borders.top = Borders.THICK
borders.bottom = Borders.THICK
pattern = Pattern()
pattern.pattern = Pattern.SOLID_PATTERN
pattern.pattern_fore_colour = 0x0A
style = XFStyle()
style.num_format_str='YYYY-MM-DD'
style.font = fnt
style.borders = borders
style.pattern = pattern
book = Workbook()
sheet = book.add_sheet('A Date')
sheet.write(1,1,date(2009,3,18),style)
book.save('simple.xls')

from datetime import date
from xlwt import Workbook, easyxf
book = Workbook()
sheet = book.add_sheet('A Date')
sheet.write(1,1,date(2009,3,18),easyxf(
    'font: name Arial;'
    'borders: left thick, right thick, top thick, bottom thick;'
    'pattern: pattern solid, fore_colour red;',
    num_format_str='YYYY-MM-DD'
    ))
sheet.write(0, 0, "hello")
book.save('simple.xls')


The human readable text breaks roughly as follows, in pseudo-regular expression syntax:
 (<element>:(<attribute> <value>,)+;)

from xlwt import Workbook, easyxf
from xlwt.Utils import rowcol_to_cell
row = easyxf('pattern: pattern solid, fore_colour blue')
col = easyxf('pattern: pattern solid, fore_colour green')
cell = easyxf('pattern: pattern solid, fore_colour red')
book = Workbook()
sheet = book.add_sheet('Precedence')
for i in range(0,10,2):
    sheet.row(i).set_style(row)
for i in range(0,10,2):
    sheet.col(i).set_style(col)
for i in range(10):
    sheet.write(i,i,None,cell)
sheet = book.add_sheet('Hiding')
for rowx in range(10):
    for colx in range(10):
        sheet.write(rowx,colx,rowcol_to_cell(rowx,colx))
for i in range(0,10,2):
    sheet.row(i).hidden = True
    sheet.col(i).hidden = True
sheet = book.add_sheet('Row height and Column width')
for i in range(10):
    sheet.write(0,i,0)
for i in range(10):
    sheet.row(i).set_style(easyxf('font:height '+str(200*i)))
    sheet.col(i).width = 256*i
book.save('simple.xls')

from xlwt import Utils
print 'AA ->',Utils.col_by_name('AA')
print 'A ->',Utils.col_by_name('A')
print 'A1 ->',Utils.cell_to_rowcol('A1')
print '$A$1 ->',Utils.cell_to_rowcol('$A$1')
print 'A1 ->',Utils.cell_to_rowcol2('A1')
print (0,0),'->',Utils.rowcol_to_cell(0,0)
print (0,0,False,True),'->',
print Utils.rowcol_to_cell(0,0,False,True)
print (0,0,True,True),'->',
print Utils.rowcol_to_cell(
          row=0,col=0,row_abs=True,col_abs=True
          )
print '1:3 ->',Utils.cellrange_to_rowcol_pair('1:3')
print 'B:G ->',Utils.cellrange_to_rowcol_pair('B:G')
print 'A2:B7 ->',Utils.cellrange_to_rowcol_pair('A2:B7')
print 'A1 ->',Utils.cellrange_to_rowcol_pair('A1')
print (0,0,100,100),'->',
print Utils.rowcol_pair_to_cellrange(0,0,100,100)
print (0,0,100,100,True,False,False,False),'->',
print Utils.rowcol_pair_to_cellrange(
          row1=0,col1=0,row2=100,col2=100,
          row1_abs=True,col1_abs=False,
          row2_abs=False,col2_abs=True
          )
for name in (
    '',"'quoted'","O'hare","X"*32,"[]:\\?/*\x00"
    ):
    print 'Is %r a valid sheet name?' % name,
    if Utils.valid_sheet_name(name):
        print "Yes"
    else:
        print "No"

from xlwt import Workbook
w = Workbook()
ws = w.add_sheet('Image')
ws.insert_bitmap('abc.bmp', 0, 0)
w.save("simple.xls")

from xlwt import Workbook,easyxf
style = easyxf(
    'pattern: pattern solid, fore_colour red;'
    'align: vertical center, horizontal center;'
    )
w = Workbook()
ws = w.add_sheet('Merged')
ws.write_merge(1,5,1,5,'Merged',style)
w.save("simple.xls")

from xlwt import Workbook,easyxf
tl = easyxf('border: left thick, top thick')
t = easyxf('border: top thick')
tr = easyxf('border: right thick, top thick')
r = easyxf('border: right thick')
br = easyxf('border: right thick, bottom thick')
b = easyxf('border: bottom thick')
bl = easyxf('border: left thick, bottom thick')
l = easyxf('border: left thick')
w = Workbook()
ws = w.add_sheet('Border')
ws.write(1,1,style=tl)
ws.write(1,2,style=t)
ws.write(1,3,style=tr)
ws.write(2,3,style=r)
ws.write(3,3,style=br)
ws.write(3,2,style=b)
ws.write(3,1,style=bl)
ws.write(2,1,style=l)
w.save("simple.xls")

from xlwt import Workbook
data = [
    ['','','2008','','2009'],
    ['','','Jan','Feb','Jan','Feb'],
    ['Company X'],
    ['','Division A'],
    ['','',100,200,300,400],
    ['','Division B'],
    ['','',100,99,98,50],
    ['Company Y'],
    ['','Division A'],
    ['','',100,100,100,100],
    ['','Division B'],
    ['','',100,101,102,103],
    ]
w = Workbook()
ws = w.add_sheet('Outlines')
for i,row in enumerate(data):
    for j,cell in enumerate(row):
        ws.write(i,j,cell)
ws.row(2).level = 1
ws.row(3).level = 2
ws.row(4).level = 3
ws.row(5).level = 2
ws.row(6).level = 3
ws.row(7).level = 1
ws.row(8).level = 2
ws.row(9).level = 3
ws.row(10).level = 2
ws.row(11).level = 3
ws.col(2).level = 1
ws.col(3).level = 2
ws.col(4).level = 1
ws.col(5).level = 2
w.save("simple.xls")

from xlwt import Workbook
w = Workbook()
ws = w.add_sheet('Normal')
ws.write(0,0,'Some text')
ws.normal_magn = 75
ws = w.add_sheet('Page Break Preview')
ws.write(0,0,'Some text')
ws.preview_magn = 150
ws.page_preview = True
w.save("simple.xls")


from xlrd import open_workbook
from xlwt import easyxf
from xlutils.copy import copy
rb = open_workbook('source.xls',formatting_info=True)
rs = rb.sheet_by_index(0)
wb = copy(rb)
ws = wb.get_sheet(0)
plain = easyxf('')
for i,cell in enumerate(rs.col(2)):
    if not i:
        continue
    ws.write(i,2,cell.value,plain)
for i,cell in enumerate(rs.col(4)):
    if not i:
        continue
    ws.write(i,4,cell.value-1000)
wb.save('output.xls')		

from xlrd import open_workbook
from xlwt import Workbook
from xlutils.copy import copy
rb = open_workbook("simple.xls")
wb = copy(rb)
ws = wb.add_sheet('hello')
wb.save("simple.xls")

# encoding=utf-8
from xlwt import Workbook, easyxf

def fillInExcel(ws, cellFormat,rowNumber, val0, val1, val2, val3, val4):
    ws.write(rowNumber, 0, val0, formatForValue)
    ws.write(rowNumber, 1, val1, formatForValue)
    ws.write(rowNumber, 2, val2, formatForValue)
    ws.write(rowNumber, 3, val3, formatForValue)
    ws.write(rowNumber, 4, val4, formatForValue)

if __name__ == '__main__':
    wb = Workbook(encoding = "utf-8")#��������������壬�����ﲻ��Ҫ��utf-8���룬Ĭ��wb = Workbook()����
    ws = wb.add_sheet("Test")
    formatForTitle = easyxf('font: name Times New Roman; font: bold True; font: height 300;'
                            'borders: left thin, right thin, top thin, bottom thin;'
                            'alignment: wrap True; alignment: horizontal left; alignment: vertical center;'
                            )
    row0 = ws.row(0)
    row0.write(0, "Moc", formatForTitle)
    row0.write(1, "Attribute", formatForTitle)
    row0.write(2, "Parameter Comparison", formatForTitle)
    row0.write(3, "Template Application", formatForTitle)
    row0.write(4, "Summary Customization(Only for New BaseStation)", formatForTitle)
    
    formatForValue = easyxf('font: name %s;'
                            'borders: left thin, right thin, top thin, bottom thin;'
                            'alignment: horizontal left; alignment: vertical center;' % (u"����")
                            )
  
    for i in range(1, 1000):
        fillInExcel(ws, formatForValue, i, "It", "is", "a", "nice", "day")#��д����������504ʱ��һ��Ҫ��format��Ϊ��������д�뺯��fillInExcel�У�������ú�����ֱ����forѭ����дҲ���ԣ��������format������fillInExcel�У���ᱨ���ʽ������
    
    ws.col(0).width = 5000#width:18.86
    ws.col(1).width = 6000
    ws.col(2).width = 7000
    ws.col(3).width = 8000
    ws.col(4).width = 8000
    wb.save("test.xls")
#-----------------------------------------------------------------------------------------
#��д�ļ�����b�������ƶ�ȡ�����Ա����ļ���ʽ
def getEnterSymbol(line):#����ֻ��һ�����ݣ�û�л��з����ļ��������������
    if len(line) == 1:#������г���Ϊ1����ôҪô��һ�����з���Ҫô��һ��ֻ��һ���ַ����ļ���
        return line
    else:
        lastTwoSymbol = line[-2:]
        if lastTwoSymbol == "\r\n":
            return lastTwoSymbol
        else:
            return  lastTwoSymbol[1]       

if __name__ == '__main__':
    testFile = open(r"c:\Users\Ben\Desktop\test.txt", "rb")
    newTestFile = open(r"c:\Users\Ben\Desktop\newtest.txt", "wb")
    line = testFile.readline()
    enterSymbol = getEnterSymbol(line)
    lineNumber = 1
    while line:        
        newTestFile.write("line%s%s:" % (str(lineNumber), enterSymbol))
        newTestFile.write(line)
        line = testFile.readline()
        lineNumber += 1
    testFile.close()
    newTestFile.close()
��
if __name__ == '__main__':
    testFile = open(r"c:\Users\Ben\Desktop\test.txt", "rb")
    newTestFile = open(r"c:\Users\Ben\Desktop\newtest.txt", "wb")
    lines = testFile.readlines()
    enterSymbol = getEnterSymbol(lines[0])
    lineNumber = 1
    for line in lines:
        newTestFile.write("line%s%s:" % (str(lineNumber), enterSymbol))
        newTestFile.write(line)
        #line = testFile.readline()
        lineNumber += 1
    testFile.close()
    newTestFile.close()
#-----------------------------------------------------------------------------------------
�ڴ�����־�ļ���ʱ�򣬳����������������������־�ļ��޴󣬲�����һ���԰������ļ����뵽�ڴ��н��д���������Ҫ��һ̨�����ڴ�Ϊ 2GB �Ļ����ϴ���һ�� 2GB ����־�ļ������ǿ���ϣ��ÿ��ֻ�������� 200MB �����ݡ�
�� Python �У����õ� File ����ֱ���ṩ��һ�� readlines(sizehint) ������������������顣������Ĵ���Ϊ����

file = open('test.log', 'r')
sizehint = 209715200   # 200M
position = 0
lines = file.readlines(sizehint)
while not file.tell() - position < 0:
	position = file.tell()
	lines = file.readlines(sizehint)

ÿ�ε��� readlines(sizehint) �������᷵�ش�Լ 200MB �����ݣ����������صı�Ȼ���������������ݣ����������£����ص����ݵ��ֽ�������΢�� sizehint ָ����ֵ��һ�㣨�����һ�ε��� readlines(sizehint) ������ʱ�򣩡�ͨ������£�Python ���Զ����û�ָ���� sizehint ��ֵ�������ڲ������С����������

file��python��һ����������ͣ���������python�����ж��ⲿ���ļ����в�������python��һ�ж��Ƕ���fileҲ�����⣬file��file�ķ��������ԡ�������������δ���һ��file����
file(name[, mode[, buffering]]) 
file()�������ڴ���һ��file��������һ��������open()�����ܸ�����һЩ�����������ú��������������Ĳ������������������ַ�������ʽ���ݵġ�name���ļ������֡�
mode�Ǵ򿪵�ģʽ����ѡ��ֵΪr w a U���ֱ�������Ĭ�ϣ� д ���֧�ָ��ֻ��з���ģʽ����w��aģʽ���ļ��Ļ�������ļ������ڣ���ô���Զ����������⣬��wģʽ��һ���Ѿ����ڵ��ļ�ʱ��ԭ���ļ������ݻᱻ��գ���Ϊһ��ʼ�ļ��Ĳ����ı�������ļ��Ŀ�ͷ�ģ���ʱ�����д���������ɻ��ԭ�е����ݸ�Ĩ����������ʷ��ԭ�򣬻��з��ڲ�ͬ��ϵͳ���в�ͬģʽ�������� unix����һ��\n������windows���ǡ�\r\n������Uģʽ���ļ�������֧�����еĻ���ģʽ��Ҳ��˵��\r�� '\n' '\r\n'���ɱ�ʾ���У�����һ��tuple������������ļ����õ����Ļ��з�����������˵�����ж���ģʽ������python��ͳһ��\n���档��ģʽ�ַ��ĺ��棬�����Լ���+ b t�����ֱ�ʶ���ֱ��ʾ���Զ��ļ�ͬʱ���ж�д�������ö�����ģʽ���ı�ģʽ��Ĭ�ϣ����ļ���
buffering���Ϊ0��ʾ�����л���;���Ϊ1��ʾ���С��л��塰;�����һ������1������ʾ�������Ĵ�С��Ӧ�������ֽ�Ϊ��λ�ġ�

file�������Լ������Ժͷ�������������file�����ԡ�
closed #����ļ��Ƿ��Ѿ��رգ���close()��д 
encoding #�ļ����� 
mode #��ģʽ 
name #�ļ��� 
newlines #�ļ����õ��Ļ���ģʽ����һ��tuple 
softspace #boolean�ͣ�һ��Ϊ0����˵����print 

F.read([size]) #sizeΪ��ȡ�ĳ��ȣ���byteΪ��λ 
F.readline([size]) 
#��һ�У����������size���п��ܷ��ص�ֻ��һ�е�һ���� 
F.readlines([size]) 
#���ļ�ÿһ����Ϊһ��list��һ����Ա�����������list����ʵ�����ڲ���ͨ��ѭ������readline()��ʵ�ֵġ�����ṩsize������size�Ǳ�ʾ��ȡ���ݵ��ܳ���Ҳ����˵����ֻ�����ļ���һ���֡� 
F.write(str) 
#��strд���ļ��У�write()��������str�����һ�����з� 
F.writelines(seq) 
#��seq������ȫ��д���ļ��С��������Ҳֻ����ʵ��д�룬������ÿ�к�������κζ����� 
file������������
F.close() 
#�ر��ļ���python����һ���ļ����ú��Զ��ر��ļ���������һ����û�б�֤����û��������Լ��رյ�ϰ�ߡ����һ���ļ��ڹرպ󻹶�����в��������ValueError 
F.flush() 
#�ѻ�����������д��Ӳ�� 
F.fileno() 
#����һ�������͵ġ��ļ���ǩ�� 
F.isatty() 
#�ļ��Ƿ���һ���ն��豸�ļ���unixϵͳ�еģ� 
F.tell() 
#�����ļ�������ǵĵ�ǰλ�ã����ļ��Ŀ�ͷΪԭ�� 
F.next() 
#������һ�У������ļ��������λ�Ƶ���һ�С���һ��file����for ... in file���������ʱ�����ǵ���next()������ʵ�ֱ����ġ� 
F.seek(offset[,whence]) 
#���ļ����������Ƶ�offset��λ�á����offsetһ����������ļ��Ŀ�ͷ������ģ�һ��Ϊ������������ṩ��whence�����Ͳ�һ���ˣ�whence����Ϊ0��ʾ��ͷ��ʼ���㣬1��ʾ�Ե�ǰλ��Ϊԭ����㡣2��ʾ���ļ�ĩβΪԭ����м��㡣��Ҫע�⣬����ļ���a��a+��ģʽ�򿪣�ÿ�ν���д����ʱ���ļ�������ǻ��Զ����ص��ļ�ĩβ�� 
F.truncate([size]) 
#���ļ��óɹ涨�Ĵ�С��Ĭ�ϵ��ǲõ���ǰ�ļ�������ǵ�λ�á����size���ļ��Ĵ�С��Ҫ������ϵͳ�Ĳ�ͬ�����ǲ��ı��ļ���Ҳ��������0���ļ�������Ӧ�Ĵ�С��Ҳ��������һЩ��������ݼ���ȥ�� 
#-----------------------------------------------------------------------------------------
#�����쳣
if __name__ == '__main__':
    try:
        a = "abc" + 1
    except Exception, e:
        print e
�����cannot concatenate 'str' and 'int' objects
#-----------------------------------------------------------------------------------------
#����
# encoding=utf-8
'''
Created on 2012-12-27

@author: Ben
'''    

if __name__ == '__main__':
    print "���"
	

# encoding=utf-8
'''
Created on 2012-12-27

@author: Ben
'''    
#��ʱ������ļ���ʽΪutf-8����������������ʾ
if __name__ == '__main__':
    f = open("test.txt", "r")
    l = f.readline()
    print l
    f.close()

#��ʱ������ļ���ʽΪansi������Ҫ��gbk����
# encoding=utf-8
if __name__ == '__main__':
    f = open("test.txt", "r")
    l = f.readline().decode("gbk")
    print l
    f.close()
����
# encoding=utf-8
import codecs
if __name__ == '__main__':
    f = open("test.txt")
    (encoder, decoder, reader, writer) = codecs.lookup("gbk")
    f = reader(f)
    l = f.readline()
    print l
    f.close()
����
# encoding=utf-8
import codecs

if __name__ == '__main__':
    f = codecs.open("test.txt", "r", "gbk")
    l = f.readline()
    print l
    f.close()

#ͨ�����´��룬���Դ���������utf-8���뻹��ansi������ļ�
# encoding=utf-8
import codecs

if __name__ == '__main__':
    try:
        f = codecs.open("test.txt", "r", "utf-8")#f = codecs.open("test.txt", "r", encoding="utf-8")
        l = f.readline()
        print l
        f.close()
    except Exception, e:
        print e
        f.close()
        f = codecs.open("test.txt", "r", "gbk")
        l = f.readline()
        print l
        f.close()
		
#дutf-8����ʱ����������ģ���utf-8��bom��ʽ�����ֻ��ascii�ַ��������ļ�����ansi��ʽ
# encoding=utf-8
import codecs

if __name__ == '__main__':
    f = codecs.open("test.txt", "w", "utf-8")
    txt = unicode("���\n", "utf-8")
    f.write(txt)
    f.close()
	
#���ǿ��Ҫ��������û�����ģ�����utf-8��ʽ���������±�д
# encoding=utf-8
import codecs

if __name__ == '__main__':
    f = codecs.open("test.txt", "w", "utf-8-sig")
    txt = unicode("qwer\n")
    f.write(txt)
    f.close()

#���û̫������ʲô��
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import codecs

txt = u"qwer"
file=codecs.open("test","w","utf-8")
file.write(txt)
file.close()

#ʲô��ʽ���������Զ�Ϊansi��ʽ�����ϣ�����ֵ�����ʽ������Ҫ��дʱ����wb���������ֻ��з���windows:\r\n��unix:\n��mac:\r�������е����⣬��ôд����windows��
# encoding=utf-8
import codecs

if __name__ == '__main__':
    f = codecs.open("test.txt", "wb")
    txt = unicode("hello\n")
    f.write(txt)
    f.close()
	
# -*- coding: UTF-8 -*- ���Ǹ�ע����
��������˵�����PythonԴ�����ļ���ʹ�õı��롣ȱʡ�������ĳ�����Ҫʹ��ascii����д�������������д���ĵĻ���python������һ��ᱨ����������������õ��ļ����룬python�ͻ��Զ������ٱ���
������ʽ������д�ɣ�
#coding=utf-8
��
#coding:utf-8	
#-----------------------------------------------------------------------------------------
#����־û�
# encoding=utf-8
'''
Created on 2012-12-27

@author: Ben
'''    
import pickle

if __name__ == '__main__':
    l = ["a", "b", "c"]
    f = open("test.txt", "w")
    pickle.dump(l, f)
    f.close()
#    nameMap = {"Mayian":["a", "b", "c"], "Dingben":["d", "e"]}
#    f = open("test.txt", "w")
#    pickle.dump(nameMap, f)
#    f.close()
	
# encoding=utf-8
import pickle

if __name__ == '__main__':
    f = open("test.txt", "r")
    l = pickle.load(f)
    print l
    f.close()
#    f = open("test.txt", "r")
#    nameMap = pickle.load(f)
#    print nameMap
#    f.close()

#shelve
# encoding=utf-8
import shelve

if __name__ == '__main__':
    dbase = shelve.open("test.sh")
    dbase["a"] = "A"
    dbase["b"] = "B"
    dbase.close()
	
# encoding=utf-8
import shelve

if __name__ == '__main__':
    dbase = shelve.open("test.sh")
    for key, value in dbase.items():
        print key, value
    dbase.close()
#-----------------------------------------------------------------------------------------
#win32com����excel
import win32com.client

if __name__ == '__main__':
    xlApp = win32com.client.DispatchEx("Excel.Application")
    xlBook = xlApp.Workbooks.Add()
    xlBook.SaveAs(r"d:\Program Files\eclipse workspace\PyLearngingProgram\main\Hello.xlsx")   
    xlBook.Close(SaveChanges=0)
    del xlApp

	
import win32com.client

if __name__ == '__main__':
    xlApp = win32com.client.DispatchEx("Excel.Application")
    excelPath = r"d:\Program Files\eclipse workspace\PyLearngingProgram\main\Hello.xlsx"
    xlBook = xlApp.Workbooks.Open(excelPath)
    sht = xlBook.Worksheets("Sheet1")
    sht.Cells(1, 1).value = "Hello"
    xlBook.Save()   
    xlBook.Close(SaveChanges=0)
    del xlApp
#-----------------------------------------------------------------------------------------
#��ˮ����
# encoding=utf-8
import sys
import copy 
#transfer matrix
gmat = 0 
maxint = 99999999 
def build_state_matrix(size):
    global gmat
    global maxint    
    gmat = size*[maxint]
    tmp = []
    for i in range(size):
        tmp.append(copy.deepcopy(gmat))        
    gmat = tmp 

def draw_lines(mystate,a,b):
    global gmat    
    aa = mystate[0]
    bb = mystate[1]    
    state_ord = aa*(b+1) + bb    
    if (aa!=0) :    
    #clean a , so (aa,bb)--> (0,bb)       
        next_state_ord =  bb        
        gmat[state_ord][next_state_ord] = 1    
    if (bb!=0) :    
    #clean b , so (aa,bb)--> (aa,0)        
        next_state_ord =  aa*(b+1)        
        gmat[state_ord][next_state_ord] = 1    
    if (aa!=a) :    
    #full a , so (aa,bb)--> (a,bb)        
        next_state_ord =  a*(b+1) + bb        
        gmat[state_ord][next_state_ord] = 1    
    if (bb!=b) :    
    #full b , so (aa,bb)--> (aa,b)        
        next_state_ord =  aa*(b+1) + b        
        gmat[state_ord][next_state_ord] = 1    
    if (aa!=0) :    
    #a to b , 1 final,a get empty    
    #(aa,bb)--->(0,bb+aa)        
        if((bb+aa)<=b) :            
            next_state_ord = bb+aa ;            
            gmat[state_ord][next_state_ord] = 1        
        #a to b , 2 final,b get full    
        #(aa,bb)--->(aa+bb-b,b,)        
        if((aa+bb-b)>=0) :            
            next_state_ord = (aa+bb-b)*(b+1) + b            
            gmat[state_ord][next_state_ord] = 1    
    if (bb!=0) :    
    #b to a , 1 final,b get empty    
    #(aa,bb)--->(aa+bb,0)        
        if((bb+aa)<=a) :            
            next_state_ord = (aa+bb)*(b+1) ;            
            gmat[state_ord][next_state_ord] = 1       
        #a to b , 2 final,a get full    
        #(aa,bb)--->(a,aa+bb-a)        
        if((aa+bb-a)>=0) :            
            next_state_ord = a*(b+1) + (aa+bb-a)            
            gmat[state_ord][next_state_ord] = 1                 

def dijkstra(mat,s,e,n) :    
    path = []    
    pair = {}.fromkeys(range(n))    
    global maxint    
    d = n*[maxint]    
    flag = n*[0]    
    for i in range(n) :        
        if(mat[s][i]==1) :            
            d[i] = 1             
            pair[i] = s        
        else : 
            d[i] = maxint     
    flag[s] = 1    
    while True :        
        mint = maxint        
        key = -1        
        for i in range(n) :            
            if(flag[i]==1) : continue            
            if ( d[i] < mint ) :                
                mint = d[i] ;                
                key = i        
        if(key==-1) : 
            return -1        
        flag[key]=1        
        if(key==e) : break        
        for i in range(n) :            
            if (i==key) : 
                continue             
            if ( mat[key][i] !=1 ) : 
                continue             
            if ( d[i] > (d[key]+mat[key][i]) ) :                
                d[i] = d[key] + mat[key][i]                 
                pair[i] = key    
    path.insert(0,e)    
    parent = pair[e]    
    while (parent!=0) :        
        path.insert(0,parent)        
        parent = pair[parent]    
    path.insert(0,0)    
    return path    

def pour(a,b,c) :    
    #1st step,build state     
    size = (a+1)*(b+1)    
    build_state_matrix(size)    
    #(0,0) is the 0th state    
    #(m,n) is the m*(b+1) + n     
    #2nd step,draw the arrows    
    global gmat    
    for aa in range(a+1) :        
        for bb in range(b+1) :            
            #(aa,bb)            
            #six operation            
            # clean A            
            # clean B            
            # fill A            
            # fill B            
            # A to B            
            # B to A            
            draw_lines((aa,bb),a,b)                
    #dijkstra to find the best way    
    s = 0 #(0,0)    
    e = c  #(0,c)    
    if c <= b :        
        path1 = dijkstra(gmat,s,e,size)    
    else : 
        path1 = -1    
    e = c*(b+1) #(c,0)    
    if c <= a :        
        path2 = dijkstra(gmat,s,e,size)    
    else : 
        path2 = -1        
    
    if( path1 == -1 and path2 == -1 ) :        
        print "No solution!"        
        return []    
    if ( path1 == -1 ) :        
        path = path2    
    elif ( path2 == -1 ) :       
        path = path1    
    else :        
        if(len(path1)<len(path2)) :            
            path = path1        
        else : path = path2    
    print len(path)    
    for i in path :        
        bi = i%(b+1)        
        ai = (i-bi)/(b+1)        
        print "(","%d" % ai,"%d" % bi ,")" ,     
        print      

if __name__ == "__main__" :    
#    a = int(sys.argv[1])    
#    b = int(sys.argv[2])    
#    c = int(sys.argv[3])    
    a = 5
    b = 3
    c = 2
    pour(a,b,c)
#-----------------------------------------------------------------------------------------
#python�ĺ������ݣ�����C++�ĺ���ָ��
def printHello(msg):
    print msg
    
def functionHandler(functionName, msg):
    functionName(msg)
    
if __name__ == '__main__': 
    functionHandler(printHello, "hello, world")
#-----------------------------------------------------------------------------------------
py2exeʹ�����
from distutils.core import setup
import py2exe

setup(windows=[{"script": "D:/Program Files/eclipse workspace/PyLearngingProgram/main/main.py", "icon_resources": [(1, "D:/Installations/python-2/py2exe/1.ico")]}])
#setup(console=[{"script": "D:/Program Files/eclipse workspace/PyLearngingProgram/main/main.py", "icon_resources": [(1, "D:/Installations/python-2/py2exe/1.ico")]}])
#-----------------------------------------------------------------------------------------
���self�������ΪC++�е�this
def thisTest(test):
    print test.l

class Test(object):
    def __init__(self):
        self.l = ["abc", "def"]
    
    def show(self):
        thisTest(self)
        
if __name__ == '__main__': 
    test = Test()
    test.show()
#-----------------------------------------------------------------------------------------
*var�ڶ��庯������ʱ���ǿɱ������������˼���ڵ���ʱ����Ҫunpack tuple����˼���磺
t = (1,2,3)ֱ�Ӵ�����һ������������*t�����3��������

#@PydevCodeAnalysisIgnore
#encoding=utf-8
#!/usr/bin/env python

def testStar(a, b, c, d):
    print a, b, c, d
    
def testStar2(*arg):
    print arg
    
if __name__ == '__main__':     
    a = (1, 2, 3, 4)
    testStar(*a)
    testStar2(a)
    testStar2(*a)
    
    b = [11, 12, 13, 14]
    testStar(*b)
    testStar2(b)
    testStar2(*b)

������£�
1 2 3 4
((1, 2, 3, 4),)
(1, 2, 3, 4)
11 12 13 14
([11, 12, 13, 14],)
(11, 12, 13, 14)
#-----------------------------------------------------------------------------------------
#��Ԫ�����������ʵ��
���濴һ��python�Ľ������һ��
def trans(v):  
	return 1 if v==0 else v  
Ҳ����if else���ļ�д��ʽ����˼һ�������ף�������������   
   
�����ǽ����������
def trans(v):  
    return v==0 and 1 or v  
�õ����������������ԡ�
����һ�£����v����0Ϊtrue�����1�������㣬Ϊtrue���򲻽��к���Ļ����㣬ֱ�ӷ���1�����v����0Ϊfalse�����1�������㣬Ϊfalse���������л����㣬����v��
#-----------------------------------------------------------------------------------------
#����xml
'''
Created on 2013-9-29

@author: Ben
'''

import xml.dom.minidom as Dom  
  
class XmlGenerator:  
    def __init__(self, xml_name):  
        self.doc = Dom.Document()  
        self.xml_name = xml_name  
          
    def createNode(self, node_name):  
        return self.doc.createElement(node_name)  
      
    def addNode(self, node, prev_node = None):  
        cur_node = node  
        if prev_node is not None:  
            prev_node.appendChild(cur_node)  
        else:  
            self.doc.appendChild(cur_node)  
        return cur_node  
  
    def setNodeAttr(self, node, att_name, value):  
        cur_node = node  
        cur_node.setAttribute(att_name, value)  
  
    def setNodeValue(self, cur_node, value):  
        node_data = self.doc.createTextNode(value)  
        cur_node.appendChild(node_data)  
  
    def genXml(self):  
        f = open(self.xml_name, "w")  
        f.write(self.doc.toprettyxml(indent = "\t", newl = "\n", encoding = "utf-8"))  
        f.close()  
          
if __name__ == "__main__":  
    myXMLGenerator = XmlGenerator("book_store.xml")  
      
    #xml root node  
    node_book_store = myXMLGenerator.createNode("book_store")  
    myXMLGenerator.setNodeAttr(node_book_store, "name", "new hua")  
    myXMLGenerator.setNodeAttr(node_book_store, "website", "http://www.ourunix.org")  
    myXMLGenerator.addNode(node = node_book_store)  
      
    #book01  
    node_book_01 = myXMLGenerator.createNode("book")
    myXMLGenerator.setNodeAttr(node_book_01, "index", "1")
      
    node_book_01_name = myXMLGenerator.createNode("name")  
    myXMLGenerator.setNodeValue(node_book_01_name, "Hamlet")  
    myXMLGenerator.addNode(node_book_01_name, node_book_01)  
      
    node_book_01_author = myXMLGenerator.createNode("author")  
    myXMLGenerator.setNodeValue(node_book_01_author, "William Shakespeare")  
    myXMLGenerator.addNode(node_book_01_author, node_book_01)  
  
    node_book_01_price = myXMLGenerator.createNode("price")  
    myXMLGenerator.setNodeValue(node_book_01_price, "$20")  
    myXMLGenerator.addNode(node_book_01_price, node_book_01)  
  
    node_book_01_grade = myXMLGenerator.createNode("grade")  
    myXMLGenerator.setNodeValue(node_book_01_grade, "good")  
    myXMLGenerator.addNode(node_book_01_grade, node_book_01)  
      
    myXMLGenerator.addNode(node_book_01, node_book_store)  
      
    #book 02  
    node_book_02 = myXMLGenerator.createNode("book")  
    myXMLGenerator.setNodeAttr(node_book_02, "index", "2")
      
    node_book_02_name = myXMLGenerator.createNode("name")  
    myXMLGenerator.setNodeValue(node_book_02_name, "shuihu")  
    myXMLGenerator.addNode(node_book_02_name, node_book_02)  
      
    node_book_02_author = myXMLGenerator.createNode("author")  
    myXMLGenerator.setNodeValue(node_book_02_author, "naian shi")  
    myXMLGenerator.addNode(node_book_02_author, node_book_02)  
  
    node_book_02_price = myXMLGenerator.createNode("price")  
    myXMLGenerator.setNodeValue(node_book_02_price, "$200")  
    myXMLGenerator.addNode(node_book_02_price, node_book_02)  
  
    node_book_02_grade = myXMLGenerator.createNode("grade")  
    myXMLGenerator.setNodeValue(node_book_02_grade, "good")  
    myXMLGenerator.addNode(node_book_02_grade, node_book_02)  
      
    myXMLGenerator.addNode(node_book_02, node_book_store)  
      
    #gen  
    myXMLGenerator.genXml()  
#-----------------------------------------------------------------------------------------
#enumerate����
enumerate Found at: __builtin__
enumerate(iterable[, start]) -> iterator for index, value of iterable
    
    Return an enumerate object.  iterable must be another object that supports
    iteration.  The enumerate object yields pairs containing a count (from
    start, which defaults to zero) and a value yielded by the iterable argument.
    enumerate is useful for obtaining an indexed list:
    (0, seq[0]), (1, seq[1]), (2, seq[2]), ...
	
import sys
rows = (("hello", "world", "good"), ("nice", "day", "how"))

if __name__ == '__main__':
    for item in rows:
        #index = f.list.InsertStringItem(sys.maxint, item[0])
        print sys.maxint, item[0]
        for col, text in enumerate(item[1:]):
            #self.list.SetStringItem(index, col + 1, text)
            print col, text
������£�
2147483647 hello
0 world
1 good
2147483647 nice
0 day
1 how

#-----------------------------------------------------------------------------------------
#ʹ��ElementTree����xml
import sys
import time
import string
 
import xml.etree.ElementTree as etree

#����Ĭ���ַ���ΪUTF8 ��Ȼ��Щʱ��ת��������
default_encoding = 'utf-8'
if sys.getdefaultencoding() != default_encoding:
    reload(sys)
    sys.setdefaultencoding(default_encoding)

def create_xml():

    data = etree.Element("data")
    #1 interface_version
    interface_version_txt = '5'
    interface_version = etree.SubElement(data, 'interface_version')
    interface_version.text = interface_version_txt
    #2 site
    site_txt = 'www.xxx.com'
    site = etree.SubElement(data, 'site')
    site.text = site_txt
    #3 lastmod
    lastmod_txt = time.strftime('%Y-%m-%d', time.localtime())
    lastmod = etree.SubElement(data, 'lastmod')
    lastmod.text = lastmod_txt
    #5 app
    app = etree.SubElement(data, 'app')
    #6 title 
    title_txt = u'%s' % '���Ļ���ð��'
    #title_txt = etree.CDATA(title_txt)
    title = etree.SubElement(app, 'title')
    title.text = title_txt
    #7 appid
    appid = etree.SubElement(app, 'appid')
    appid.text = '%s' % '222'

    #dataxml = etree.tostring(data, pretty_print=True, encoding="UTF-8", method="xml", xml_declaration=True, standalone=None)
    dataxml = etree.tostring(data, encoding="UTF-8", method="xml")
    print dataxml
    
     
if __name__ == '__main__':
    create_xml()

������£�
<?xml version='1.0' encoding='UTF-8'?>
<data><interface_version>5</interface_version><site>www.xxx.com</site><lastmod>2016-06-11</lastmod><app><title>���Ļ���ð��</title><appid>222</appid></app></data>


#coding:utf-8

'''
Created on 2016��6��10��

@author: Ben
'''

from xml.etree.ElementTree import Element, ElementTree

if __name__ == '__main__':
    root = Element('bookstore')
    tree = ElementTree(root)
    
    #����1���ӽڵ�
    child0 = Element('book', {'category' : "COOKING"} )
    root.append(child0)
    
    #����2���ӽڵ�
    child00 = Element('title', {'language' : "English"} )
    child00.text = 'Everyday Italian' #2���ӽڵ��ı�
    child0.append(child00)
    
    tree.write('test.xml', encoding="UTF-8", method="xml", xml_declaration=True)
    

http://pycoders-weekly-chinese.readthedocs.io/en/latest/issue6/processing-xml-in-python-with-element-tree.html
ElementTree �� һ�� API ������ʵ��
ElementTree ��������Ϊ�˴��� XML ������ Python ��׼����������ʵ�֡�һ���Ǵ� Python ʵ������ xml.etree.ElementTree ������һ�����ٶȿ�һ��� xml.etree.cElementTree ����Ҫ��ס�� ����ʹ�� C ����ʵ�ֵ����֣���Ϊ���ٶȸ��죬�������ĵ��ڴ���١������ĵ�����û�� _elementtree (��ע��4) ��ô����Ҫ��������

try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET
#-----------------------------------------------------------------------------------------
#lambda���ʽ
http://www.cnblogs.com/evening/archive/2012/03/29/2423554.html
# encoding=utf-8


def add(a, b):
    return a + b


if __name__ == "__main__":
    g = lambda x: x + 1
    print g(1)

    print (lambda x: x + 2)(10)

    foo = [2, 18, 9, 22, 17, 24, 8, 12, 27]
    listAfterFilter = filter(lambda x: x % 3 == 0, foo)
    print listAfterFilter

    l1 = [1, 2, 3]
    l2 = [4, 5, 6]
    print "one sequence: ", map(lambda x: x + 2, l1)  # ������ӳ��
    print "two sequence: ", map(lambda x, y: x + y, l1, l2)  # �������߶�����ӳ��
    print "three sequences: ", map(lambda x, y, z: x + y + z, l1, l2, [7, 8, 9])

    print reduce(lambda x, y: x + y, l1)  # �޳�ʼֵ
    print reduce(lambda x, y: x + y, l1, 10)  # �г�ʼֵ
    print reduce(add, l1, 20)  # ʹ�ú���
    print reduce(lambda sumLen, s: sumLen + len(s), ["abcd", "de", "f"], 0) #����������󳤶Ⱥͣ���Ҫ����ʼֵ�����м��ۼ�ֵ��������lambda���ʽ�ĵ�һ��������������ṩ��ʼֵ0���򽫱���TypeError: cannot concatenate 'str' and 'int' objects

    print all([1, 2, 3]), all([0, 1, 2, 3]) #�Ƿ����еĵ�������ΪTrue
    print any([1, 2, 3]), any([0, 1, 2, 3]) #�Ƿ�������һ����������ΪTrue
�����
2
12
[18, 9, 24, 12, 27]
one sequence:  [3, 4, 5]
two sequence:  [5, 7, 9]
three sequences:  [12, 15, 18]
6
16
26
7
True False
True True

#-----�ַ�����split��join����
# encoding=utf-8


if __name__ == "__main__":
    s = "a b   c"
    print s.split()  # �޲�����ʹ��" "�ָ���["a", "b", "c"]
    print " ".join(s.split()) #�õ�һ��str�������������ж��� a b c
#-----�Ƿ�Ҫʹ��lambda���ʽ
# encoding=utf-8


def getProcess(collapse):
    return collapse and (lambda s: " ".join(s.split())) or (lambda s: s)


def getStr(collapse, s):
    if collapse:
        return " ".join(s.split())
    else:
        return s


if __name__ == "__main__":
    # print getProcess(True)("a   b  c")
    processFunc = lambda collapse: collapse and (lambda s: " ".join(s.split())) or (lambda s: s) #lambda���ʽ�����������һ��boolֵ������ֵ��һ��lambda���ʽ
    testStr = "a   b  c"
    print processFunc(True)(testStr) #�ȴ���True���õ�����һ��lambda���ʽ��lambda s: " ".join(s.split())������lambda���ʽ����"a   b  c"��������ѹ���ո���
    # print getStr(True, testStr) #ͬ���Ĳ�������Ҫ5�ж���+ʵ��
    print processFunc(False)(testStr) #�ȴ���False���õ���һ��lambdag���ʽ��lambda s: s���������������

    
ͨ�������ӣ����Ƿ��֣�lambda��ʹ�ô������˴��룬ʹ�����������������ֵ��ע����ǣ������һ���̶��Ͻ��ʹ���Ŀɶ��ԡ�������Ƿǳ���Ϥpython���˻����Դ˸е�������⡣
--lambda ������һ����������
--lambda �����������������Ч�ʵ���ߣ�ֻ��ʹ�������ࡣ
--�������ʹ��for...in...if����ɵģ��������lambda��
--���ʹ��lambda��lambda�ڲ�Ҫ����ѭ��������У�����Ը���庯������ɣ�ʹ�����ÿ������Ժ͸��õĿɶ��ԡ�
�ܽ᣺lambda ��Ϊ�˼��ٵ��к����Ķ�������ڵġ�
#-----------------------------------------------------------------------------------------
#��lambda���ʽ�ж�С�����Ƿ�ƥ��
# encoding=utf-8

def balance(chars):
    def f(cnt, c):
        if cnt >= 0:
            if c == '(':
                cnt += 1
            elif c == ')':
                cnt -= 1
        return cnt

    return 0 == reduce(lambda cnt, c: f(cnt, c), chars, 0)  # ��reduce����Ҫ��ĵ�һ������function�У�Ҫ���һ����������ʼ����ֵ
    # return 0 == reduce(lambda c, cnt: f(cnt, c), chars, 0) #�����lambda���ʽ����Ϊ lambda c, cnt������ʼֵ0�ḳֵ��c�����ַ�����һ���ַ��ḳֵ��cnt


def balance_recurse(chars, cnt=0):
    if len(chars) == 0:
        return cnt == 0

    if chars[0] == "(":
        cnt += 1
    elif chars[0] == ")":
        cnt -= 1
    return cnt >= 0 and balance_recurse(chars[1:], cnt)


PARENS = (("(", ")"), ("[", "]"), ("{", "}"))
def balance_recurse_mul(chars, cnt=[0] * len(PARENS)):
    if len(chars) == 0:
        return all(map(lambda b: b == 0, cnt))
    for i, paren in enumerate(PARENS):  # ����index, seq[index]��������
        if chars[0] == paren[0]:
            cnt[i] += 1
        elif chars[0] == paren[1]:
            cnt[i] -= 1
    return all(map(lambda b: b >= 0, cnt)) and balance_recurse_mul(chars[1:], cnt)


if __name__ == "__main__":
    print balance(")()(")
    print balance("(())")
    print balance("()()")
    
    print balance_recurse(")()(")
    print balance_recurse("(())")
    print balance_recurse("()()")

    print balance_recurse_mul("{([)]}")

�����
False
True
True
False
True
True
True
#-----------------------------------------------------------------------------------------
#dict�����ֹ��췽ʽ
# coding=utf-8

if __name__ == "__main__":
    d = {'a': 10, 'b': 20}
    print d

    d = dict(a=10, b=20)
    print d

    l = [("a", 10), ("b", 20), ('c', 30)]
    d = dict(l)
    print d
�����
{'a': 10, 'b': 20}
{'a': 10, 'b': 20}
{'a': 10, 'c': 30, 'b': 20}
#-----------------------------------------------------------------------------------------
#����ѧϰhttp://python.jobbole.com/81336/
# coding=utf-8
import urllib2

if __name__ == "__main__":
    response = urllib2.urlopen("http://www.baidu.com")
    print response.read()

#post��ʽ��¼ 
# coding=utf-8
import urllib
import urllib2

if __name__ == "__main__":
    values = {"username": "bigben0204", "password": "ilovesjtu8624"}
    data = urllib.urlencode(values)
    url = "https://passport.csdn.net/account/login?from=http://my.csdn.net/my/mycsdn"
    request = urllib2.Request(url, data)
    response = urllib2.urlopen(request)
    print response.read()

#get��ʽ��¼
# coding=utf-8
import urllib
import urllib2

if __name__ == "__main__":
    values = {'username': "bigben0204", 'password': "ilovesjtu8624"}
    data = urllib.urlencode(values)
    url = "http://passport.csdn.net/account/login"
    geturl = url + "?" + data
    request = urllib2.Request(geturl)
    response = urllib2.urlopen(request)
    print response.read()
    
#����Headers
# coding=utf-8
import urllib
import urllib2

if __name__ == "__main__":
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0'
    referer = 'https://passport.csdn.net/account/verify'
    headers = {'User-Agent': user_agent,
               'Referer': referer,
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
               # 'Accept-Encoding': 'gzip, deflate, sdch',
               'Accept-Language': 'zh-CN,zh;q=0.8',
               'Connection': 'keep-alive',
               }

    values = {'username': "bigben0204", 'password': "ilovesjtu8624"}
    data = urllib.urlencode(values)

    url = "https://passport.csdn.net/account/login?from=http://my.csdn.net/my/mycsdn"
    request = urllib2.Request(url, data, headers)
    response = urllib2.urlopen(request)
    page = response.read()
    # print page
    file_object = open('test.html', 'w')
    file_object.write(page)
    file_object.close()

#ʹ��DebugLog
# coding=utf-8
import urllib2

if __name__ == "__main__":
    httpHandler = urllib2.HTTPHandler(debuglevel=1)
    httpsHandler = urllib2.HTTPSHandler(debuglevel=1)
    opener = urllib2.build_opener(httpHandler, httpsHandler)
    urllib2.install_opener(opener)
    response = urllib2.urlopen('http://www.baidu.com')
    page = response.read()
    # print page
    file_object = open('test.html', 'w')
    file_object.write(page)
    file_object.close()
�����
send: 'GET / HTTP/1.1\r\nAccept-Encoding: identity\r\nHost: www.baidu.com\r\nConnection: close\r\nUser-Agent: Python-urllib/2.7\r\n\r\n'
reply: 'HTTP/1.1 200 OK\r\n'
header: Date: Tue, 09 Jan 2018 15:19:11 GMT
header: Content-Type: text/html; charset=utf-8
header: Transfer-Encoding: chunked
header: Connection: Close
header: Vary: Accept-Encoding
header: Set-Cookie: BAIDUID=FE00878D716349FA8FF9E5C929E7F6F2:FG=1; expires=Thu, 31-Dec-37 23:55:55 GMT; max-age=2147483647; path=/; domain=.baidu.com
header: Set-Cookie: BIDUPSID=FE00878D716349FA8FF9E5C929E7F6F2; expires=Thu, 31-Dec-37 23:55:55 GMT; max-age=2147483647; path=/; domain=.baidu.com
header: Set-Cookie: PSTM=1515511151; expires=Thu, 31-Dec-37 23:55:55 GMT; max-age=2147483647; path=/; domain=.baidu.com
header: Set-Cookie: BDSVRTM=0; path=/
header: Set-Cookie: BD_HOME=0; path=/
header: Set-Cookie: H_PS_PSSID=1431_19036_21085_20718; path=/; domain=.baidu.com
header: P3P: CP=" OTI DSP COR IVA OUR IND COM "
header: Cache-Control: private
header: Cxy_all: baidu+b8c667b8da39e37523e1b3045d5162a9
header: Expires: Tue, 09 Jan 2018 15:18:59 GMT
header: X-Powered-By: HPHP
header: Server: BWS/1.1
header: X-UA-Compatible: IE=Edge,chrome=1
header: BDPAGETYPE: 1
header: BDQID: 0xd5adebaf000038b6
header: BDUSERID: 0

#URLError
# coding=utf-8
import urllib2

if __name__ == "__main__":
    requset = urllib2.Request('http://www.xxxxx.com')
    try:
        urllib2.urlopen(requset)
    except urllib2.URLError, e:
        print e.reason
�����
[Errno 10060] 

#2.HTTPError
# coding=utf-8
import urllib2

if __name__ == "__main__":
    req = urllib2.Request('http://blog.csdn.net/cqcre')
    try:
        urllib2.urlopen(req)
    except urllib2.HTTPError, e:
        print e.code
    except urllib2.URLError, e:
        print e.reason
    else:
        print "OK"
#��ǰ�������ж�
# coding=utf-8
import urllib2

if __name__ == "__main__":
    req = urllib2.Request('http://blog.csdn.net/cqcre')
    try:
        urllib2.urlopen(req)
    except urllib2.URLError, e:
        if hasattr(e, "code"):
            print e.code
        if hasattr(e, "reason"):
            print e.reason
    else:
        print "OK"
        

#��ȡCookie���浽����
# coding=utf-8
import cookielib
import urllib2

if __name__ == "__main__":
    cookie = cookielib.CookieJar()
    # ����urllib2���HTTPCookieProcessor����������cookie������
    handler = urllib2.HTTPCookieProcessor(cookie)
    # ͨ��handler������opener
    opener = urllib2.build_opener(handler)
    # �˴���open����ͬurllib2��urlopen������Ҳ���Դ���request
    response = opener.open('http://www.baidu.com')
    for item in cookie:
        print 'Name = ' + item.name
        print 'Value = ' + item.value

#����Cookie���ļ�
# coding=utf-8
import cookielib
import urllib2

if __name__ == "__main__":
    # ���ñ���cookie���ļ���ͬ��Ŀ¼�µ�cookie.txt
    filename = 'cookie.txt'
    # ����һ��MozillaCookieJar����ʵ��������cookie��֮��д���ļ�
    cookie = cookielib.MozillaCookieJar(filename)
    # ����urllib2���HTTPCookieProcessor����������cookie������
    handler = urllib2.HTTPCookieProcessor(cookie)
    # ͨ��handler������opener
    opener = urllib2.build_opener(handler)
    # ����һ������ԭ��ͬurllib2��urlopen
    response = opener.open("http://www.baidu.com")
    # ����cookie���ļ�
    cookie.save(ignore_discard=True, ignore_expires=True)
    
#���ļ��л�ȡCookie������
# coding=utf-8
import cookielib
import urllib2

if __name__ == "__main__":
    # ����MozillaCookieJarʵ������
    cookie = cookielib.MozillaCookieJar()
    # ���ļ��ж�ȡcookie���ݵ�����
    cookie.load('cookie.txt', ignore_discard=True, ignore_expires=True)
    # ���������request
    req = urllib2.Request("http://www.baidu.com")
    # ����urllib2��build_opener��������һ��opener
    opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cookie))
    response = opener.open(req)
    print response.read()
    

#����cookieģ����վ��¼
import urllib
import urllib2
import cookielib
 
filename = 'cookie.txt'
#����һ��MozillaCookieJar����ʵ��������cookie��֮��д���ļ�
cookie = cookielib.MozillaCookieJar(filename)
opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cookie))
postdata = urllib.urlencode({
			'stuid':'201200131012',
			'pwd':'23342321'
		})
#��¼����ϵͳ��URL
loginUrl = 'http://jwxt.sdu.edu.cn:7890/pls/wwwbks/bks_login2.login'
#ģ���¼������cookie���浽����
result = opener.open(loginUrl,postdata)
#����cookie��cookie.txt��
cookie.save(ignore_discard=True, ignore_expires=True)
#����cookie���������һ����ַ������ַ�ǳɼ���ѯ��ַ
gradeUrl = 'http://jwxt.sdu.edu.cn:7890/pls/wwwbks/bkscjcx.curscopre'
#������ʳɼ���ѯ��ַ
result = opener.open(gradeUrl)
print result.read()

����һ������cookie��opener���ڷ��ʵ�¼��URLʱ������¼���cookie����������Ȼ���������cookie������������ַ��
���¼֮����ܲ鿴�ĳɼ���ѯѽ����ѧ�ڿα�ѽ�ȵ���ַ��ģ���¼����ôʵ����

4.Python Reģ��
Python �Դ���reģ�飬���ṩ�˶�������ʽ��֧�֡���Ҫ�õ��ķ����о�����
#����pattern����
re.compile(string[,flag])  
#����Ϊƥ�����ú���
re.match(pattern, string[, flags])
re.search(pattern, string[, flags])
re.split(pattern, string[, maxsplit])
re.findall(pattern, string[, flags])
re.finditer(pattern, string[, flags])
re.sub(pattern, repl, string[, count])
re.subn(pattern, repl, string[, count])

? re.I(ȫƴ��IGNORECASE): ���Դ�Сд��������������д������ͬ��
? re.M(ȫƴ��MULTILINE): ����ģʽ���ı�'^'��'$'����Ϊ���μ���ͼ��
? re.S(ȫƴ��DOTALL): ������ƥ��ģʽ���ı�'.'����Ϊ
? re.L(ȫƴ��LOCALE): ʹԤ���ַ��� \w \W \b \B \s \S ȡ���ڵ�ǰ�����趨
? re.U(ȫƴ��UNICODE): ʹԤ���ַ��� \w \W \b \B \s \S \d \D ȡ����unicode������ַ�����
? re.X(ȫƴ��VERBOSE): ��ϸģʽ�����ģʽ��������ʽ�����Ƕ��У����Կհ��ַ��������Լ���ע�͡�

//
# coding=utf-8
import re

if __name__ == "__main__":
    pattern = re.compile(r'hello')
    # ʹ��re.matchƥ���ı������ƥ�������޷�ƥ��ʱ������None
    result1 = re.match(pattern, 'hello hello')
    result2 = re.match(pattern, 'helloo CQC!') #�����ͷƥ�䣬���aahello���޷�ƥ��
    result3 = re.match(pattern, 'helo CQC!')
    result4 = re.match(pattern, 'hello CQC!')

    # ���1ƥ��ɹ�
    if result1:
        # ʹ��Match��÷�����Ϣ
        print result1.group()
        print result1.string
        print result1.re #ͬ pattern
        print result1.endpos
        print result1.lastindex
        print result1.lastgroup
    else:
        print '1ƥ��ʧ�ܣ�'

    # ���2ƥ��ɹ�
    if result2:
        # ʹ��Match��÷�����Ϣ
        print result2.group()
    else:
        print '2ƥ��ʧ�ܣ�'

    # ���3ƥ��ɹ�
    if result3:
        # ʹ��Match��÷�����Ϣ
        print result3.group()
    else:
        print '3ƥ��ʧ�ܣ�'

    # ���4ƥ��ɹ�
    if result4:
        # ʹ��Match��÷�����Ϣ
        print result4.group()
    else:
        print '4ƥ��ʧ�ܣ�'
�����
hello
hello hello
<_sre.SRE_Pattern object at 0x04A1B840>
11
None
None
hello
3ƥ��ʧ�ܣ�
hello

ƥ�����
1.��һ��ƥ�䣬pattern������ʽΪ��hello��������ƥ���Ŀ���ַ���stringҲΪhello����ͷ��β��ȫƥ�䣬ƥ��ɹ���
2.�ڶ���ƥ�䣬stringΪhelloo CQC����stringͷ��ʼƥ��pattern��ȫ����ƥ�䣬patternƥ�������ͬʱƥ����ֹ�������o CQC����ƥ�䣬����ƥ��ɹ�����Ϣ��
3.������ƥ�䣬stringΪhelo CQC����stringͷ��ʼƥ��pattern�����ֵ� ��o�� ʱ�޷����ƥ�䣬ƥ����ֹ������None
4.���ĸ�ƥ�䣬ͬ�ڶ���ƥ��ԭ����ʹ�����˿ո��Ҳ������Ӱ�졣

#����
# coding=utf-8
import re

if __name__ == "__main__":
    m = re.match(r'(\w+) (\w+)(?P<sign>.*)', 'hello world!') #!hello world!��ƥ�䲻��

    print "m.string:", m.string
    print "m.re:", m.re
    print "m.pos:", m.pos
    print "m.endpos:", m.endpos
    print "m.lastindex:", m.lastindex #ָ���һ��������������(?P<sign>.*)��������
    print "m.lastgroup:", m.lastgroup #ָ���һ������ı���
    print "m.group():", m.group() #ƥ�䵽���ַ���
    print "m.group(1,2):", m.group(1, 2)
    print "m.groups():", m.groups()
    print "m.groupdict():", m.groupdict()
    print "m.start(2):", m.start(2) #��2��������ʼλ��
    print "m.end(2):", m.end(2) #��2���������λ��
    print "m.span(2):", m.span(2) #��2��������ʼ�ͽ���λ��
    print r"m.expand(r'\g \g\g'):", m.expand(r'\2 \1\3')
�����
m.string: hello world!
m.re: <_sre.SRE_Pattern object at 0x02719760>
m.pos: 0
m.endpos: 12
m.lastindex: 3
m.lastgroup: sign
m.group(): hello world!
m.group(1,2): ('hello', 'world')
m.groups(): ('hello', 'world', '!')
m.groupdict(): {'sign': '!'}
m.start(2): 6
m.end(2): 11
m.span(2): (6, 11)
m.expand(r'\g \g\g'): world hello!

��2��re.search(pattern, string[, flags])
search������match�����������ƣ���������match()����ֻ���re�ǲ�����string�Ŀ�ʼλ��ƥ�䣬search()��ɨ������string����ƥ�䣬match����ֻ����0λ��ƥ��ɹ��Ļ����з��أ�������ǿ�ʼλ��ƥ��ɹ��Ļ���match()�ͷ���None��ͬ����search�����ķ��ض���ͬ��match()���ض���ķ��������ԡ�������һ�����Ӹ���һ��
# coding=utf-8
import re

if __name__ == "__main__":
    # ��������ʽ�����Pattern����
    pattern = re.compile(r'world')
    # ʹ��search()����ƥ����Ӵ�����������ƥ����Ӵ�ʱ������None
    # ���������ʹ��match()�޷��ɹ�ƥ��
    match = re.search(pattern, 'hello world!')
    if match:
        # ʹ��Match��÷�����Ϣ
        print match.group()
    ### ��� ###
    # world

��3��re.split(pattern, string[, maxsplit])
�����ܹ�ƥ����Ӵ���string�ָ�󷵻��б�maxsplit����ָ�����ָ��������ָ����ȫ���ָ����ͨ����������Ӹ���һ�¡�
# coding=utf-8
import re

if __name__ == "__main__":
    pattern = re.compile(r'\d+')
    print re.split(pattern, 'one1two2three3four4')

    ### ��� ###
    # ['one', 'two', 'three', 'four', '']

��4��re.findall(pattern, string[, flags])
����string�����б���ʽ����ȫ����ƥ����Ӵ�������ͨ���������������һ��
# coding=utf-8
import re

if __name__ == "__main__":
    pattern = re.compile(r'\d+')
    print re.findall(pattern, 'one1two2three3four4')

    ### ��� ###
    # ['1', '2', '3', '4']

��5��re.finditer(pattern, string[, flags])
����string������һ��˳�����ÿһ��ƥ������Match���󣩵ĵ�����������ͨ�����������������һ��
# coding=utf-8
import re

if __name__ == "__main__":
    pattern = re.compile(r'\d+')
    for m in re.finditer(pattern, 'one1two2three3four4'):
        print m.group(),

    ### ��� ###
    # 1 2 3 4

��6��re.sub(pattern, repl, string[, count])
ʹ��repl�滻string��ÿһ��ƥ����Ӵ��󷵻��滻����ַ�����
��repl��һ���ַ���ʱ������ʹ��\id��\g��\g���÷��飬������ʹ�ñ��0��
��repl��һ������ʱ���������Ӧ��ֻ����һ��������Match���󣩣�������һ���ַ��������滻�����ص��ַ����в��������÷��飩��
count����ָ������滻��������ָ��ʱȫ���滻��
# coding=utf-8
import re


def func(m):
    return m.group(1).title() + ' ' + m.group(2).title()


if __name__ == "__main__":
    pattern = re.compile(r'(\w+) (\w+)')
    s = 'i say, hello world!'

    print re.sub(pattern, r'\2 \1', s)
    print re.sub(pattern, func, s)
    ### output ###
    # say i, world hello!
    # I Say, Hello World!

��7��re.subn(pattern, repl, string[, count])
���� (sub(repl, string[, count]), �滻����)��
# coding=utf-8
import re


def func(m):
    return m.group(1).title() + ' ' + m.group(2).title()


if __name__ == "__main__":
    pattern = re.compile(r'(\w+) (\w+)')
    s = 'i say, hello world!'

    print re.subn(pattern, func, s) #pattern.subn(func, s)
    print re.subn(pattern, r'\2 \1', s) #pattern.subn(r'\2 \1', s)

    ### output ###
    # ('I Say, Hello World!', 2)
    # ('say i, world hello!', 2)

5. Python Reģ�����һ��ʹ�÷�ʽ
���������ǽ�����7�����߷���������match��search�ȵȣ��������÷�ʽ���� re.match��re.search�ķ�ʽ����ʵ��������һ�ֵ��÷�ʽ������ͨ��pattern.match��pattern.search���ã����� ���ñ㲻�ý�pattern��Ϊ��һ�����������ˣ�������������ýԿɡ�
Pattern.match(string[, pos[, endpos]]) | re.match(pattern, string[, flags])
Pattern.search(string[, pos[, endpos]]) | re.search(pattern, string[, flags])
Pattern.split(string[, maxsplit]) | re.split(pattern, string[, maxsplit])
Pattern.findall(string[, pos[, endpos]]) | re.findall(pattern, string[, flags])
Pattern.finditer(string[, pos[, endpos]]) | re.finditer(pattern, string[, flags])
Pattern.sub(repl, string[, count]) | re.sub(pattern, repl, string[, count])
Pattern.subn(repl, string[, count]) |re.sub(pattern, repl, string[, count])

4. ���� Beautiful Soup ����
# coding=utf-8

from bs4 import BeautifulSoup

html = """
<html><head><title>The Dormouse's story</title></head>
<body>
<p class="title" name="dromouse"><b>The Dormouse's story</b></p>
<p class="story">Once upon a time there were three little sisters; and their names were
<a href="http://example.com/elsie" class="sister" id="link1"><!-- Elsie --></a>,
<a href="http://example.com/lacie" class="sister" id="link2">Lacie</a> and
<a href="http://example.com/tillie" class="sister" id="link3">Tillie</a>;
and they lived at the bottom of a well.</p>
<p class="story">...</p>
"""

if __name__ == "__main__":
    soup = BeautifulSoup(html, "lxml") #�������lxml������ʾ��ǰ�ڱ���ʹ��lxml��ȥ������������ʾ��һ�£�soup = BeautifulSoup(open('index.html'))
    print soup.prettify()
�����
<html>
 <head>
  <title>
   The Dormouse's story
  </title>
 </head>
 <body>
  <p class="title" name="dromouse">
   <b>
    The Dormouse's story
   </b>
  </p>
  <p class="story">
   Once upon a time there were three little sisters; and their names were
   <a class="sister" href="http://example.com/elsie" id="link1">
    <!-- Elsie -->
   </a>
   ,
   <a class="sister" href="http://example.com/lacie" id="link2">
    Lacie
   </a>
   and
   <a class="sister" href="http://example.com/tillie" id="link3">
    Tillie
   </a>
   ;
and they lived at the bottom of a well.
  </p>
  <p class="story">
   ...
  </p>
 </body>
</html>

5. �Ĵ��������
Beautiful Soup������HTML�ĵ�ת����һ�����ӵ����νṹ,ÿ���ڵ㶼��Python����,���ж�����Թ���Ϊ4��:
Tag
NavigableString
BeautifulSoup
Comment
�������ǽ���һһ����
��1��Tag
Tag ��ʲô��ͨ�׵㽲���� HTML �е�һ������ǩ������
print soup.title
print soup.head
print soup.a
print soup.p
�����
<title>The Dormouse's story</title>
<head><title>The Dormouse's story</title></head>
<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>
<p class="title" name="dromouse"><b>The Dormouse's story</b></p>
���ǿ������� soup�ӱ�ǩ�����ɵػ�ȡ��Щ��ǩ�����ݣ��ǲ��Ǹо���������ʽ������ˣ�������һ���ǣ������ҵ��������������еĵ�һ������Ҫ��ı�ǩ�����Ҫ��ѯ���еı�ǩ�������ں�����н��ܡ�

���ǿ�����֤һ����Щ���������
print type(soup.a)
# <class 'bs4.element.Tag'>

���� Tag������������Ҫ�����ԣ��� name �� attrs���������Ƿֱ�������һ��
--name
print soup.name
print soup.head.name
#[document]
#head
soup ������Ƚ����⣬���� name ��Ϊ [document]�����������ڲ���ǩ�������ֵ��Ϊ��ǩ��������ơ�

--attrs
print soup.p.attrs
#{'class': ['title'], 'name': 'dromouse'}
��������ǰ� p ��ǩ���������Դ�ӡ����˳������õ���������һ���ֵ䡣

���������Ҫ������ȡĳ�����ԣ������������������ǻ�ȡ���� class ��ʲô
print soup.p['class']
#['title']
����������������get�������������Ե����ƣ������ǵȼ۵�
print soup.p.get('class')
#['title']
���ǿ��Զ���Щ���Ժ����ݵȵȽ����޸ģ�����
soup.p['class']="newClass"
print soup.p
#<p class="newClass" name="dromouse"><b>The Dormouse's story</b></p>
�����Զ�������Խ���ɾ��������
del soup.p['class']
print soup.p
#<p name="dromouse"><b>The Dormouse's story</b></p>

��2��NavigableString
��Ȼ�����Ѿ��õ��˱�ǩ�����ݣ���ô�������ˣ�����Ҫ���ȡ��ǩ�ڲ���������ô���أ��ܼ򵥣��� .string ���ɣ�����
print soup.p.string
#The Dormouse's story

�������Ǿ����ɻ�ȡ���˱�ǩ��������ݣ����������������ʽҪ���鷳������������һ�� NavigableString����������� ���Ա������ַ���������������û��ǳ���Ӣ�����ְɡ�
print type(soup.p.string)
#<class 'bs4.element.NavigableString'>

��3��BeautifulSoup
BeautifulSoup �����ʾ����һ���ĵ���ȫ������.�󲿷�ʱ��,���԰������� Tag ������һ������� Tag�����ǿ��Էֱ��ȡ�������ͣ����ƣ��Լ�����������һ��
print type(soup.name)
#<type 'unicode'>
print soup.name 
# [document]
print soup.attrs 
#{} ���ֵ�

��4��Comment
Comment ������һ���������͵� NavigableString ������ʵ�����������Ȼ������ע�ͷ��ţ�����������úô����������ܻ�����ǵ��ı�����������벻�����鷳��
������һ����ע�͵ı�ǩ
print soup.a
print soup.a.string
print type(soup.a.string)
Python �C ��������
��ҳ�������¹۵��붯̬����֪ʶϵ�н̳�ʵ����Ŀ�������ܹ�����ԴPythonС��
�������� > Python - �������� > �������� > ʵ����Ŀ > Python�������ţ�8����Beautiful Soup���÷�
Python�������ţ�8����Beautiful Soup���÷�

2015/04/25 �� ʵ����Ŀ, ϵ�н̳� �� 6 ���� �� ����
������ 27
�������ߣ� �������� - ����� ��δ��������ɣ���ֹת�أ�
��ӭ���벮������ ר�����ߡ�
Python�������ţ�1��������
Python�������ţ�2������������˽�
Python�������ţ�3����Urllib��Ļ���ʹ��
Python�������ţ�4����Urllib��ĸ߼��÷�
Python�������ţ�5����URLError�쳣����
Python�������ţ�6����Cookie��ʹ��
Python�������ţ�7����������ʽ
Python�������ţ�8����Beautiful Soup���÷�
��һ�����ǽ�����������ʽ������������ʵ��������ģ����һ������ƥ�����в�أ��ǿ��ܳ���ʹ������õ�ѭ��֮�У������е�С�����Ҳ��д����� ��ʽ��д���õò�������û��ϵ�����ǻ���һ����ǿ��Ĺ��ߣ���Beautiful Soup�����������ǿ��Ժܷ������ȡ��HTML��XML��ǩ�е����ݣ�ʵ���Ƿ��㣬��һ�ھ�������һ��������һ��Beautiful Soup��ǿ��ɡ�

1. Beautiful Soup�ļ��

����˵��Beautiful Soup��python��һ���⣬����Ҫ�Ĺ����Ǵ���ҳץȡ���ݡ��ٷ��������£�

Beautiful Soup�ṩһЩ�򵥵ġ�pythonʽ�ĺ����������������������޸ķ������ȹ��ܡ�����һ�������䣬ͨ�������ĵ�Ϊ�û��ṩ��Ҫץȡ�����ݣ���Ϊ�򵥣����Բ���Ҫ���ٴ���Ϳ���д��һ��������Ӧ�ó���

Beautiful Soup�Զ��������ĵ�ת��ΪUnicode���룬����ĵ�ת��Ϊutf-8���롣�㲻��Ҫ���Ǳ��뷽ʽ�������ĵ�û��ָ��һ�����뷽ʽ����ʱ��Beautiful Soup�Ͳ����Զ�ʶ����뷽ʽ�ˡ�Ȼ���������Ҫ˵��һ��ԭʼ���뷽ʽ�Ϳ����ˡ�

Beautiful Soup�ѳ�Ϊ��lxml��html6libһ����ɫ��python��������Ϊ�û������ṩ��ͬ�Ľ������Ի�ǿ�����ٶȡ�
�ϻ�����˵����������һ�°�~

2. Beautiful Soup ��װ

Beautiful Soup 3 Ŀǰ�Ѿ�ֹͣ�������Ƽ������ڵ���Ŀ��ʹ��Beautiful Soup 4���������Ѿ�����ֲ��BS4�ˣ�Ҳ����˵����ʱ������Ҫ import bs4 ���������������õİ汾�� Beautiful Soup 4.3.2 (���BS4)�������˵ BS4 �� Python3 ��֧�ֲ����ã��������õ��� Python2.7.7�������С����õ��� Python3 �汾�����Կ������� BS3 �汾��

������õ����°��Debain��Ubuntu,��ô����ͨ��ϵͳ���������������װ���������������°汾��Ŀǰ��4.2.1�汾

Python

sudo apt-get install Python-bs4
1
sudo apt-get install Python-bs4
����밲װ���µİ汾����ֱ�����ذ�װ�����ֶ���װ��Ҳ��ʮ�ַ���ķ������������Ұ�װ���� Beautiful Soup 4.3.2

Beautiful Soup 3.2.1Beautiful Soup 4.3.2

�������֮���ѹ

����������������ɰ�װ

Python

sudo python setup.py install
1
sudo python setup.py install
����ͼ��ʾ��֤����װ�ɹ���

2015-03-11 00:15:41 ����Ļ��ͼ

Ȼ����Ҫ��װ lxml

Python

sudo apt-get install Python-lxml
1
sudo apt-get install Python-lxml
Beautiful Soup֧��Python��׼���е�HTML������,��֧��һЩ�������Ľ�������������ǲ���װ������ Python ��ʹ�� PythonĬ�ϵĽ�������lxml ����������ǿ���ٶȸ��죬�Ƽ���װ��

3. ����Beautiful Soup ֮��

�������ȷ���ٷ��ĵ����ӣ�������������Щ�࣬Ҳ���������ڴ˱�������һ���������Ҳο���

�ٷ��ĵ�

4. ���� Beautiful Soup ����

���ȱ���Ҫ���� bs4 ��

Python

from bs4 import BeautifulSoup
1
from bs4 import BeautifulSoup
���Ǵ���һ���ַ�����������������Ǳ����������ʾ

Python

html = """
<html><head><title>The Dormouse's story</title></head>
<body>
<p class="title" name="dromouse"><b>The Dormouse's story</b></p>
<p class="story">Once upon a time there were three little sisters; and their names were
<a href="http://example.com/elsie" class="sister" id="link1"><!-- Elsie --></a>,
<a href="http://example.com/lacie" class="sister" id="link2">Lacie</a> and
<a href="http://example.com/tillie" class="sister" id="link3">Tillie</a>;
and they lived at the bottom of a well.</p>
<p class="story">...</p>
"""
1
2
3
4
5
6
7
8
9
10
11
html = """
<html><head><title>The Dormouse's story</title></head>
<body>
<p class="title" name="dromouse"><b>The Dormouse's story</b></p>
<p class="story">Once upon a time there were three little sisters; and their names were
<a href="http://example.com/elsie" class="sister" id="link1"><!-- Elsie --></a>,
<a href="http://example.com/lacie" class="sister" id="link2">Lacie</a> and
<a href="http://example.com/tillie" class="sister" id="link3">Tillie</a>;
and they lived at the bottom of a well.</p>
<p class="story">...</p>
"""
���� beautifulsoup ����

Python

soup = BeautifulSoup(html)
1
soup = BeautifulSoup(html)
���⣬���ǻ������ñ��� HTML �ļ���������������

Python

soup = BeautifulSoup(open('index.html'))
1
soup = BeautifulSoup(open('index.html'))
������������ǽ����� index.html �ļ��򿪣����������� soup ����

������������ӡһ�� soup ��������ݣ���ʽ�����

Python

print soup.prettify()
1
print soup.prettify()
Python

&lt;html&gt;
 &lt;head&gt;
  &lt;title&gt;
   The Dormouse's story
  &lt;/title&gt;
 &lt;/head&gt;
 &lt;body&gt;
  &lt;p class="title" name="dromouse"&gt;
   &lt;b&gt;
    The Dormouse's story
   &lt;/b&gt;
  &lt;/p&gt;
  &lt;p class="story"&gt;
   Once upon a time there were three little sisters; and their names were
   &lt;a class="sister" href="http://example.com/elsie" id="link1"&gt;
    &lt;!-- Elsie --&gt;
   &lt;/a&gt;
   ,
   &lt;a class="sister" href="http://example.com/lacie" id="link2"&gt;
    Lacie
   &lt;/a&gt;
   and
   &lt;a class="sister" href="http://example.com/tillie" id="link3"&gt;
    Tillie
   &lt;/a&gt;
   ;
and they lived at the bottom of a well.
  &lt;/p&gt;
  &lt;p class="story"&gt;
   ...
  &lt;/p&gt;
 &lt;/body&gt;
&lt;/html&gt;
1
2
3
4
5
6
7
8
9
10
11
12
13
14
15
16
17
18
19
20
21
22
23
24
25
26
27
28
29
30
31
32
33
&lt;html&gt;
 &lt;head&gt;
  &lt;title&gt;
   The Dormouse's story
  &lt;/title&gt;
 &lt;/head&gt;
 &lt;body&gt;
  &lt;p class="title" name="dromouse"&gt;
   &lt;b&gt;
    The Dormouse's story
   &lt;/b&gt;
  &lt;/p&gt;
  &lt;p class="story"&gt;
   Once upon a time there were three little sisters; and their names were
   &lt;a class="sister" href="http://example.com/elsie" id="link1"&gt;
    &lt;!-- Elsie --&gt;
   &lt;/a&gt;
   ,
   &lt;a class="sister" href="http://example.com/lacie" id="link2"&gt;
    Lacie
   &lt;/a&gt;
   and
   &lt;a class="sister" href="http://example.com/tillie" id="link3"&gt;
    Tillie
   &lt;/a&gt;
   ;
and they lived at the bottom of a well.
  &lt;/p&gt;
  &lt;p class="story"&gt;
   ...
  &lt;/p&gt;
 &lt;/body&gt;
&lt;/html&gt;
���ϱ�������������ʽ����ӡ�����������ݣ�������������õ���С�����Ҫ�Ǻÿ���

5. �Ĵ��������

Beautiful Soup������HTML�ĵ�ת����һ�����ӵ����νṹ,ÿ���ڵ㶼��Python����,���ж�����Թ���Ϊ4��:

Tag
NavigableString
BeautifulSoup
Comment
�������ǽ���һһ����

��1��Tag

Tag ��ʲô��ͨ�׵㽲���� HTML �е�һ������ǩ������
<title>The Dormouse's story</title>

&lt;a class="sister" href="http://example.com/elsie" id="link1"&gt;Elsie&lt;/a&gt;
1
&lt;a class="sister" href="http://example.com/elsie" id="link1"&gt;Elsie&lt;/a&gt;
����� title a �ȵ� HTML ��ǩ����������������ݾ��� Tag����������������һ�������� Beautiful Soup ������ػ�ȡ Tags

����ÿһ�δ�����ע�Ͳ��ּ�Ϊ���н��
print soup.title
#<title>The Dormouse's story</title>
print soup.head
#<head><title>The Dormouse's story</title></head>
print soup.a
#<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>
print soup.p #ͬsoup.body.p
#<p class="title" name="dromouse"><b>The Dormouse's story</b></p>

���ǿ������� soup�ӱ�ǩ�����ɵػ�ȡ��Щ��ǩ�����ݣ��ǲ��Ǹо���������ʽ������ˣ�������һ���ǣ������ҵ��������������еĵ�һ������Ҫ��ı�ǩ�����Ҫ��ѯ���еı�ǩ�������ں�����н��ܡ�

���ǿ�����֤һ����Щ���������
print type(soup.a)
#<class 'bs4.element.Tag'>

���� Tag������������Ҫ�����ԣ��� name �� attrs���������Ƿֱ�������һ��
name
print soup.name
print soup.head.name
#[document]
#head
soup ������Ƚ����⣬���� name ��Ϊ [document]�����������ڲ���ǩ�������ֵ��Ϊ��ǩ��������ơ�

attrs
print soup.p.attrs
#{'class': ['title'], 'name': 'dromouse'}
��������ǰ� p ��ǩ���������Դ�ӡ����˳������õ���������һ���ֵ䡣

���������Ҫ������ȡĳ�����ԣ������������������ǻ�ȡ���� class ��ʲô
print soup.p['class']
#['title']

����������������get�������������Ե����ƣ������ǵȼ۵�
print soup.p.get('class')
#['title']

���ǿ��Զ���Щ���Ժ����ݵȵȽ����޸ģ�����
soup.p['class']="newClass"
print soup.p

#<p class="newClass" name="dromouse"><b>The Dormouse's story</b></p>
�����Զ�������Խ���ɾ��������

del soup.p['class']
print soup.p
#<p name="dromouse"><b>The Dormouse's story</b></p>

�����������޸�ɾ���Ĳ������������ǵ���Ҫ��;���ڴ˲�����ϸ�����ˣ��������Ҫ����鿴ǰ���ṩ�Ĺٷ��ĵ�

��2��NavigableString

��Ȼ�����Ѿ��õ��˱�ǩ�����ݣ���ô�������ˣ�����Ҫ���ȡ��ǩ�ڲ���������ô���أ��ܼ򵥣��� .string ���ɣ�����
print soup.p.string
#The Dormouse's story
�������Ǿ����ɻ�ȡ���˱�ǩ��������ݣ����������������ʽҪ���鷳������������һ�� NavigableString����������� ���Ա������ַ���������������û��ǳ���Ӣ�����ְɡ�
print type(soup.p.string)
#<class 'bs4.element.NavigableString'>
�����һ����������
print type(soup.p.string)
#<class 'bs4.element.NavigableString'>

��3��BeautifulSoup

BeautifulSoup �����ʾ����һ���ĵ���ȫ������.�󲿷�ʱ��,���԰������� Tag ������һ������� Tag�����ǿ��Էֱ��ȡ�������ͣ����ƣ��Լ�����������һ��
print type(soup.name)
#<type 'unicode'> #�������н����<type 'str'>
print soup.name 
# [document]
print soup.attrs 
#{} ���ֵ�

��4��Comment
Comment ������һ���������͵� NavigableString ������ʵ�����������Ȼ������ע�ͷ��ţ�����������úô����������ܻ�����ǵ��ı�����������벻�����鷳��
������һ����ע�͵ı�ǩ
print soup.a
print soup.a.string
print type(soup.a.string)

���н������
<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>
 Elsie 
<class 'bs4.element.Comment'>

a ��ǩ�������ʵ������ע�ͣ���������������� .string ������������ݣ����Ƿ������Ѿ���ע�ͷ���ȥ���ˣ���������ܻ�����Ǵ�������Ҫ���鷳��
�������Ǵ�ӡ������������ͣ���������һ�� Comment ���ͣ����ԣ�������ʹ��ǰ�����һ���жϣ��жϴ�������
if type(soup.a.string) == bs4.element.Comment:
    print soup.a.string
����Ĵ����У����������ж����������ͣ��Ƿ�Ϊ Comment ���ͣ�Ȼ���ٽ����������������ӡ�����

6. �����ĵ���
��1��ֱ���ӽڵ�
Ҫ�㣺.contents .children ����
.contents
tag �� .content ���Կ��Խ�tag���ӽڵ����б�ķ�ʽ���
print soup.head.contents 
#[<title>The Dormouse's story</title>]
�����ʽΪ�б����ǿ������б���������ȡ����ĳһ��Ԫ��
print soup.head.contents[0]
#<title>The Dormouse's story</title>

.children
�����صĲ���һ�� list���������ǿ���ͨ��������ȡ�����ӽڵ㡣
���Ǵ�ӡ��� .children ��һ�£����Է�������һ�� list ����������
print soup.head.children
# <listiterator object at 0x0000000005237E48>
for child in soup.head.children:
    print child
# <title>The Dormouse's story</title>

for child in soup.body.children:
    print child
    print "type: ", type(child) #����˺ܶ���У���Ϊ��NavigableString
�����

type:  <class 'bs4.element.NavigableString'>
<p class="title" name="dromouse"><b>The Dormouse's story</b></p>
type:  <class 'bs4.element.Tag'>


type:  <class 'bs4.element.NavigableString'>
<p class="story">Once upon a time there were three little sisters; and their names were
<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>,
<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a> and
<a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>;
and they lived at the bottom of a well.</p>
type:  <class 'bs4.element.Tag'>


type:  <class 'bs4.element.NavigableString'>
<p class="story">...</p>
type:  <class 'bs4.element.Tag'>


type:  <class 'bs4.element.NavigableString'>

��������еĻ��з�ȥ�������ӡ����
<p class="title" name="dromouse"><b>The Dormouse's story</b></p>
type:  <class 'bs4.element.Tag'>
<p class="story">Once upon a time there were three little sisters; and their names were
<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>,
<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a> and
<a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>;
and they lived at the bottom of a well.</p>
type:  <class 'bs4.element.Tag'>
<p class="story">...</p>
type:  <class 'bs4.element.Tag'>

��2����������ڵ�
֪ʶ�㣺.descendants ����
.descendants
.contents �� .children ���Խ�����tag��ֱ���ӽڵ㣬.descendants ���Կ��Զ�����tag������ڵ���еݹ�ѭ������ children���ƣ�����Ҳ��Ҫ������ȡ���е����ݡ�
for child in soup.descendants:
    print child
���н�����£����Է��֣����еĽڵ㶼����ӡ�����ˣ����������� HTML��ǩ����δ� head ��ǩһ�������룬�Դ����ơ�

3���ڵ�����
֪ʶ�㣺.string ����
���tagֻ��һ�� NavigableString �����ӽڵ�,��ô���tag����ʹ�� .string �õ��ӽڵ㡣���һ��tag����һ���ӽڵ�,��ô���tagҲ����ʹ�� .string ����,�������뵱ǰΨһ�ӽڵ�� .string �����ͬ��
ͨ�׵�˵���ǣ����һ����ǩ����û�б�ǩ�ˣ���ô .string �ͻ᷵�ر�ǩ��������ݡ������ǩ����ֻ��Ψһ��һ����ǩ�ˣ���ô .string Ҳ�᷵������������ݡ�����
print soup.head.string
#The Dormouse's story
print soup.title.string #ͬsoup.title.get_text()
#The Dormouse's story

���tag�����˶���ӽڵ�,tag���޷�ȷ����string ����Ӧ�õ����ĸ��ӽڵ������, .string ���������� None
print soup.html.string
# None

��4���������
֪ʶ�㣺 .strings .stripped_strings ����
.strings
��ȡ������ݣ�������Ҫ������ȡ���������������
for string in soup.html.strings:
    print (repr(string))
�����
u"The Dormouse's story"
u'\n'
u'\n'
u"The Dormouse's story"
u'\n'
u'Once upon a time there were three little sisters; and their names were\n'
u',\n'
u'Lacie'
u' and\n'
u'Tillie'
u';\nand they lived at the bottom of a well.'
u'\n'
u'...'

.stripped_strings
������ַ����п��ܰ����˺ܶ�ո�����,ʹ�� .stripped_strings ����ȥ������հ�����
for string in soup.stripped_strings:
    print(repr(string))
    # u"The Dormouse's story"
    # u"The Dormouse's story"
    # u'Once upon a time there were three little sisters; and their names were'
    # u'Elsie'
    # u','
    # u'Lacie'
    # u'and'
    # u'Tillie'
    # u';\nand they lived at the bottom of a well.'
    # u'...'

��5�����ڵ�
֪ʶ�㣺 .parent ����
p = soup.p
print p.parent.name
#body

content = soup.head.title.string
print content.parent.name
#title

��6��ȫ�����ڵ�
֪ʶ�㣺.parents ����
ͨ��Ԫ�ص� .parents ���Կ��Եݹ�õ�Ԫ�ص����и����ڵ㣬����
content = soup.head.title.string
for parent in  content.parents:
    print parent.name
title
head
html
[document]

��7���ֵܽڵ�
֪ʶ�㣺.next_sibling .previous_sibling ����
�ֵܽڵ�������Ϊ�ͱ��ڵ㴦��ͳһ���Ľڵ㣬.next_sibling ���Ի�ȡ�˸ýڵ����һ���ֵܽڵ㣬.previous_sibling ����֮�෴������ڵ㲻���ڣ��򷵻� None
ע�⣺ʵ���ĵ��е�tag�� .next_sibling �� .previous_sibling ����ͨ�����ַ�����հף���Ϊ�հ׻��߻���Ҳ���Ա�����һ���ڵ㣬���Եõ��Ľ�������ǿհ׻��߻���
print repr(soup.p.next_sibling)
#       ʵ�ʸô�Ϊ�հ� u'\n'
print soup.p.prev_sibling #ԭ������ƴ��prev_sibling���Ǹ�û�е����ԣ����Դ����None
# None   û��ǰһ���ֵܽڵ㣬���� None
print repr(soup.p.previous_sibling)
# u'\n'
print soup.p.previous_sibling.previous_sibling
# None
print soup.p.next_sibling.next_sibling
# <p class="story">Once upon a time there were three little sisters; and their names were
# <a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>,
# <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a> and
# <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>;
# and they lived at the bottom of a well.</p>
# ��һ���ڵ����һ���ֵܽڵ������ǿ��Կ����Ľڵ�

��8��ȫ���ֵܽڵ�
֪ʶ�㣺.next_siblings .previous_siblings ����
ͨ�� .next_siblings �� .previous_siblings ���Կ��ԶԵ�ǰ�ڵ���ֵܽڵ�������
for sibling in soup.a.next_siblings:
    print repr(sibling)
# u',\n'
# <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>
# u' and\n'
# <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>
# u';\nand they lived at the bottom of a well.'

��9��ǰ��ڵ�
֪ʶ�㣺.next_element .previous_element ����
�� .next_sibling .previous_sibling ��ͬ����������������ֵܽڵ㣬���������нڵ㣬���ֲ��
���� head �ڵ�Ϊ
<head><title>The Dormouse's story</title></head>
��ô������һ���ڵ���� title�����ǲ��ֲ�ι�ϵ��
print soup.head.next_element
#<title>The Dormouse's story</title>

�Ƚ����£�
print soup.p.previous_sibling.previous_sibling
# None
print soup.p.previous_element.previous_element
# <body>
# <p class="title" name="dromouse"><b>The Dormouse's story</b></p>
# <p class="story">Once upon a time there were three little sisters; and their names were
# <a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>,
# <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a> and
# <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>;
# and they lived at the bottom of a well.</p>
# <p class="story">...</p></body>

��10������ǰ��ڵ�
֪ʶ�㣺.next_elements .previous_elements ����
ͨ�� .next_elements �� .previous_elements �ĵ������Ϳ�����ǰ���������ĵ��Ľ�������,�ͺ����ĵ����ڱ�����һ��
aTags = soup.find_all("a")
print aTags
last_a_tag = aTags[len(aTags) - 1]
for element in last_a_tag.next_elements:
    print(repr(element))
�����
[<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>, <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>, <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>]
u'Tillie'
u';\nand they lived at the bottom of a well.'
u'\n'
<p class="story">...</p>
u'...'
u'\n'

7.�����ĵ���
��1��find_all( name , attrs , recursive , text , **kwargs )
find_all() ����������ǰtag������tag�ӽڵ�,���ж��Ƿ���Ϲ�����������
1��name ����
name �������Բ�����������Ϊ name ��tag,�ַ�������ᱻ�Զ����Ե�
A.���ַ���
��򵥵Ĺ��������ַ���.�����������д���һ���ַ�������,Beautiful Soup��������ַ�������ƥ�������,������������ڲ����ĵ������е�<b>��ǩ
print soup.find_all('b')
# [<b>The Dormouse's story</b>]
print soup.find_all('a')
#[<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>, <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>, <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>]

B.��������ʽ
�������������ʽ��Ϊ����,Beautiful Soup��ͨ��������ʽ�� match() ��ƥ������.�����������ҳ�������b��ͷ�ı�ǩ,���ʾ<body>��<b>��ǩ��Ӧ�ñ��ҵ�
import re
for tag in soup.find_all(re.compile("^b")):
    print(tag.name)
# body
# b

C.���б�
��������б����,Beautiful Soup�Ὣ���б�����һԪ��ƥ������ݷ���.��������ҵ��ĵ�������<a>��ǩ��<b>��ǩ
for tag in soup.find_all(["a", "b"]):
    print tag
# [<b>The Dormouse's story</b>,
#  <a class="sister" href="http://example.com/elsie" id="link1">Elsie</a>,
#  <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>,
#  <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>]

D.�� True
True ����ƥ���κ�ֵ,���������ҵ����е�tag,���ǲ��᷵���ַ����ڵ�
for tag in soup.find_all(True):
    print tag.name
# html
# head
# title
# body
# p
# b
# p
# a
# a
# a
# p

E.������
���û�к��ʹ�����,��ô�����Զ���һ������,����ֻ����һ��Ԫ�ز��� [4] ,�������������� True ��ʾ��ǰԪ��ƥ�䲢�ұ��ҵ�,��������򷴻� False
���淽��У���˵�ǰԪ��,������� class ����ȴ������ id ����,��ô������ True:
def has_class_but_no_id(tag):
    return tag.has_attr('class') and not tag.has_attr('id')

for tag in soup.find_all(has_class_but_no_id):
    print tag.name
# p
# p
# p

2��keyword ����
ע�⣺���һ��ָ�����ֵĲ��������������õĲ�����,����ʱ��Ѹò�������ָ������tag������������,�������һ������Ϊ id �Ĳ���,Beautiful Soup������ÿ��tag�ġ�id������
for tag in soup.find_all(id='link2'):
    print tag
# <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>

������� href ����,Beautiful Soup������ÿ��tag�ġ�href������
for tag in soup.find_all(href=re.compile('elsie')):
    print tag
# <a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>

ʹ�ö��ָ�����ֵĲ�������ͬʱ����tag�Ķ������
for tag in soup.find_all(href=re.compile('elsie'), id='link1'):
    print tag
# <a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>

�������������� class ���ˣ����� class �� python �Ĺؼ��ʣ�����ô�죿�Ӹ��»��߾Ϳ���
for tag in soup.find_all("a", class_="sister"): #ͬsoup.find_all("a", attrs={"class": "sister"})
    print tag
# <a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>
# <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>
# <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>

��Щtag��������������ʹ��,����HTML5�е� data-* ����
���ǿ���ͨ�� find_all() ������ attrs ��������һ���ֵ���������������������Ե�tag
for tag in soup.find_all("a", attrs={"class": "sister"}):
    print tag
# <a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>
# <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>
# <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>

3��text ����
ͨ�� text �������������ĵ��е��ַ�������.�� name �����Ŀ�ѡֵһ��, text �������� �ַ��� , ������ʽ , �б�, True
print soup.find_all(text="Elsie") #�ڱ��أ�û������������ΪElsie��ע���ַ���
# [u'Elsie']
print soup.find_all(text=["Tillie", "Elsie", "Lacie"])
# [u'Elsie', u'Lacie', u'Tillie']
print soup.find_all(text=re.compile("Dormouse"))
# [u"The Dormouse's story", u"The Dormouse's story"]

4��limit ����
find_all() ��������ȫ���������ṹ,����ĵ����ܴ���ô���������.������ǲ���Ҫȫ�����,����ʹ�� limit �������Ʒ��ؽ��������.Ч����SQL�е�limit�ؼ�������,���������Ľ�������ﵽ limit ������ʱ,��ֹͣ�������ؽ��.
�ĵ�������3��tag������������,�����ֻ������2��,��Ϊ���������˷�������
soup.find_all("a", limit=2)
[<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>, <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>]

5��recursive ����
����tag�� find_all() ����ʱ,Beautiful Soup�������ǰtag����������ڵ�,���ֻ������tag��ֱ���ӽڵ�,����ʹ�ò��� recursive=False .
soup.html.find_all("title")
# [<title>The Dormouse's story</title>]
 
soup.html.find_all("title", recursive=False)
# []

��2��find( name , attrs , recursive , text , **kwargs )
���� find_all() ����Ψһ�������� find_all() �����ķ��ؽ����ֵ����һ��Ԫ�ص��б�,�� find() ����ֱ�ӷ��ؽ��
print soup.html.find("a", recursive=True)
# <a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>
print soup.html.find("a", recursive=False)
# None

��3��find_parents() find_parent()
find_all() �� find() ֻ������ǰ�ڵ�������ӽڵ�,���ӽڵ��. find_parents() �� find_parent() ����������ǰ�ڵ�ĸ����ڵ�,������������ͨtag������������ͬ,�����ĵ������ĵ�����������
# coding=utf-8
from bs4 import BeautifulSoup

html = """
<html><head><title>The Dormouse's story</title></head>
<body class="body">
<p class="title" name="dromouse"><b>The Dormouse's story</b></p>
<p class="story">Once upon a time there were three little sisters; and their names were
<a href="http://example.com/elsie" class="sister" id="link1"><!-- Elsie --></a>,
<a href="http://example.com/lacie" class="sister" id="link2">Lacie</a> and
<a href="http://example.com/tillie" class="sister" id="link3">Tillie</a>;
and they lived at the bottom of a well.</p>
<p class="story">...</p>
"""


def has_class(tag):
    return tag.has_attr('class')


if __name__ == "__main__":
    soup = BeautifulSoup(html, "lxml")
    print soup.a.find_parents(has_class)
    print soup.a.find_parent(has_class)
�����
[<p class="story">Once upon a time there were three little sisters; and their names were\n<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>,\n<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a> and\n<a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>;\nand they lived at the bottom of a well.</p>, <body class="body">\n<p class="title" name="dromouse"><b>The Dormouse's story</b></p>\n<p class="story">Once upon a time there were three little sisters; and their names were\n<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>,\n<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a> and\n<a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>;\nand they lived at the bottom of a well.</p>\n<p class="story">...</p>\n</body>]
<p class="story">Once upon a time there were three little sisters; and their names were
<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>,
<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a> and
<a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>;
and they lived at the bottom of a well.</p>

4��find_next_siblings() find_next_sibling()
��2������ͨ�� .next_siblings ���ԶԵ� tag �����к���������ֵ� tag �ڵ���е���, find_next_siblings() �����������з��������ĺ�����ֵܽڵ�,find_next_sibling() ֻ���ط��������ĺ���ĵ�һ��tag�ڵ�
print soup.a.find_next_siblings(has_class)
# [<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>, <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>]
print soup.a.find_next_sibling(has_class)
# <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>

��5��find_previous_siblings() find_previous_sibling()
��2������ͨ�� .previous_siblings ���ԶԵ�ǰ tag ��ǰ��������ֵ� tag �ڵ���е���, find_previous_siblings() �����������з���������ǰ����ֵܽڵ�, find_previous_sibling() �������ص�һ������������ǰ����ֵܽڵ�

��6��find_all_next() find_next()
��2������ͨ�� .next_elements ���ԶԵ�ǰ tag ��֮��� tag ���ַ������е���, find_all_next() �����������з��������Ľڵ�, find_next() �������ص�һ�����������Ľڵ�
print soup.a.find_all_next(has_class)
# [<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>, <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>, <p class="story">...</p>]
print soup.a.find_next(has_class)
# <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>

��7��find_all_previous() �� find_previous()
��2������ͨ�� .previous_elements ���ԶԵ�ǰ�ڵ�ǰ��� tag ���ַ������е���, find_all_previous() �����������з��������Ľڵ�, find_previous()�������ص�һ�����������Ľڵ�

8.CSSѡ����
������д CSS ʱ����ǩ�������κ����Σ�����ǰ�ӵ㣬id��ǰ�� #������������Ҳ�����������Ƶķ�����ɸѡԪ�أ��õ��ķ����� soup.select()������������ list
��1��ͨ����ǩ������
print soup.select("title")
# [<title>The Dormouse's story</title>]
print soup.select("a")
# [<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>, <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>, <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>]
print soup.select("b")
# [<b>The Dormouse's story</b>]

��2��ͨ����������
print soup.select(".sister")
[<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>, <a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>, <a class="sister" href="http://example.com/tillie" id="link3">Tillie</a>]

3��ͨ�� id ������
print soup.select("#link1")
#[<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>]

��4����ϲ���
��ϲ��Ҽ���д class �ļ�ʱ����ǩ����������id�����е����ԭ����һ���ģ�������� p ��ǩ�У�id ���� link1�����ݣ�������Ҫ�ÿո�ֿ�
print soup.select('p #link2')
#[<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>]
ֱ���ӱ�ǩ����
print soup.select('html > head')
# [<head><title>The Dormouse's story</title></head>]
print soup.select('html > title')
# []
���Բ����ӱ�ǩ
print soup.select("html title")
# [<title>The Dormouse's story</title>]

��5�����Բ���
����ʱ�����Լ�������Ԫ�أ�������Ҫ����������������ע�����Ժͱ�ǩ����ͬһ�ڵ㣬�����м䲻�ܼӿո񣬷�����޷�ƥ�䵽��
print soup.select('a[id="link1"]')
#[<a class="sister" href="http://example.com/elsie" id="link1"><!-- Elsie --></a>]
ͬ����������Ȼ�������������ҷ�ʽ��ϣ�����ͬһ�ڵ�Ŀո������ͬһ�ڵ�Ĳ��ӿո�
print soup.select('p a[id="link2"]') #soup.select('body a[id="link2"]')
#[<a class="sister" href="http://example.com/lacie" id="link2">Lacie</a>]




#-----------------------------------------------------------------------------------------
#python str��repr������ http://www.cnpythoner.com/post/251.html
����str(),repr()��``���������Ժ͹��ܷ��涼�ǳ����ƣ���ʵ��repr()��``��������ȫһ�������飬���Ƿ��ص���һ������ġ��ٷ����ַ�����ʾ��Ҳ����˵�����������¿���ͨ����ֵ���㣨ʹ���ڽ�����eval()�����µõ��ö���

��str()��������ͬ��str()����������һ������Ŀɶ��Ժõ��ַ�����ʾ�����ķ��ؽ��ͨ���޷�����eval()��ֵ�������ʺ�����print����������Ҫ�ٴ����ѵ��ǣ�����������repr()���ص��ַ������ܹ��� eval()�ڽ������õ�ԭ���Ķ��� Ҳ����˵ repr() ����� Python�Ƚ��Ѻã���str()��������û��Ƚ��Ѻá�

��Ȼ��ˣ��ܶ�����������ߵ������Ȼ������ȫһ���ġ� ��ҿ��Կ�������Ĵ��룬�����жԱ�

>>> s = 'Hello, world.'
>>> str(s)
'Hello, world.'
>>> repr(s)
"'Hello, world.'"
>>> str(0.1)
'0.1'
>>> repr(0.1)
'0.10000000000000001'
>>> x = 10 * 3.25
>>> y = 200 * 200
>>> s = 'The value of x is ' + repr(x) + ', and y is ' + repr(y) + '...'
>>> print s
The value of x is 32.5, and y is 40000...
>>> # The repr() of a string adds string quotes and backslashes:
... hello = 'hello, world\n'
>>> hellos = repr(hello)
>>> print hellos
'hello, world\n'
>>> # The argument to repr() may be any Python object:
... repr((x, y, ('spam', 'eggs')))
"(32.5, 40000, ('spam', 'eggs'))"


#--------------------BeautifulSoup ��ȡĳ��tag��ǩ���������
# �ο�http://blog.csdn.net/willib/article/details/52246086
# ��ȡ������վ��IP�Ͷ˿�
# coding=utf-8
import re
import urllib2

from bs4 import BeautifulSoup

if __name__ == "__main__":
    user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5)'
    headers = {'User-Agent': user_agent}
    request = urllib2.Request("http://www.xicidaili.com/nn/1", headers=headers)
    response = urllib2.urlopen(request)
    page = response.read()

    soup = BeautifulSoup(page, "lxml")
    trTags = soup.find_all("tr", attrs={"class": re.compile('odd|')}) #soup.find_all("tr", class_=re.compile('odd|'))
    for tr in trTags:
        tdTags = tr.find_all("td")
        print tdTags[1].string
        print tdTags[2].string
�����
222.136.171.248
8118
171.211.210.126
808
223.199.171.39
8118
39.86.42.244
8118
202.120.29.28
...

#-----------------------------------------------------------------------------------------
# With ��䣬�Զ��ͷ���Դ
#�ο�https://www.ibm.com/developerworks/cn/opensource/os-cn-pythonwith/
#-----------------------------------------------------------------------------------------
#���ͼ��
def my_abs(x):
    if not isinstance(x, (int, float)):
        raise TypeError('bad operand type')
    if x >= 0:
        return x
    else:
        return -x
����˲�����������������Ĳ������ͣ������Ϳ����׳�һ������

>>> my_abs('A')
Traceback (most recent call last):
  File "<stdin>", line 1, in <module>
  File "<stdin>", line 3, in my_abs
TypeError: bad operand type
#-----------------------------------------------------------------------------------------
#Ĭ�ϲ���Ҫָ�򲻱���� https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431752945034eb82ac80a3e64b9bb4929b16eeed1eb9000
�ȶ���һ������������һ��list�����һ��END�ٷ��أ�
def add_end(L=[]):
    L.append('END')
    return L
������������ʱ������ƺ�����

>>> add_end([1, 2, 3])
[1, 2, 3, 'END']
>>> add_end(['x', 'y', 'z'])
['x', 'y', 'z', 'END']
����ʹ��Ĭ�ϲ�������ʱ��һ��ʼ���Ҳ�ǶԵģ�

>>> add_end()
['END']
���ǣ��ٴε���add_end()ʱ������Ͳ����ˣ�

>>> add_end()
['END', 'END']
>>> add_end()
['END', 'END', 'END']
�ܶ��ѧ�ߺ��ɻ�Ĭ�ϲ�����[]�����Ǻ����ƺ�ÿ�ζ�����ס�ˡ��ϴ������'END'���list��

ԭ��������£�

Python�����ڶ����ʱ��Ĭ�ϲ���L��ֵ�ͱ���������ˣ���[]����ΪĬ�ϲ���LҲ��һ����������ָ�����[]��ÿ�ε��øú���������ı���L�����ݣ����´ε���ʱ��Ĭ�ϲ��������ݾͱ��ˣ������Ǻ�������ʱ��[]�ˡ�

 ����Ĭ�ϲ���Ҫ�μ�һ�㣺Ĭ�ϲ�������ָ�򲻱����

def add_end(L=None):
    if L is None:
        L = []
    L.append('END')
    return L
���ڣ����۵��ö��ٴΣ������������⣺

>>> add_end()
['END']
>>> add_end()
['END']

#-----------------------------------------------------------------------------------------
#Ҫ����봫��ؼ��֣�ʹ��*
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431752945034eb82ac80a3e64b9bb4929b16eeed1eb9000
def person(name, age, *, city, job):
    print(name, age, city, job)

def person(name, age, *args, city, job): #���˿ɱ����֮�󣬾Ͳ���Ҫһ��������*�ָ���
    print(name, age, args, city, job)
    
if __name__ == "__main__":
    person("dingben", 28, city="sh", job="Engineer")

���ԣ��������⺯����������ͨ������func(*args, **kw)����ʽ���������������Ĳ�������ζ���ġ�
#-----------------------------------------------------------------------------------------
#Slice 
https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431756919644a792ee4ead724ef7afab3f7f771b04f5000
ǰ10������ÿ����ȡһ����
>>> L[:10:2]
[0, 2, 4, 6, 8]

��������ÿ5��ȡһ����
>>> L[::5]
[0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95]

����ʲô����д��ֻд[:]�Ϳ���ԭ������һ��list��
>>> L[:]
[0, 1, 2, 3, ..., 99]

�ַ���'xxx'Ҳ���Կ�����һ��list��ÿ��Ԫ�ؾ���һ���ַ�����ˣ��ַ���Ҳ��������Ƭ������ֻ�ǲ�����������ַ�����
>>> 'ABCDEFG'[:3]
'ABC'
>>> 'ABCDEFG'[::2]
'ACEG'
#-----------------------------------------------------------------------------------------
#����
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014317793224211f408912d9c04f2eac4d2af0d5d3d7b2000
���ԣ�������ʹ��forѭ��ʱ��ֻҪ������һ���ɵ�������forѭ���Ϳ����������У������ǲ�̫���ĸö��󾿾���list���������������͡�

��ô������ж�һ�������ǿɵ��������أ�������ͨ��collectionsģ���Iterable�����жϣ�

>>> from collections import Iterable
>>> isinstance('abc', Iterable) # str�Ƿ�ɵ���
True
>>> isinstance([1,2,3], Iterable) # list�Ƿ�ɵ���
True
>>> isinstance(123, Iterable) # �����Ƿ�ɵ���
False

���һ��С���⣬���Ҫ��listʵ������Java�������±�ѭ����ô�죿Python���õ�enumerate�������԰�һ��list�������-Ԫ�ضԣ������Ϳ�����forѭ����ͬʱ����������Ԫ�ر���
>>> for i, value in enumerate(['A', 'B', 'C']):
...     print(i, value)
...
0 A
1 B
2 C
#-----------------------------------------------------------------------------------------
#�б�������
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431779637539089fd627094a43a8a7c77e6102e3a811000
����ѭ��̫���������б�����ʽ�������һ��������ѭ�����������list��

>>> [x * x for x in range(1, 11)]
[1, 4, 9, 16, 25, 36, 49, 64, 81, 100]
д�б�����ʽʱ����Ҫ���ɵ�Ԫ��x * x�ŵ�ǰ�棬�����forѭ�����Ϳ��԰�list����������ʮ�����ã���д���Σ��ܿ�Ϳ�����Ϥ�����﷨��

forѭ�����滹���Լ���if�жϣ��������ǾͿ���ɸѡ����ż����ƽ����

>>> [x * x for x in range(1, 11) if x % 2 == 0]
[4, 16, 36, 64, 100]
������ʹ������ѭ������������ȫ���У�

>>> [m + n for m in 'ABC' for n in 'XYZ']
['AX', 'AY', 'AZ', 'BX', 'BY', 'BZ', 'CX', 'CY', 'CZ']
#-----------------------------------------------------------------------------------------
#������
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014317799226173f45ce40636141b6abc8424e12b5fb27000
ͨ���б�����ʽ�����ǿ���ֱ�Ӵ���һ���б����ǣ��ܵ��ڴ����ƣ��б������϶������޵ġ����ң�����һ������100���Ԫ�ص��б�����ռ�úܴ�Ĵ洢�ռ䣬������ǽ�����Ҫ����ǰ�漸��Ԫ�أ��Ǻ���������Ԫ��ռ�õĿռ䶼�װ��˷��ˡ�

���ԣ�����б�Ԫ�ؿ��԰���ĳ���㷨����������������Ƿ������ѭ���Ĺ����в��������������Ԫ���أ������Ͳ��ش���������list���Ӷ���ʡ�����Ŀռ䡣��Python�У�����һ��ѭ��һ�߼���Ļ��ƣ���Ϊ��������generator��

Ҫ����һ��generator���кܶ��ַ�������һ�ַ����ܼ򵥣�ֻҪ��һ���б�����ʽ��[]�ĳ�()���ʹ�����һ��generator��

>>> L = [x * x for x in range(10)]
>>> L
[0, 1, 4, 9, 16, 25, 36, 49, 64, 81]
>>> g = (x * x for x in range(10))
>>> g
<generator object <genexpr> at 0x1022ef630>
����L��g�����������������[]��()��L��һ��list����g��һ��generator��

���ǿ���ֱ�Ӵ�ӡ��list��ÿһ��Ԫ�أ���������ô��ӡ��generator��ÿһ��Ԫ���أ�

���Ҫһ��һ����ӡ����������ͨ��next()�������generator����һ������ֵ��

>>> next(g)
0
>>> next(g)
1
>>> next(g)
4
>>> next(g)
9
>>> next(g)
16
>>> next(g)
25
>>> next(g)
36
>>> next(g)
49
>>> next(g)
64
>>> next(g)
81
>>> next(g)
Traceback (most recent call last):
  File "<stdin>", line 1, in <module>
StopIteration
���ǽ�����generator��������㷨��ÿ�ε���next(g)���ͼ����g����һ��Ԫ�ص�ֵ��ֱ�����㵽���һ��Ԫ�أ�û�и����Ԫ��ʱ���׳�StopIteration�Ĵ���

��Ȼ���������ֲ��ϵ���next(g)ʵ����̫��̬�ˣ���ȷ�ķ�����ʹ��forѭ������ΪgeneratorҲ�ǿɵ�������

>>> g = (x * x for x in range(10))
>>> for n in g:
...     print(n)
... 
0
1
4
9
16
25
36
49
64
81


Ҳ����˵������ĺ�����generator��һ��֮ң��Ҫ��fib�������generator��ֻ��Ҫ��print(b)��Ϊyield b�Ϳ����ˣ�

def fib(max):
    n, a, b = 0, 0, 1
    while n < max:
        yield b
        a, b = b, a + b
        n = n + 1
    return 'done'
    
ͬ���ģ��Ѻ����ĳ�generator�����ǻ����ϴ���������next()����ȡ��һ������ֵ������ֱ��ʹ��forѭ����������

>>> for n in fib(6):
...     print(n)
...
1
1
2
3
5
8
������forѭ������generatorʱ�������ò���generator��return���ķ���ֵ�������Ҫ�õ�����ֵ�����벶��StopIteration���󣬷���ֵ������StopIteration��value�У�

>>> g = fib(6)
>>> while True:
...     try:
...         x = next(g)
...         print('g:', x)
...     except StopIteration as e:
...         print('Generator return value:', e.value)
...         break
...
g: 1
g: 1
g: 2
g: 3
g: 5
g: 8
Generator return value: done


������Ƕ������£�

          1
         / \
        1   1
       / \ / \
      1   2   1
     / \ / \ / \
    1   3   3   1
   / \ / \ / \ / \
  1   4   6   4   1
 / \ / \ / \ / \ / \
1   5   10  10  5   1
��ÿһ�п���һ��list����дһ��generator�����������һ�е�list��
# -*- coding: utf-8 -*-
def triangles():
    l = [1]
    yield l
    for rowIndex in range(1, 10):
        l = [l[i] + l[i + 1] for i in range(len(l) - 1)]
        l.insert(0, 1)
        l.append(1)
        yield l

if __name__ == "__main__":
    for t in triangles():
        print(t)
�����
[1],
[1, 1],
[1, 2, 1],
[1, 3, 3, 1],
[1, 4, 6, 4, 1],
[1, 5, 10, 10, 5, 1],
[1, 6, 15, 20, 15, 6, 1],
[1, 7, 21, 35, 35, 21, 7, 1],
[1, 8, 28, 56, 70, 56, 28, 8, 1],
[1, 9, 36, 84, 126, 126, 84, 36, 9, 1]
#-----------------------------------------------------------------------------------------
#������ https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143178254193589df9c612d2449618ea460e7a672a366000
if __name__ == "__main__":
    print(isinstance([], Iterable))
    print(isinstance({}, Iterable))
    print(isinstance("abc", Iterable))
    print(isinstance((x for x in range(10)), Iterable))
    print(isinstance(100, Iterable))
�����
True
True
True
True
False

����������������������forѭ���������Ա�next()�������ϵ��ò�������һ��ֵ��ֱ������׳�StopIteration�����ʾ�޷�����������һ��ֵ�ˡ�
���Ա�next()�������ò����Ϸ�����һ��ֵ�Ķ����Ϊ��������Iterator��
����ʹ��isinstance()�ж�һ�������Ƿ���Iterator����
if __name__ == "__main__":
    print(isinstance([], Iterator))
    print(isinstance({}, Iterator))
    print(isinstance("abc", Iterator))
    print(isinstance((x for x in range(10)), Iterator))
    print(isinstance(100, Iterator))
�����
False
False
False
True
False

����������Iterator���󣬵�list��dict��str��Ȼ��Iterable��ȴ����Iterator��
��list��dict��str��Iterable���Iterator����ʹ��iter()������
if __name__ == "__main__":
    print(isinstance(iter([]), Iterator))
    print(isinstance(iter({}), Iterator))
    print(isinstance(iter("abc"), Iterator))
    print(isinstance((x for x in range(10)), Iterator))
�����
True
True
True
True

������ΪPython��Iterator�����ʾ����һ����������Iterator������Ա�next()�������ò����Ϸ�����һ�����ݣ�ֱ��û������ʱ�׳�StopIteration���󡣿��԰����������������һ���������У�������ȴ������ǰ֪�����еĳ��ȣ�ֻ�ܲ���ͨ��next()����ʵ�ְ��������һ�����ݣ�����Iterator�ļ����Ƕ��Եģ�ֻ������Ҫ������һ������ʱ���Ż���㡣

Iterator�������Ա�ʾһ�����޴��������������ȫ����Ȼ������ʹ��list����Զ�����ܴ洢ȫ����Ȼ���ġ�

if __name__ == "__main__":
    l = list(range(10))
    iterL = iter(l)
    try:
        while True:
            print(next(iterL))
    except StopIteration as e:
        print(e.value)
�����
0
1
2
3
4
5
6
7
8
9
None

С��

���ǿ�������forѭ���Ķ�����Iterable���ͣ�

���ǿ�������next()�����Ķ�����Iterator���ͣ����Ǳ�ʾһ�����Լ�������У�

��������������list��dict��str����Iterable������Iterator����������ͨ��iter()�������һ��Iterator����

Python��forѭ�������Ͼ���ͨ�����ϵ���next()����ʵ�ֵģ����磺

for x in [1, 2, 3, 4, 5]:
    pass
ʵ������ȫ�ȼ��ڣ�

# ���Ȼ��Iterator����:
it = iter([1, 2, 3, 4, 5])
# ѭ��:
while True:
    try:
        # �����һ��ֵ:
        x = next(it)
    except StopIteration:
        # ����StopIteration���˳�ѭ��
        break
        
#range�ɵ��������ǲ��ǵ���������list�ɱ�Ϊlist����
print(isinstance(range(10), Iterable)) #True
print(isinstance(range(10), Iterator)) #False
t = iter(range(10))
print(isinstance(t, Iterable)) #True
print(isinstance(t, Iterator)) #True

#�������п�����list����list
g = (x * x for x in range(10))
print(g, list(g))
<generator object <genexpr> at 0x000002657F96E360> [0, 1, 4, 9, 16, 25, 36, 49, 64, 81]
#-----------------------------------------------------------------------------------------
#����ʽ��� https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014317848428125ae6aa24068b4c50a7e71501ab275d52000
����ʽ��̾���һ�ֳ���̶Ⱥܸߵı�̷�ʽ������ĺ���ʽ������Ա�д�ĺ���û�б�������ˣ�����һ��������ֻҪ������ȷ���ģ��������ȷ���ģ����ִ��������ǳ�֮Ϊû�и����á�������ʹ�ñ����ĳ���������ԣ����ں����ڲ��ı���״̬��ȷ����ͬ�������룬���ܵõ���ͬ���������ˣ����ֺ������и����õġ�

����ʽ��̵�һ���ص���ǣ�����Ѻ���������Ϊ����������һ����������������һ��������

Python�Ժ���ʽ����ṩ����֧�֡�����Python����ʹ�ñ�������ˣ�Python���Ǵ�����ʽ������ԡ�
#-----------------------------------------------------------------------------------------
# map/reduce
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014317852443934a86aa5bb5ea47fbbd5f35282b331335000
# -*- coding: utf-8 -*-
def f(x):
    return x * x

if __name__ == "__main__":
    r = map(f, list(range(10)))
    print(list(r))
    print(r)
    print(isinstance(r, Iterator))
    print(isinstance(r, Iterable))
�����
[0, 1, 4, 9, 16, 25, 36, 49, 64, 81]
<map object at 0x000002BB2F31A7F0>
True
True

map()����ĵ�һ��������f�����������������ڽ��r��һ��Iterator��Iterator�Ƕ������У����ͨ��list()�����������������ж��������������һ��list��
#����������python3��python2��һ����python3 map�õ�����һ��map����python2��ȡ������һ��list����


�ٿ�reduce���÷���reduce��һ������������һ������[x1, x2, x3, ...]�ϣ�������������������������reduce�ѽ�����������е���һ��Ԫ�����ۻ����㣬��Ч�����ǣ�
reduce(f, [x1, x2, x3, x4]) = f(f(f(x1, x2), x3), x4)

# -*- coding: utf-8 -*-
from functools import reduce

if __name__ == "__main__":
    r = reduce(lambda x, y : x + y, range(101))
    print(r) #5050 = sum(range(101))

���Ҫ������[1, 3, 5, 7, 9]�任������13579��reduce�Ϳ��������ó���
r = reduce(lambda x, y: x * 10 + y, range(1, 10, 2)) #13579

#ͬ��
def fn(x, y):
    return x * 10 + y
if __name__ == "__main__":
    r = reduce(fn, [1, 3, 5, 7, 9])
    print(r)
    
������ӱ���û����ô������ǣ�������ǵ��ַ���strҲ��һ�����У�������������ԼӸĶ������map()�����ǾͿ���д����strת��Ϊint�ĺ�����
from functools import reduce

def fn(x, y):
    return x * 10 + y

def char2num(s):
    digits = {'0': 0, '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9}
    return digits[s]

if __name__ == "__main__":
    r = reduce(fn, map(char2num, '13579'))
    print(r) #13579
    
    
�����һ��str2int�ĺ������ǣ�
from functools import reduce

DIGITS = {'0': 0, '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9}

def str2int(s):
    def fn(x, y):
        return x * 10 + y
    def char2num(s):
        return DIGITS[s]
    return reduce(fn, map(char2num, s))

#��lambda���ʽ��
# -*- coding: utf-8 -*-
from functools import reduce

DIGITS = {'0': 0, '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9}


def str2int(s):
    def char2num(s):
        return DIGITS[s]

    return reduce(lambda x, y: x * 10 + y, map(char2num, s))


if __name__ == "__main__":
    print(str2int('1357913579135791357913579135791357913579135791357913579135791357913579135791357913579'))
�����
1357913579135791357913579135791357913579135791357913579135791357913579135791357913579
#-----------------------------------------------------------------------------------------
#filter()
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431821084171d2e0f22e7cc24305ae03aa0214d0ef29000
�ɼ���filter()����߽׺������ؼ�������ȷʵ��һ����ɸѡ��������

ע�⵽filter()�������ص���һ��Iterator��Ҳ����һ���������У�����Ҫǿ��filter()��ɼ���������Ҫ��list()����������н��������list��

#����
# -*- coding: utf-8 -*-
def filterEven(number):
    return number % 2 == 0


if __name__ == "__main__":
    l = (x for x in range(10))
    oddList = filter(filterEven, l)#oddList��һ���������У��˴���û�е���filterEven����oddListֵ��������forѭ�����ٵ���filterEven����
    for x in oddList:
        print(x)


��filter������

����������һ�������ǰ���ɸ���������㷨��������ǳ��򵥣�

���ȣ��г���2��ʼ��������Ȼ��������һ�����У�

2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, ...

ȡ���еĵ�һ����2����һ����������Ȼ����2�����е�2�ı���ɸ����

3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, ...

ȡ�����еĵ�һ����3����һ����������Ȼ����3�����е�3�ı���ɸ����

5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, ...

ȡ�����еĵ�һ����5��Ȼ����5�����е�5�ı���ɸ����

7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, ...

����ɸ��ȥ���Ϳ��Եõ����е�������

��Python��ʵ������㷨�������ȹ���һ����3��ʼ���������У�

def _odd_iter():
    n = 1
    while True:
        n = n + 2
        yield n
ע������һ����������������һ���������С�

Ȼ����һ��ɸѡ������

def _not_divisible(n):
    return lambda x: x % n > 0
��󣬶���һ�������������Ϸ�����һ��������

def primes():
    yield 2
    it = _odd_iter() # ��ʼ����
    while True:
        n = next(it) # �������еĵ�һ����
        yield n
        it = filter(_not_divisible(n), it) # ����������
����������ȷ��ص�һ������2��Ȼ������filter()���ϲ���ɸѡ����µ����С�

����primes()Ҳ��һ���������У����Ե���ʱ��Ҫ����һ���˳�ѭ����������

# ��ӡ1000���ڵ�����:
for n in primes():
    if n < 1000:
        print(n)
    else:
        break
ע�⵽Iterator�Ƕ��Լ�������У��������ǿ�����Python��ʾ��ȫ����Ȼ��������ȫ�����������������У�������ǳ���ࡣ

#ȫ������
# -*- coding: utf-8 -*-


def oddIter():
    n = 1
    while True:
        n += 2
        yield n


def notDivisible(n):
    return lambda x: x % n != 0


def primes():
    yield 2
    it = oddIter()
    while True:
        n = next(it)
        yield n
        it = filter(notDivisible(n), it)#�������е�����ʱ��û�м���ÿ�������Ƿ���Ա�����������������ȡnext��ʱ��Ž��м��㣬��������nΪ5ʱ��������Ƿ�5 % 3 != 0��nΪ7ʱ������� 7 % 3 != 0 && 7 % 5 != 0����nΪ9ʱ��������9 % 3 != 0����n�����������һ��ֵ11������Ҫ����11 % 3 != 0 && 11 % 5 != 0 && 11 % 7 != 0����������...


if __name__ == "__main__":
    for n in primes():
        if n < 1000:
            print(n)
        else:
            break

#�ж��Ƿ��ǻ������֣�
# -*- coding: utf-8 -*-
from math import floor


def is_palindrome(n):
    s = str(n)
    lenS = len(s)
    for i in range(0, floor(lenS / 2)):
        if s[i] != s[-1 - i]:
            return False
    return True


if __name__ == "__main__":
    output = filter(is_palindrome, range(1, 1000))
    print(list(output))
#-----------------------------------------------------------------------------------------
# sorted https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014318230588782cac105d0d8a40c6b450a232748dc854000
print(sorted(['bob', 'about', 'Zoo', 'Credit'], key=str.lower))
�����
['about', 'bob', 'Credit', 'Zoo']

Ҫ���з������򣬲��ظĶ�key���������Դ������������reverse=True��
sorted(['bob', 'about', 'Zoo', 'Credit'], key=str.lower, reverse=True)

# -*- coding: utf-8 -*-

def by_name(t):
    return t[0]


def by_score(t):
    return t[1]


if __name__ == "__main__":
    L = [('Bob', 75), ('Adam', 92), ('Bart', 66), ('Lisa', 88)]
    print(sorted(L, key=by_name))
    print(sorted(L, key=by_score))
�����
[('Adam', 92), ('Bart', 66), ('Bob', 75), ('Lisa', 88)]
[('Bart', 66), ('Bob', 75), ('Lisa', 88), ('Adam', 92)]
#-----------------------------------------------------------------------------------------
#���� https://www.cnblogs.com/pengwang57/p/7183468.html
Python list����sort()������������Ҳ������python���õ�ȫ��sorted()�������Կɵ������������������µ����С�

1���������

�򵥵����������Ƿǳ����׵ġ�ֻ��Ҫ����sorted()������������һ���µ�list���µ�list��Ԫ�ػ���С�������(__lt__)������

>>> sorted([5, 2, 3, 1, 4])
[1, 2, 3, 4, 5]
������Ҳ����ʹ��list.sort()���������򣬴�ʱlist�������޸ġ�ͨ���˷�������sorted()���㣬��������㲻��Ҫ����ԭ����list���˷���������Ч��

>>> a = [5, 2, 3, 1, 4]
>>> a.sort()
>>> a
[1, 2, 3, 4, 5]
������һ����ͬ����list.sort()��������������list�У��෴��sorted()���������еĿɵ������ж���Ч��

>>> sorted({1: 'D', 2: 'B', 3: 'B', 4: 'E', 5: 'A'})
[1, 2, 3, 4, 5]
����

2��key����/����

��python2.4��ʼ��list.sort()��sorted()����������key������ָ��һ���������˺�������ÿ��Ԫ�رȽ�ǰ�����á� ����ͨ��keyָ���ĺ����������ַ����Ĵ�Сд��

>>> sorted("This is a test string from Andrew".split(), key=str.lower)
['a', 'Andrew', 'from', 'is', 'string', 'test', 'This']
����key������ֵΪһ���������˺���ֻ��һ�������ҷ���һ��ֵ�������бȽϡ���������ǿ��ٵ���Ϊkeyָ���ĺ�����׼ȷ�ض�ÿ��Ԫ�ص��á�

 

���㷺��ʹ��������ø��Ӷ����ĳЩֵ���Ը��Ӷ���������������磺

>>> student_tuples = [
        ('john', 'A', 15),
        ('jane', 'B', 12),
        ('dave', 'B', 10),
]
>>> sorted(student_tuples, key=lambda student: student[2])   # sort by age
[('dave', 'B', 10), ('jane', 'B', 12), ('john', 'A', 15)]
����ͬ���ļ�����ӵ���������Եĸ��Ӷ���Ҳ���ã����磺

���ƴ���
>>> class Student:
        def __init__(self, name, grade, age):
                self.name = name
                self.grade = grade
                self.age = age
        def __repr__(self):
                return repr((self.name, self.grade, self.age))
>>> student_objects = [
        Student('john', 'A', 15),
        Student('jane', 'B', 12),
        Student('dave', 'B', 10),
]
>>> sorted(student_objects, key=lambda student: student.age)   # sort by age
[('dave', 'B', 10), ('jane', 'B', 12), ('john', 'A', 15)]
���ƴ���
3��Operator ģ�麯��

�����key������ʹ�÷ǳ��㷺�����python�ṩ��һЩ����ĺ�����ʹ�÷��ʷ����������׺Ϳ��١�operatorģ����itemgetter��attrgetter����2.6��ʼ��������methodcaller������ʹ����Щ����������Ĳ�������ø��Ӽ��Ϳ��٣�

>>> from operator import itemgetter, attrgetter
>>> sorted(student_tuples, key=itemgetter(2))
[('dave', 'B', 10), ('jane', 'B', 12), ('john', 'A', 15)]
>>> sorted(student_objects, key=attrgetter('age'))
[('dave', 'B', 10), ('jane', 'B', 12), ('john', 'A', 15)]
����operatorģ�黹����༶���������磬����grade��Ȼ������age������

>>> sorted(student_tuples, key=itemgetter(1,2))
[('john', 'A', 15), ('dave', 'B', 10), ('jane', 'B', 12)]
>>> sorted(student_objects, key=attrgetter('grade', 'age'))
[('john', 'A', 15), ('dave', 'B', 10), ('jane', 'B', 12)]


#-----------------------------------------------------------------------------------------
#���� #https://docs.python.org/3/howto/sorting.html
>>> def reverse_numeric(x, y):
...     return y - x
>>> sorted([5, 2, 4, 1, 3], cmp=reverse_numeric) 
[5, 4, 3, 2, 1]

When porting code from Python 2.x to 3.x, the situation can arise when you have the user supplying a comparison function and you need to convert that to a key function. The following wrapper makes that easy to do:

def cmp_to_key(mycmp):
    'Convert a cmp= function into a key= function'
    class K:
        def __init__(self, obj, *args):
            self.obj = obj
        def __lt__(self, other):
            return mycmp(self.obj, other.obj) < 0
        def __gt__(self, other):
            return mycmp(self.obj, other.obj) > 0
        def __eq__(self, other):
            return mycmp(self.obj, other.obj) == 0
        def __le__(self, other):
            return mycmp(self.obj, other.obj) <= 0
        def __ge__(self, other):
            return mycmp(self.obj, other.obj) >= 0
        def __ne__(self, other):
            return mycmp(self.obj, other.obj) != 0
    return K
To convert to a key function, just wrap the old comparison function:

>>>
>>> sorted([5, 2, 4, 1, 3], key=cmp_to_key(reverse_numeric))
[5, 4, 3, 2, 1]

In Python 3.2, the functools.cmp_to_key() function was added to the functools module in the standard library.
#-----------------------------------------------------------------------------------------
#���غ���
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431835236741e42daf5af6514f1a8917b8aaadff31bf000

������ʵ��һ���ɱ��������͡�ͨ������£���͵ĺ�������������ģ�

def calc_sum(*args):
    ax = 0
    for n in args:
        ax = ax + n
    return ax
���ǣ��������Ҫ������ͣ������ں���Ĵ����У�������Ҫ�ټ�����ô�죿���Բ�������͵Ľ�������Ƿ�����͵ĺ�����

def lazy_sum(*args):
    def sum():
        ax = 0
        for n in args:
            ax = ax + n
        return ax
    return sum
�����ǵ���lazy_sum()ʱ�����صĲ�������ͽ����������ͺ�����

>>> f = lazy_sum(1, 3, 5, 7, 9)
>>> f
<function lazy_sum.<locals>.sum at 0x101c6ed90>
���ú���fʱ��������������͵Ľ����

>>> f()
25

����������У������ں���lazy_sum���ֶ����˺���sum�����ң��ڲ�����sum���������ⲿ����lazy_sum�Ĳ����;ֲ���������lazy_sum���غ���sumʱ����ز����ͱ����������ڷ��صĺ����У����ֳ�Ϊ���հ���Closure�����ĳ���ṹӵ�м����������

����ע��һ�㣬�����ǵ���lazy_sum()ʱ��ÿ�ε��ö��᷵��һ���µĺ�������ʹ������ͬ�Ĳ�����

>>> f1 = lazy_sum(1, 3, 5, 7, 9)
>>> f2 = lazy_sum(1, 3, 5, 7, 9)
>>> f1==f2
False
f1()��f2()�ĵ��ý������Ӱ�졣


�հ�

ע�⵽���صĺ������䶨���ڲ������˾ֲ�����args�����ԣ���һ������������һ�����������ڲ��ľֲ����������º������ã����ԣ��հ��������򵥣�ʵ�������ɲ����ס�

��һ����Ҫע��������ǣ����صĺ�����û������ִ�У�����ֱ��������f()��ִ�С���������һ�����ӣ�

def count():
    fs = []
    for i in range(1, 4):
        def f():
             return i*i
        fs.append(f)
    return fs

f1, f2, f3 = count()
������������У�ÿ��ѭ������������һ���µĺ�����Ȼ�󣬰Ѵ�����3�������������ˡ�

�������Ϊ����f1()��f2()��f3()���Ӧ����1��4��9����ʵ�ʽ���ǣ�

>>> f1()
9
>>> f2()
9
>>> f3()
9
ȫ������9��ԭ������ڷ��صĺ��������˱���i��������������ִ�С��ȵ�3������������ʱ�����������õı���i�Ѿ������3��������ս��Ϊ9��

���رհ�ʱ�μ�һ�㣺���غ�����Ҫ�����κ�ѭ�����������ߺ����ᷢ���仯�ı�����
 
 ���һ��Ҫ����ѭ��������ô�죿�������ٴ���һ���������øú����Ĳ�����ѭ��������ǰ��ֵ�����۸�ѭ������������θ��ģ��Ѱ󶨵�����������ֵ���䣺

def count():
    def f(j):
        def g():
            return j*j
        return g
    fs = []
    for i in range(1, 4):
        fs.append(f(i)) # f(i)���̱�ִ�У����i�ĵ�ǰֵ������f()
    return fs
�ٿ��������

>>> f1, f2, f3 = count()
>>> f1()
1
>>> f2()
4
>>> f3()
9

���رհ�ʱ�μ�һ�㣺���غ�����Ҫ�����κ�ѭ�����������ߺ����ᷢ���仯�ı�����
���һ��Ҫ����ѭ��������ô�죿�������ٴ���һ���������øú����Ĳ�����ѭ��������ǰ��ֵ�����۸�ѭ������������θ��ģ��Ѱ󶨵�����������ֵ���䣺
def count():
    fs = []
    for i in range(1, 4):
        def g():
            return i * i
        fs.append(g)
    return fs

def count2():
    def f(j):
        def g():
            return j * j
        return g

    fs = []
    for i in range(1, 4):
        fs.append(f(i))
    return fs


if __name__ == "__main__":
    f1, f2, f3 = count()
    print(f1(), f2(), f3())
    f1, f2, f3 = count2()
    print(f1(), f2(), f3())


#���ñհ�����һ��������������ÿ�ε��������ص���������
Python 3.x������nonlocal�ؼ��֣��������ڱ�ʶ�ⲿ������ı�����
# -*- coding: utf-8 -*-

def createCounter():
    i = 0
    def counter():
        nonlocal i
        i += 1
        return i
    return counter

if __name__ == "__main__":
    counter = createCounter()
    print(counter(), counter(), counter())
#-----------------------------------------------------------------------------------------
#��������
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431843456408652233b88b424613aa8ec2fe032fd85a000
def build(x, y):
    return lambda: x * x + y * y

if __name__ == "__main__":
    f = build(3, 4)
    print(f())
#-----------------------------------------------------------------------------------------
#װ���� https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014318435599930270c0381a3b44db991cd6d858064ac0000
���ں���Ҳ��һ�����󣬶��Һ���������Ա���ֵ�����������ԣ�ͨ������Ҳ�ܵ��øú�����

>>> def now():
...     print('2015-3-25')
...
>>> f = now
>>> f()
2015-3-25
����������һ��__name__���ԣ������õ����������֣�

>>> now.__name__
'now'
>>> f.__name__
'now'

���ڣ���������Ҫ��ǿnow()�����Ĺ��ܣ����磬�ں�������ǰ���Զ���ӡ��־�����ֲ�ϣ���޸�now()�����Ķ��壬�����ڴ��������ڼ䶯̬���ӹ��ܵķ�ʽ����֮Ϊ��װ��������Decorator����

�����ϣ�decorator����һ�����غ����ĸ߽׺��������ԣ�����Ҫ����һ���ܴ�ӡ��־��decorator�����Զ������£�

def log(func):
    def wrapper(*args, **kw):
        print('call %s():' % func.__name__)
        return func(*args, **kw)
    return wrapper
�۲������log����Ϊ����һ��decorator�����Խ���һ��������Ϊ������������һ������������Ҫ����Python��@�﷨����decorator���ں����Ķ��崦��

@log
def now():
    print('2015-3-25')
����now()����������������now()������������������now()����ǰ��ӡһ����־��

>>> now()
call now():
2015-3-25
��@log�ŵ�now()�����Ķ��崦���൱��ִ������䣺

now = log(now)
����log()��һ��decorator������һ�����������ԣ�ԭ����now()������Ȼ���ڣ�ֻ������ͬ����now����ָ�����µĺ��������ǵ���now()��ִ���º���������log()�����з��ص�wrapper()������

wrapper()�����Ĳ���������(*args, **kw)����ˣ�wrapper()�������Խ�����������ĵ��á���wrapper()�����ڣ����ȴ�ӡ��־���ٽ����ŵ���ԭʼ������

# -*- coding: utf-8 -*-
def log(func):
    def wrapper(*args, **kw):
        print('call %s():' % func.__name__)
        return func(*args, **kw)

    return wrapper

@log
def now():
    print("abc")


if __name__ == "__main__":
    now()
    print(now.__name__)
�����
call now():
abc
wrapper

���decorator������Ҫ����������Ǿ���Ҫ��дһ������decorator�ĸ߽׺�����д����������ӡ����磬Ҫ�Զ���log���ı���
def log(text):
    def decorator(func):
        def wrapper(*args, **kw):
            print('%s %s():' % (text, func.__name__))
            return func(*args, **kw)
        return wrapper
    return decorator
���3��Ƕ�׵�decorator�÷����£�

@log('execute')
def now():
    print('2015-3-25')
ִ�н�����£�

>>> now()
execute now():
2015-3-25
������Ƕ�׵�decorator��ȣ�3��Ƕ�׵�Ч���������ģ�

>>> now = log('execute')(now)
�����������������䣬����ִ��log('execute')�����ص���decorator�������ٵ��÷��صĺ�����������now����������ֵ������wrapper������
��������decorator�Ķ��嶼û�����⣬���������һ������Ϊ���ǽ��˺���Ҳ�Ƕ�������__name__�����ԣ�����ȥ������decoratorװ��֮��ĺ��������ǵ�__name__�Ѿ���ԭ����'now'�����'wrapper'��
>>> now.__name__
'wrapper'

��Ϊ���ص��Ǹ�wrapper()�������־���'wrapper'�����ԣ���Ҫ��ԭʼ������__name__�����Ը��Ƶ�wrapper()�����У�������Щ��������ǩ���Ĵ���ִ�оͻ����
����Ҫ��дwrapper.__name__ = func.__name__�����Ĵ��룬Python���õ�functools.wraps���Ǹ�����µģ����ԣ�һ��������decorator��д�����£�

import functools

def log(func):
    @functools.wraps(func)
    def wrapper(*args, **kw):
        print('call %s():' % func.__name__)
        return func(*args, **kw)
    return wrapper
������Դ�������decorator��

import functools

def log(text):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kw):
            print('%s %s():' % (text, func.__name__))
            return func(*args, **kw)
        return wrapper
    return decorator
import functools�ǵ���functoolsģ�顣ģ��ĸ����Ժ򽲽⡣���ڣ�ֻ���ס�ڶ���wrapper()��ǰ�����@functools.wraps(func)���ɡ�
#-----------------------------------------------------------------------------------------
#ƫ���� https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143184474383175eeea92a8b0439fab7b392a8a32f8fa000
Python��functoolsģ���ṩ�˺ܶ����õĹ��ܣ�����һ������ƫ������Partial function����Ҫע�⣬�����ƫ��������ѧ�����ϵ�ƫ������һ����

�ڽ��ܺ���������ʱ�����ǽ�����ͨ���趨������Ĭ��ֵ�����Խ��ͺ������õ��Ѷȡ���ƫ����Ҳ����������һ�㡣�������£�

int()�������԰��ַ���ת��Ϊ���������������ַ���ʱ��int()����Ĭ�ϰ�ʮ����ת����

>>> int('12345')
12345
��int()�������ṩ�����base������Ĭ��ֵΪ10���������base�������Ϳ�����N���Ƶ�ת����

>>> int('12345', base=8)
5349
>>> int('12345', 16)
74565
����Ҫת�������Ķ������ַ�����ÿ�ζ�����int(x, base=2)�ǳ��鷳�����ǣ������뵽�����Զ���һ��int2()�ĺ�����Ĭ�ϰ�base=2����ȥ��

def int2(x, base=2):
    return int(x, base)
����������ת�������ƾͷǳ������ˣ�

>>> int2('1000000')
64
>>> int2('1010101')
85
functools.partial���ǰ������Ǵ���һ��ƫ�����ģ�����Ҫ�����Լ�����int2()������ֱ��ʹ������Ĵ��봴��һ���µĺ���int2��

>>> import functools
>>> int2 = functools.partial(int, base=2)
>>> int2('1000000')
64
>>> int2('1010101')
85
���ԣ����ܽ�functools.partial�����þ��ǣ���һ��������ĳЩ�������̶�ס��Ҳ��������Ĭ��ֵ��������һ���µĺ�������������º�������򵥡�

ע�⵽������µ�int2�����������ǰ�base���������趨Ĭ��ֵΪ2����Ҳ�����ں�������ʱ��������ֵ��

>>> int2('1000000', base=10)
1000000
��󣬴���ƫ����ʱ��ʵ���Ͽ��Խ��պ�������*args��**kw��3�������������룺

int2 = functools.partial(int, base=2)
ʵ���Ϲ̶���int()�����Ĺؼ��ֲ���base��Ҳ���ǣ�

int2('10010')
�൱�ڣ�

kw = { 'base': 2 }
int('10010', **kw)
�����룺

max2 = functools.partial(max, 10)
ʵ���ϻ��10��Ϊ*args��һ�����Զ��ӵ���ߣ�Ҳ���ǣ�

max2(5, 6, 7)
�൱�ڣ�

args = (10, 5, 6, 7)
max(*args)
���Ϊ10��

С��

�������Ĳ�������̫�࣬��Ҫ��ʱ��ʹ��functools.partial���Դ���һ���µĺ���������º������Թ̶�סԭ�����Ĳ��ֲ������Ӷ��ڵ���ʱ���򵥡�


# -*- coding: utf-8 -*-
import functools


def test(a=3, b=4):
    return a + b


if __name__ == "__main__":
    print(test())
    new_test = functools.partial(test, a=7, b=8)
    print(new_test(a=10, b=11))
    print(new_test(10, 11))#�൱��new_test(10, 11, a=7, b=8)
#-----------------------------------------------------------------------------------------
#ʹ��ģ�� https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431845183474e20ee7e7828b47f7b7607f2dc1e90dbb000
������

��һ��ģ���У����ǿ��ܻᶨ��ܶຯ���ͱ��������еĺ����ͱ�������ϣ��������ʹ�ã��еĺ����ͱ�������ϣ��������ģ���ڲ�ʹ�á���Python�У���ͨ��_ǰ׺��ʵ�ֵġ�

�����ĺ����ͱ������ǹ����ģ�public�������Ա�ֱ�����ã����磺abc��x123��PI�ȣ�

����__xxx__�����ı�����������������Ա�ֱ�����ã�������������;�����������__author__��__name__�������������helloģ�鶨����ĵ�ע��Ҳ�������������__doc__���ʣ������Լ��ı���һ�㲻Ҫ�����ֱ�������

����_xxx��__xxx�����ĺ�����������Ƿǹ����ģ�private������Ӧ�ñ�ֱ�����ã�����_abc��__abc�ȣ�

֮��������˵��private�����ͱ�������Ӧ�á���ֱ�����ã������ǡ����ܡ���ֱ�����ã�����ΪPython��û��һ�ַ���������ȫ���Ʒ���private��������������ǣ��ӱ��ϰ���ϲ�Ӧ������private�����������

private�����������Ӧ�ñ��������ã���������ʲô���أ��뿴���ӣ�

def _private_1(name):
    return 'Hello, %s' % name

def _private_2(name):
    return 'Hi, %s' % name

def greeting(name):
    if len(name) > 3:
        return _private_1(name)
    else:
        return _private_2(name)
������ģ���﹫��greeting()�����������ڲ��߼���private�������������ˣ�����������greeting()�������ù����ڲ���private����ϸ�ڣ���Ҳ��һ�ַǳ����õĴ����װ�ͳ���ķ���������

�ⲿ����Ҫ���õĺ���ȫ�������private��ֻ���ⲿ��Ҫ���õĺ����Ŷ���Ϊpublic��
#-----------------------------------------------------------------------------------------
#ģ������·��
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143186362353505516c5d4e38456fb225c18cc5b54ffb000
 
��������ͼ����һ��ģ��ʱ��Python����ָ����·����������Ӧ��.py�ļ�������Ҳ������ͻᱨ��

>>> import mymodule
Traceback (most recent call last):
  File "<stdin>", line 1, in <module>
ImportError: No module named mymodule
Ĭ������£�Python��������������ǰĿ¼�������Ѱ�װ������ģ��͵�����ģ�飬����·�������sysģ���path�����У�

>>> import sys
>>> sys.path
['', '/Library/Frameworks/Python.framework/Versions/3.6/lib/python36.zip', '/Library/Frameworks/Python.framework/Versions/3.6/lib/python3.6', ..., '/Library/Frameworks/Python.framework/Versions/3.6/lib/python3.6/site-packages']
�������Ҫ����Լ�������Ŀ¼�������ַ�����

һ��ֱ���޸�sys.path�����Ҫ������Ŀ¼��

>>> import sys
>>> sys.path.append('/Users/michael/my_py_scripts')
���ַ�����������ʱ�޸ģ����н�����ʧЧ��

�ڶ��ַ��������û�������PYTHONPATH���û������������ݻᱻ�Զ���ӵ�ģ������·���С����÷�ʽ������Path�����������ơ�ע��ֻ��Ҫ������Լ�������·����Python�Լ����������·������Ӱ�졣
#-----------------------------------------------------------------------------------------
#��������
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014318650247930b1b21d7d3c64fe38c4b5a80d4469ad7000

#!/usr/bin/env python3
# -*- coding: utf-8 -*-

class Student(object):
    def __init__(self, name, score):
        self.__name = name
        self.__score = score

    def print_score(self):
        print('%s: %s' % (self.__name, self.__score))


if __name__ == "__main__":
    ben = Student("Ben", 59)
    ben.print_score()
    #print(ben.name) #�޷�����
    print(ben._Student__name)#���Է���
����󣬶����ⲿ������˵��ûʲô�䶯�������Ѿ��޷����ⲿ����ʵ������.__name��ʵ������.__score�ˣ�

��Ҫע����ǣ���Python�У�����������__xxx__�ģ�Ҳ������˫�»��߿�ͷ��������˫�»��߽�β�ģ��������������������ǿ���ֱ�ӷ��ʵģ�����private���������ԣ�������__name__��__score__�����ı�������

��Щʱ����ῴ����һ���»��߿�ͷ��ʵ��������������_name��������ʵ�������ⲿ�ǿ��Է��ʵģ����ǣ�����Լ���׳ɵĹ涨�����㿴�������ı���ʱ����˼���ǣ�����Ȼ�ҿ��Ա����ʣ����ǣ��������Ϊ˽�б�������Ҫ������ʡ���

˫�»��߿�ͷ��ʵ�������ǲ���һ�����ܴ��ⲿ�����أ���ʵҲ���ǡ�����ֱ�ӷ���__name����ΪPython�����������__name�����ĳ���_Student__name�����ԣ���Ȼ����ͨ��_Student__name������__name������

>>> bart._Student__name
'Bart Simpson'
����ǿ�ҽ����㲻Ҫ��ô�ɣ���Ϊ��ͬ�汾��Python���������ܻ��__name�ĳɲ�ͬ�ı�������

#-----------------------------------------------------------------------------------------
#�̳кͶ�̬ https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431865288798deef438d865e4c2985acff7e9fad15e3000
��̬���� vs ��̬����

���ھ�̬���ԣ�����Java����˵�������Ҫ����Animal���ͣ�����Ķ��������Animal���ͻ����������࣬���򣬽��޷�����run()������

����Python�����Ķ�̬������˵����һ����Ҫ����Animal���͡�����ֻ��Ҫ��֤����Ķ�����һ��run()�����Ϳ����ˣ�

class Timer(object):
    def run(self):
        print('Start...')
����Ƕ�̬���Եġ�Ѽ�����͡���������Ҫ���ϸ�ļ̳���ϵ��һ������ֻҪ����������Ѽ�ӣ�����·����Ѽ�ӡ��������Ϳ��Ա�������Ѽ�ӡ�

Python�ġ�file-like object������һ��Ѽ�����͡����������ļ���������һ��read()���������������ݡ����ǣ�������ֻҪ��read()������������Ϊ��file-like object������ຯ�����յĲ������ǡ�file-like object�����㲻һ��Ҫ�����������ļ�������ȫ���Դ����κ�ʵ����read()�����Ķ���
#-----------------------------------------------------------------------------------------
#��ȡ������Ϣ
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431866385235335917b66049448ab14a499afd5b24db000
type()�������ص���ʲô�����أ������ض�Ӧ��Class���͡��������Ҫ��if������жϣ�����Ҫ�Ƚ�����������type�����Ƿ���ͬ��

>>> type(123)==type(456)
True
>>> type(123)==int
True
>>> type('abc')==type('123')
True
>>> type('abc')==str
True
>>> type('abc')==type(123)
False

�жϻ����������Ϳ���ֱ��дint��str�ȣ������Ҫ�ж�һ�������Ƿ��Ǻ�����ô�죿����ʹ��typesģ���ж���ĳ�����
>>> import types
>>> def fn():
...     pass
...
>>> type(fn)==types.FunctionType
True
>>> type(abs)==types.BuiltinFunctionType
True
>>> type(lambda x: x)==types.LambdaType
True
>>> type((x for x in range(10)))==types.GeneratorType
True

ʹ��isinstance()

����class�ļ̳й�ϵ��˵��ʹ��type()�ͺܲ����㡣����Ҫ�ж�class�����ͣ�����ʹ��isinstance()������

���ǻع��ϴε����ӣ�����̳й�ϵ�ǣ�

object -> Animal -> Dog -> Husky
��ô��isinstance()�Ϳ��Ը������ǣ�һ�������Ƿ���ĳ�����͡��ȴ���3�����͵Ķ���

>>> a = Animal()
>>> d = Dog()
>>> h = Husky()
Ȼ���жϣ�

>>> isinstance(h, Husky)
True
û�����⣬��Ϊh����ָ��ľ���Husky����

���жϣ�

>>> isinstance(h, Dog)
True
h��Ȼ������Husky���ͣ�������Husky�Ǵ�Dog�̳������ģ����ԣ�hҲ����Dog���͡����仰˵��isinstance()�жϵ���һ�������Ƿ��Ǹ����ͱ�������λ�ڸ����͵ĸ��̳����ϡ�

��ˣ����ǿ���ȷ�ţ�h����Animal���ͣ�

>>> isinstance(h, Animal)
True
ͬ��ʵ��������Dog��dҲ��Animal���ͣ�

>>> isinstance(d, Dog) and isinstance(d, Animal)
True
���ǣ�d����Husky���ͣ�

>>> isinstance(d, Husky)
False
����type()�жϵĻ�������Ҳ������isinstance()�жϣ�

>>> isinstance('a', str)
True
>>> isinstance(123, int)
True
>>> isinstance(b'a', bytes)
True
���һ������ж�һ�������Ƿ���ĳЩ�����е�һ�֣���������Ĵ���Ϳ����ж��Ƿ���list����tuple��

>>> isinstance([1, 2, 3], (list, tuple))
True
>>> isinstance((1, 2, 3), (list, tuple))
True

ʹ��dir()

���Ҫ���һ��������������Ժͷ���������ʹ��dir()������������һ�������ַ�����list�����磬���һ��str������������Ժͷ�����

>>> dir('ABC')
['__add__', '__class__',..., '__subclasshook__', 'capitalize', 'casefold',..., 'zfill']
����__xxx__�����Ժͷ�����Python�ж�����������;�ģ�����__len__�������س��ȡ���Python�У���������len()������ͼ��ȡһ������ĳ��ȣ�ʵ���ϣ���len()�����ڲ������Զ�ȥ���øö����__len__()���������ԣ�����Ĵ����ǵȼ۵ģ�

>>> len('ABC')
3
>>> 'ABC'.__len__()
3
�����Լ�д���࣬���Ҳ����len(myObj)�Ļ������Լ�дһ��__len__()������

>>> class MyDog(object):
...     def __len__(self):
...         return 100
...
>>> dog = MyDog()
>>> len(dog)
100


���������Ժͷ����г����ǲ����ģ����getattr()��setattr()�Լ�hasattr()�����ǿ���ֱ�Ӳ���һ�������״̬��
>>> class MyObject(object):
...     def __init__(self):
...         self.x = 9
...     def power(self):
...         return self.x * self.x
...
>>> obj = MyObject()
�����ţ����Բ��Ըö�������ԣ�

>>> hasattr(obj, 'x') # ������'x'��
True
>>> obj.x
9
>>> hasattr(obj, 'y') # ������'y'��
False
>>> setattr(obj, 'y', 19) # ����һ������'y'
>>> hasattr(obj, 'y') # ������'y'��
True
>>> getattr(obj, 'y') # ��ȡ����'y'
19
>>> obj.y # ��ȡ����'y'
19
�����ͼ��ȡ�����ڵ����ԣ����׳�AttributeError�Ĵ���

>>> getattr(obj, 'z') # ��ȡ����'z'
Traceback (most recent call last):
  File "<stdin>", line 1, in <module>
AttributeError: 'MyObject' object has no attribute 'z'

���Դ���һ��default������������Բ����ڣ��ͷ���Ĭ��ֵ��

>>> getattr(obj, 'z', 404) # ��ȡ����'z'����������ڣ�����Ĭ��ֵ404
404
Ҳ���Ի�ö���ķ�����

>>> hasattr(obj, 'power') # ������'power'��
True
>>> getattr(obj, 'power') # ��ȡ����'power'
<bound method MyObject.power of <__main__.MyObject object at 0x10077a6a0>>
>>> fn = getattr(obj, 'power') # ��ȡ����'power'����ֵ������fn
>>> fn # fnָ��obj.power
<bound method MyObject.power of <__main__.MyObject object at 0x10077a6a0>>
>>> fn() # ����fn()�����obj.power()��һ����
81


С��

ͨ�����õ�һϵ�к��������ǿ��Զ�����һ��Python��������������õ����ڲ������ݡ�Ҫע����ǣ�ֻ���ڲ�֪��������Ϣ��ʱ�����ǲŻ�ȥ��ȡ������Ϣ���������ֱ��д��

sum = obj.x + obj.y
�Ͳ�Ҫд��

sum = getattr(obj, 'x') + getattr(obj, 'y')
һ����ȷ���÷����������£�

def readImage(fp):
    if hasattr(fp, 'read'):
        return readData(fp)
    return None
��������ϣ�����ļ���fp�ж�ȡͼ����������Ҫ�жϸ�fp�����Ƿ����read������������ڣ���ö�����һ��������������ڣ����޷���ȡ��hasattr()���������ó���

��ע�⣬��Python���ද̬�����У�����Ѽ�����ͣ���read()�������������fp�������һ���ļ�������Ҳ��������������Ҳ�������ڴ��е�һ���ֽ�������ֻҪread()�������ص�����Ч��ͼ�����ݣ��Ͳ�Ӱ���ȡͼ��Ĺ��ܡ�
#-----------------------------------------------------------------------------------------
#����Python��������ʵ�����Ե�����
#http://python.jobbole.com/85100/
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

class Student(object):
    COUNT = 10
    def __init__(self, name, score):
        self.__name = name
        self.__score = score

    def print_score(self):
        print('%s: %s' % (self.__name, self.__score))


if __name__ == "__main__":
    s1 = Student("Ben", 99)
    s2 = Student("Mei", 80)
    print(s1.COUNT, s2.COUNT, Student.COUNT)
    s1.COUNT += 2
    print(s1.COUNT, s2.COUNT, Student.COUNT)
    Student.COUNT += 3
    print(s1.COUNT, s2.COUNT, Student.COUNT)
�����
10 10 10
12 10 10
12 13 13

Python�����Բ��һ���

Python�����ԵĻ�ȡ����һ�����ϲ��һ��ƣ������������������˵����
Python��һ�нԶ���AAA���������obj1����ʵ�����󣬴Ӷ���ĽǶ�������AAA��obj1�������޹صĶ��󣬵��ǣ�Pythonͨ������Ĳ����������������AAA��ʵ������obj1��obj2֮��Ĺ�ϵ��
��ͼ��ʾ
        AAA
         |
       -----
      |     |  
    obj1   obj2
    
��������1�е���obj1.aaaʱ��Python���մ�obj1��AAA��˳�����µ��ϲ�������aaa��
ֵ��ע�����ʱ��obj1��û������aaa�ģ����ǣ�Python����AAA��ȥ���ң��ɹ��ҵ�������ʾ���������ԣ���������������AAA������aaaȷʵ�ǹ����������ʵ���ģ���Ȼ����ֻ�ǴӲ���������ʽģ�������ϵ��

Python�е���������

ԭ���ӵ�����Ҳָ������Ĺؼ���������2��obj1.aaa += 2��
Ϊʲô�أ�
��������ָ��obj.aaa += 2���������Ի�ȡ����������������������obj1.aaa += 2�ȼ���obj1.aaa = obj1.aaa + 2��
���е�ʽ�Ҳ��obj.aaa�������Ի�ȡ��������ǰ��������ᵽ�Ĳ��ҹ�����У�������ʱ�򣬻�ȡ������AAA������aaa�����Ե�ʽ����ֵΪ12��
�ڶ����������������ã���obj.aaa = 12���������������õ�ʱ��obj1���ʵ������û������aaa����˻�Ϊ����̬���һ������aaa��
���ڴӶ���ĽǶȣ�������ʵ�������������������Ķ������ԣ����aaa����ֻ����obj1��Ҳ����˵����ʱ�������AAA��ʵ������aaa������һ������aaa��
��ô��������3�У��ٴε���obj1.aaaʱ���������Ե��ò��ҹ������ʱ���ȡ������ʵ������obj1������aaa�������������AAA������aaa��
#-----------------------------------------------------------------------------------------
#ʹ��__slots__
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143186739713011a09b63dcbd42cc87f907a778b3ac73000ʹ��__slots__

��������£������Ƕ�����һ��class��������һ��class��ʵ�������ǿ��Ը���ʵ�����κ����Ժͷ���������Ƕ�̬���Ե�����ԡ��ȶ���class��

class Student(object):
    pass
Ȼ�󣬳��Ը�ʵ����һ�����ԣ�

>>> s = Student()
>>> s.name = 'Michael' # ��̬��ʵ����һ������
>>> print(s.name)
Michael
�����Գ��Ը�ʵ����һ��������

>>> def set_age(self, age): # ����һ��������Ϊʵ������
...     self.age = age
...
>>> from types import MethodType
>>> s.set_age = MethodType(set_age, s) # ��ʵ����һ������
>>> s.set_age(25) # ����ʵ������
>>> s.age # ���Խ��
25
���ǣ���һ��ʵ���󶨵ķ���������һ��ʵ���ǲ������õģ�

>>> s2 = Student() # �����µ�ʵ��
>>> s2.set_age(25) # ���Ե��÷���
Traceback (most recent call last):
  File "<stdin>", line 1, in <module>
AttributeError: 'Student' object has no attribute 'set_age'
Ϊ�˸�����ʵ�����󶨷��������Ը�class�󶨷�����

>>> def set_score(self, score):
...     self.score = score
...
>>> Student.set_score = set_score
��class�󶨷���������ʵ�����ɵ��ã�

>>> s.set_score(100)
>>> s.score
100
>>> s2.set_score(99)
>>> s2.score
99
ͨ������£������set_score��������ֱ�Ӷ�����class�У�����̬�����������ڳ������еĹ����ж�̬��class���Ϲ��ܣ����ھ�̬�����к���ʵ�֡�

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from types import MethodType


class Student(object):
    def __init__(self, name, score):
        self.__name = name
        self.score = score

    def print_score(self):
        print('%s: %s' % (self.__name, self.score))

def set_age(self, age):
    self.age = age

def set_name(self, name):
    self.__name = name

def set_score(self, score):
    self.score = score

def print_age(self):
    # print('Name[%s]: score[%s] age[%s]' % (self.__name, self.__score, self.age))
    print('age[%s]' % self.age)


if __name__ == "__main__":
    s1 = Student("Ben", 99)
    s1.set_age = MethodType(set_age, s1)
    s1.set_name = MethodType(set_name, s1)
    s1.print_age = MethodType(print_age, s1)
    s1.set_score = MethodType(set_score, s1)
    s1.set_age(25)
    s1.set_name("Yan") #ԭ����__name��ΪΪ_Student__name����ǰ�Ǹ�s1�������°���һ������__name����δ�ı�ԭ����_Student__name
    print(s1.__dict__)
    print(s1.age)
    s1.print_age()
    s1.print_score()

    s1.set_score("89")
    s1.print_score()
�����
{'_Student__name': 'Ben', 'score': 99, 'set_age': <bound method set_age of <__main__.Student object at 0x0000025FE8E07898>>, 'set_name': <bound method set_name of <__main__.Student object at 0x0000025FE8E07898>>, 'print_age': <bound method print_age of <__main__.Student object at 0x0000025FE8E07898>>, 'set_score': <bound method set_score of <__main__.Student object at 0x0000025FE8E07898>>, 'age': 25, '__name': 'Yan'}
25
age[25]
Ben: 99
Ben: 89

ʹ��__slots__
���ǣ����������Ҫ����ʵ����������ô�죿���磬ֻ�����Studentʵ�����name��age���ԡ�
Ϊ�˴ﵽ���Ƶ�Ŀ�ģ�Python�����ڶ���class��ʱ�򣬶���һ�������__slots__�����������Ƹ�classʵ������ӵ����ԣ�
class Student(object):
    __slots__ = ('name', 'age') # ��tuple��������󶨵���������
Ȼ���������ԣ�

>>> s = Student() # �����µ�ʵ��
>>> s.name = 'Michael' # ������'name'
>>> s.age = 25 # ������'age'
>>> s.score = 99 # ������'score'
Traceback (most recent call last):
  File "<stdin>", line 1, in <module>
AttributeError: 'Student' object has no attribute 'score'
����'score'û�б��ŵ�__slots__�У����Բ��ܰ�score���ԣ���ͼ��score���õ�AttributeError�Ĵ���

ʹ��__slots__Ҫע�⣬__slots__��������Խ��Ե�ǰ��ʵ�������ã��Լ̳е������ǲ������õģ�

>>> class GraduateStudent(Student):
...     pass
...
>>> g = GraduateStudent()
>>> g.score = 9999
������������Ҳ����__slots__������������ʵ������������Ծ��������__slots__���ϸ����__slots__��
#-----------------------------------------------------------------------------------------
#ʹ��@property
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143186781871161bc8d6497004764b398401a401d4cce000

@property��ǩ����VBA��Let_Property��Get_Property������ֱ�ӻ�ȡ�����ö�������

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from types import MethodType


class Student(object):
    def __init__(self, name, score):
        self.__name = name
        self.__score = score

    def print_score(self):
        print('%s: %s' % (self.__name, self.__score))

    @property #��ҪSetter������Getter������@score����������һ��
    def score(self):
        return self.__score

    @score.setter
    def score(self, value):
        if not isinstance(value, int):
            raise ValueError('score must be an integer!')
        if value < 0 or value > 100:
            raise ValueError('score must between 0 ~ 100!')
        self.__score = value

if __name__ == "__main__":
    s1 = Student("Ben", 99)
    # s1._Student__score = 9999
    print(s1.__dict__)
    print(s1.__dir__())
    s1.score = 97
    print(s1.score)
    s1.print_score()


#-----------------------------------------------------------------------------------------
#���ؼ̳�
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014318680104044a55f4a9dbf8452caf71e8dc68b75a18000
MixIn

�������ļ̳й�ϵʱ��ͨ�������߶��ǵ�һ�̳������ģ����磬Ostrich�̳���Bird�����ǣ������Ҫ�����롱����Ĺ��ܣ�ͨ�����ؼ̳оͿ���ʵ�֣����磬��Ostrich���˼̳���Bird�⣬��ͬʱ�̳�Runnable���������ͨ����֮ΪMixIn��

Ϊ�˸��õؿ����̳й�ϵ�����ǰ�Runnable��Flyable��ΪRunnableMixIn��FlyableMixIn�����Ƶģ��㻹���Զ������ʳ����CarnivorousMixIn��ֲʳ����HerbivoresMixIn����ĳ������ͬʱӵ�кü���MixIn��

class Dog(Mammal, RunnableMixIn, CarnivorousMixIn):
    pass
MixIn��Ŀ�ľ��Ǹ�һ�������Ӷ�����ܣ���������������ʱ���������ȿ���ͨ�����ؼ̳�����϶��MixIn�Ĺ��ܣ���������ƶ��εĸ��ӵļ̳й�ϵ��

Python�Դ��ĺܶ��Ҳʹ����MixIn���ٸ����ӣ�Python�Դ���TCPServer��UDPServer������������񣬶�Ҫͬʱ�������û��ͱ���ʹ�ö���̻���߳�ģ�ͣ�������ģ����ForkingMixIn��ThreadingMixIn�ṩ��ͨ����ϣ����ǾͿ��Դ�������ʵķ�������

���磬��дһ�������ģʽ��TCP���񣬶������£�

class MyTCPServer(TCPServer, ForkingMixIn):
    pass
��дһ�����߳�ģʽ��UDP���񣬶������£�

class MyUDPServer(UDPServer, ThreadingMixIn):
    pass
���������һ�����Ƚ���Э��ģ�ͣ����Ա�дһ��CoroutineMixIn��

class MyTCPServer(TCPServer, CoroutineMixIn):
    pass
����һ�������ǲ���Ҫ���Ӷ��Ӵ�ļ̳�����ֻҪѡ����ϲ�ͬ����Ĺ��ܣ��Ϳ��Կ��ٹ������������ࡣ
#-----------------------------------------------------------------------------------------
#������
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014319098638265527beb24f7840aa97de564ccc7f20f6000
__str__

�����ȶ���һ��Student�࣬��ӡһ��ʵ����

>>> class Student(object):
...     def __init__(self, name):
...         self.name = name
...
>>> print(Student('Michael'))
<__main__.Student object at 0x109afb190>
��ӡ��һ��<__main__.Student object at 0x109afb190>�����ÿ���

��ô���ܴ�ӡ�úÿ��أ�ֻ��Ҫ�����__str__()����������һ���ÿ����ַ����Ϳ����ˣ�

>>> class Student(object):
...     def __init__(self, name):
...         self.name = name
...     def __str__(self):
...         return 'Student object (name: %s)' % self.name
...
>>> print(Student('Michael'))
Student object (name: Michael)
������ӡ������ʵ���������ÿ����������׿���ʵ���ڲ���Ҫ�����ݡ�

����ϸ�ĵ����ѻᷢ��ֱ���ñ�������print����ӡ������ʵ�����ǲ��ÿ���

>>> s = Student('Michael')
>>> s
<__main__.Student object at 0x109afb310>
������Ϊֱ����ʾ�������õĲ���__str__()������__repr__()�����ߵ�������__str__()�����û��������ַ�������__repr__()���س��򿪷��߿������ַ�����Ҳ����˵��__repr__()��Ϊ���Է���ġ�

����취���ٶ���һ��__repr__()������ͨ��__str__()��__repr__()���붼��һ���ģ����ԣ��и�͵����д����

class Student(object):
    def __init__(self, name):
        self.name = name
    def __str__(self):
        return 'Student object (name=%s)' % self.name
    __repr__ = __str__
    

__iter__

���һ�����뱻����for ... inѭ��������list��tuple�������ͱ���ʵ��һ��__iter__()�������÷�������һ����������Ȼ��Python��forѭ���ͻ᲻�ϵ��øõ��������__next__()�����õ�ѭ������һ��ֵ��ֱ������StopIteration����ʱ�˳�ѭ����

������쳲���������Ϊ����дһ��Fib�࣬����������forѭ����

class Fib(object):
    def __init__(self):
        self.a, self.b = 0, 1 # ��ʼ������������a��b

    def __iter__(self):
        return self # ʵ��������ǵ������󣬹ʷ����Լ�

    def __next__(self):
        self.a, self.b = self.b, self.a + self.b # ������һ��ֵ
        if self.a > 100000: # �˳�ѭ��������
            raise StopIteration()
        return self.a # ������һ��ֵ
���ڣ����԰�Fibʵ��������forѭ����

>>> for n in Fib():
...     print(n)
...
1
1
2
3
5
...
46368
75025

__getitem__

Fibʵ����Ȼ��������forѭ������������list�е��񣬵��ǣ���������list��ʹ�û��ǲ��У����磬ȡ��5��Ԫ�أ�

>>> Fib()[5]
Traceback (most recent call last):
  File "<stdin>", line 1, in <module>
TypeError: 'Fib' object does not support indexing
Ҫ���ֵ���list���������±�ȡ��Ԫ�أ���Ҫʵ��__getitem__()������

class Fib(object):
    def __getitem__(self, n):
        a, b = 1, 1
        for x in range(n):
            a, b = b, a + b
        return a
���ڣ��Ϳ��԰��±�������е�����һ���ˣ�

>>> f = Fib()
>>> f[0]
1
>>> f[1]
1
>>> f[2]
2
>>> f[3]
3
>>> f[10]
89
>>> f[100]
573147844013817084101
����list�и��������Ƭ������

>>> list(range(100))[5:10]
[5, 6, 7, 8, 9]
����Fibȴ����ԭ����__getitem__()����Ĳ���������һ��int��Ҳ������һ����Ƭ����slice������Ҫ���жϣ�

class Fib(object):
    def __getitem__(self, n):
        if isinstance(n, int): # n������
            a, b = 1, 1
            for x in range(n):
                a, b = b, a + b
            return a
        if isinstance(n, slice): # n����Ƭ
            start = n.start
            stop = n.stop
            if start is None:
                start = 0
            a, b = 1, 1
            L = []
            for x in range(stop):
                if x >= start:
                    L.append(a)
                a, b = b, a + b
            return L
��������Fib����Ƭ��

>>> f = Fib()
>>> f[0:5]
[1, 1, 2, 3, 5]
>>> f[:10]
[1, 1, 2, 3, 5, 8, 13, 21, 34, 55]
����û�ж�step����������

>>> f[:10:2]
[1, 1, 2, 3, 5, 8, 13, 21, 34, 55, 89]
Ҳû�жԸ������������ԣ�Ҫ��ȷʵ��һ��__getitem__()�����кܶ๤��Ҫ���ġ�

���⣬����Ѷ��󿴳�dict��__getitem__()�Ĳ���Ҳ������һ��������key��object������str��

��֮��Ӧ����__setitem__()�������Ѷ�������list��dict���Լ��ϸ�ֵ����󣬻���һ��__delitem__()����������ɾ��ĳ��Ԫ�ء�

��֮��ͨ������ķ����������Լ����������ֵú�Python�Դ���list��tuple��dictûʲô��������ȫ�鹦�ڶ�̬���Եġ�Ѽ�����͡�������Ҫǿ�Ƽ̳�ĳ���ӿڡ�


__getattr__

��������£������ǵ�����ķ���������ʱ����������ڣ��ͻᱨ�����綨��Student�ࣺ

class Student(object):

    def __init__(self):
        self.name = 'Michael'
����name���ԣ�û���⣬���ǣ����ò����ڵ�score���ԣ����������ˣ�

>>> s = Student()
>>> print(s.name)
Michael
>>> print(s.score)
Traceback (most recent call last):
  ...
AttributeError: 'Student' object has no attribute 'score'
������Ϣ������ظ������ǣ�û���ҵ�score���attribute��

Ҫ����������󣬳��˿��Լ���һ��score�����⣬Python������һ�����ƣ��Ǿ���дһ��__getattr__()��������̬����һ�����ԡ��޸����£�

class Student(object):

    def __init__(self):
        self.name = 'Michael'

    def __getattr__(self, attr):
        if attr=='score':
            return 99
�����ò����ڵ�����ʱ������score��Python����������ͼ����__getattr__(self, 'score')�����Ի�����ԣ����������Ǿ��л��᷵��score��ֵ��

>>> s = Student()
>>> s.name
'Michael'
>>> s.score
99
���غ���Ҳ����ȫ���Եģ�

class Student(object):

    def __getattr__(self, attr):
        if attr=='age':
            return lambda: 25
ֻ�ǵ��÷�ʽҪ��Ϊ��

>>> s.age()
25
ע�⣬ֻ����û���ҵ����Ե�����£��ŵ���__getattr__�����е����ԣ�����name��������__getattr__�в��ҡ�

���⣬ע�⵽���������s.abc���᷵��None��������Ϊ���Ƕ����__getattr__Ĭ�Ϸ��ؾ���None��Ҫ��classֻ��Ӧ�ض��ļ������ԣ����Ǿ�Ҫ����Լ�����׳�AttributeError�Ĵ���

class Student(object):

    def __getattr__(self, attr):
        if attr=='age':
            return lambda: 25
        raise AttributeError('\'Student\' object has no attribute \'%s\'' % attr)
��ʵ���Ͽ��԰�һ������������Ժͷ�������ȫ����̬�������ˣ�����Ҫ�κ������ֶΡ�

������ȫ��̬���õ�������ʲôʵ�������أ����þ��ǣ����������ȫ��̬����������á�

�ٸ����ӣ�

���ںܶ���վ����REST API����������΢��������ɶ�ģ�����API��URL���ƣ�

http://api.server/user/friends
http://api.server/user/timeline/list
���ҪдSDK����ÿ��URL��Ӧ��API��дһ���������ǵ����������ң�APIһ���Ķ���SDKҲҪ�ġ�

������ȫ��̬��__getattr__�����ǿ���д��һ����ʽ���ã�

class Chain(object):

    def __init__(self, path=''):
        self._path = path

    def __getattr__(self, path):
        return Chain('%s/%s' % (self._path, path))

    def __str__(self):
        return self._path

    __repr__ = __str__
���ԣ�

>>> Chain().status.user.timeline.list #ÿһ��.���Ի�ȡ�Ķ���һ��Chain����
'/status/user/timeline/list'
����������API��ô�䣬SDK�����Ը���URLʵ����ȫ��̬�ĵ��ã����ң�����API�����Ӷ��ı䣡

����ЩREST API��Ѳ����ŵ�URL�У�����GitHub��API��

GET /users/:user/repos
����ʱ����Ҫ��:user�滻Ϊʵ���û��������������д����������ʽ���ã�

Chain().users('michael').repos
�Ϳ��Էǳ�����ص���API�ˡ�����Ȥ��ͯЬ��������д������
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

class Chain(object):
    def __init__(self, path=''):
        self._path = path

    def __getattr__(self, path):
        if path == "users":
            return lambda path: Chain('%s/users/%s' % (self._path, path))
        return Chain('%s/%s' % (self._path, path)) #���ﷵ�ص���һ��Chain()��ʵ��������û��(str)����������Ҫ���users('str')ʱ����Ҫ����һ���ɵ��õĶ��󣬼�����һ���������ú�������һ������������һ���µ�Chain����

    def __str__(self):
        return self._path

    __repr__ = __str__


if __name__ == "__main__":
    s = Chain().abc.users('michael').repos.opq
    print(s)
�����
/abc/users/michael/repos/opq


__call__

һ������ʵ���������Լ������Ժͷ����������ǵ���ʵ������ʱ��������instance.method()�����á��ܲ���ֱ����ʵ�������ϵ����أ���Python�У����ǿ϶��ġ�

�κ��ֻ࣬��Ҫ����һ��__call__()�������Ϳ���ֱ�Ӷ�ʵ�����е��á��뿴ʾ����

class Student(object):
    def __init__(self, name):
        self.name = name

    def __call__(self):
        print('My name is %s.' % self.name)
���÷�ʽ���£�

>>> s = Student('Michael')
>>> s() # self������Ҫ����
My name is Michael.
__call__()�����Զ����������ʵ������ֱ�ӵ��þͺñȶ�һ���������е���һ������������ȫ���԰Ѷ��󿴳ɺ������Ѻ������ɶ�����Ϊ������֮�䱾����ûɶ����������

�����Ѷ��󿴳ɺ�������ô����������ʵҲ�����������ڶ�̬������������Ϊ���ʵ�����������ڴ��������ģ���ôһ�������Ǿ�ģ���˶���ͺ����Ľ��ޡ�

��ô����ô�ж�һ�������Ƕ����Ǻ����أ���ʵ�������ʱ��������Ҫ�ж�һ�������Ƿ��ܱ����ã��ܱ����õĶ������һ��Callable���󣬱��纯�����������涨��Ĵ���__call__()����ʵ����

>>> callable(Student())
True
>>> callable(max)
True
>>> callable([1, 2, 3])
False
>>> callable(None)
False
>>> callable('str')
False
ͨ��callable()���������ǾͿ����ж�һ�������Ƿ��ǡ��ɵ��á�����

ѧϰ��__call__�����������Chain������޸�Ϊ��
class Chain(object):
    def __init__(self, path=''):
        self._path = path

    def __getattr__(self, path):
        return Chain('%s/%s' % (self._path, path))

    def __str__(self):
        return self._path

    def __call__(self, s):
        return Chain('%s/%s' % (self._path, s))

    __repr__ = __str__
#-----------------------------------------------------------------------------------------
#ʹ��ö����
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143191235886950998592cd3e426e91687cdae696e64b000

��������Ҫ���峣��ʱ��һ���취���ô�д����ͨ�����������壬�����·ݣ�

JAN = 1
FEB = 2
MAR = 3
...
NOV = 11
DEC = 12
�ô��Ǽ򵥣�ȱ����������int��������Ȼ�Ǳ�����

���õķ�����Ϊ������ö�����Ͷ���һ��class���ͣ�Ȼ��ÿ����������class��һ��Ψһʵ����Python�ṩ��Enum����ʵ��������ܣ�

from enum import Enum

Month = Enum('Month', ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'))
�������Ǿͻ����Month���͵�ö���࣬����ֱ��ʹ��Month.Jan������һ������������ö���������г�Ա��

for name, member in Month.__members__.items():
    print(name, '=>', member, ',', member.value)
value���������Զ�������Ա��int������Ĭ�ϴ�1��ʼ������


#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from enum import Enum


if __name__ == "__main__":
    Month = Enum('Month', ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'))
    print(type(Month.__members__))
    for name, member in Month.__members__.items():
        print(name, '=>', member, ',', member.value)
�����
<class 'mappingproxy'>
Jan => Month.Jan , 1
Feb => Month.Feb , 2
Mar => Month.Mar , 3
Apr => Month.Apr , 4
May => Month.May , 5
Jun => Month.Jun , 6
Jul => Month.Jul , 7
Aug => Month.Aug , 8
Sep => Month.Sep , 9
Oct => Month.Oct , 10
Nov => Month.Nov , 11
Dec => Month.Dec , 12

�����Ҫ����ȷ�ؿ���ö�����ͣ����Դ�Enum�������Զ����ࣺ

from enum import Enum, unique

@unique
class Weekday(Enum):
    Sun = 0 # Sun��value���趨Ϊ0
    Mon = 1
    Tue = 2
    Wed = 3
    Thu = 4
    Fri = 5
    Sat = 6
@uniqueװ�������԰������Ǽ�鱣֤û���ظ�ֵ��
������Щö�����Ϳ����������ַ�����
>>> day1 = Weekday.Mon
>>> print(day1)
Weekday.Mon
>>> print(Weekday.Tue)
Weekday.Tue
>>> print(Weekday['Tue'])
Weekday.Tue
>>> print(Weekday.Tue.value)
2
>>> print(day1 == Weekday.Mon)
True
>>> print(day1 == Weekday.Tue)
False
>>> print(Weekday(1))
Weekday.Mon
>>> print(day1 == Weekday(1))
True
>>> Weekday(7)
Traceback (most recent call last):
  ...
ValueError: 7 is not a valid Weekday
>>> for name, member in Weekday.__members__.items():
        print(name, '=>', member, member.value)
...
Sun => Weekday.Sun 0
Mon => Weekday.Mon 1
Tue => Weekday.Tue 2
Wed => Weekday.Wed 3
Thu => Weekday.Thu 4
Fri => Weekday.Fri 5
Sat => Weekday.Sat 6
�ɼ����ȿ����ó�Ա��������ö�ٳ������ֿ���ֱ�Ӹ���value��ֵ���ö�ٳ�����

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from enum import Enum, unique

@unique
class Weekday(Enum):
    Sun = 0
    Mon = 1
    Tue = 2
    Wed = 3
    Thu = 4
    Fri = 5
    Sat = 0

if __name__ == "__main__":
    day1 = Weekday.Mon
    print(day1)

#-----------------------------------------------------------------------------------------
#ʹ��Ԫ�� https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014319106919344c4ef8b1e04c48778bb45796e0335839000

type()

��̬���Ժ;�̬�������Ĳ�ͬ�����Ǻ�������Ķ��壬���Ǳ���ʱ����ģ���������ʱ��̬�����ġ�

�ȷ�˵����Ҫ����һ��Hello��class����дһ��hello.pyģ�飺

class Hello(object):
    def hello(self, name='world'):
        print('Hello, %s.' % name)
��Python����������helloģ��ʱ���ͻ�����ִ�и�ģ���������䣬ִ�н�����Ƕ�̬������һ��Hello��class���󣬲������£�

>>> from hello import Hello
>>> h = Hello()
>>> h.hello()
Hello, world.
>>> print(type(Hello))
<class 'type'>
>>> print(type(h))
<class 'hello.Hello'>
type()�������Բ鿴һ�����ͻ���������ͣ�Hello��һ��class���������;���type����h��һ��ʵ�����������;���class Hello��

����˵class�Ķ���������ʱ��̬�����ģ�������class�ķ�������ʹ��type()������

type()�����ȿ��Է���һ����������ͣ��ֿ��Դ������µ����ͣ����磬���ǿ���ͨ��type()����������Hello�࣬������ͨ��class Hello(object)...�Ķ��壺

>>> def fn(self, name='world'): # �ȶ��庯��
...     print('Hello, %s.' % name)
...
>>> Hello = type('Hello', (object,), dict(hello=fn)) # ����Hello class
>>> h = Hello()
>>> h.hello()
Hello, world.
>>> print(type(Hello))
<class 'type'>
>>> print(type(h))
<class '__main__.Hello'>
Ҫ����һ��class����type()�������δ���3��������

class�����ƣ�
�̳еĸ��༯�ϣ�ע��Python֧�ֶ��ؼ̳У����ֻ��һ�����࣬������tuple�ĵ�Ԫ��д����
class�ķ��������뺯���󶨣��������ǰѺ���fn�󶨵�������hello�ϡ�
ͨ��type()�������������ֱ��дclass����ȫһ���ģ���ΪPython����������class����ʱ��������ɨ��һ��class������﷨��Ȼ�����type()����������class��

��������£����Ƕ���class Xxx...�������࣬���ǣ�type()����Ҳ�������Ƕ�̬������������Ҳ����˵����̬���Ա���֧�������ڶ�̬�����࣬��;�̬�����зǳ���Ĳ�ͬ��Ҫ�ھ�̬���������ڴ����࣬���빹��Դ�����ַ����ٵ��ñ����������߽���һЩ���������ֽ���ʵ�֣������϶��Ƕ�̬���룬��ǳ����ӡ�

metaclass
����ʹ��type()��̬���������⣬Ҫ������Ĵ�����Ϊ��������ʹ��metaclass��
metaclass��ֱ��ΪԪ�࣬�򵥵Ľ��;��ǣ�

�����Ƕ��������Ժ󣬾Ϳ��Ը�������ഴ����ʵ�������ԣ��ȶ����࣬Ȼ�󴴽�ʵ����

������������봴�������أ��Ǿͱ������metaclass�������࣬���ԣ��ȶ���metaclass��Ȼ�󴴽��ࡣ

�����������ǣ��ȶ���metaclass���Ϳ��Դ����࣬��󴴽�ʵ����

���ԣ�metaclass�����㴴��������޸��ࡣ���仰˵������԰��࿴����metaclass���������ġ�ʵ������

metaclass��Python���������������⣬Ҳ������ʹ�õ�ħ�����롣��������£��㲻��������Ҫʹ��metaclass����������ԣ��������ݿ�����Ҳû��ϵ����Ϊ�������㲻���õ���

�����ȿ�һ���򵥵����ӣ����metaclass���Ը������Զ����MyList����һ��add������

����ListMetaclass������Ĭ��ϰ�ߣ�metaclass������������Metaclass��β���Ա�����ر�ʾ����һ��metaclass��

# metaclass�����ģ�壬���Ա����`type`����������
class ListMetaclass(type):
    def __new__(cls, name, bases, attrs):
        attrs['add'] = lambda self, value: self.append(value)
        return type.__new__(cls, name, bases, attrs)
����ListMetaclass�������ڶ������ʱ��Ҫָʾʹ��ListMetaclass�������࣬����ؼ��ֲ���metaclass��

class MyList(list, metaclass=ListMetaclass):
    pass
�����Ǵ���ؼ��ֲ���metaclassʱ��ħ������Ч�ˣ���ָʾPython�������ڴ���MyListʱ��Ҫͨ��ListMetaclass.__new__()���������ڴˣ����ǿ����޸���Ķ��壬���磬�����µķ�����Ȼ�󣬷����޸ĺ�Ķ��塣

__new__()�������յ��Ĳ��������ǣ�

��ǰ׼����������Ķ���

������֣�

��̳еĸ��༯�ϣ�

��ķ������ϡ�

����һ��MyList�Ƿ���Ե���add()������

>>> L = MyList()
>>> L.add(1)
>> L
[1]
����ͨ��listû��add()������

>>> L2 = list()
>>> L2.add(1)
Traceback (most recent call last):
  File "<stdin>", line 1, in <module>
AttributeError: 'list' object has no attribute 'add'
��̬�޸���ʲô���壿ֱ����MyList������д��add()�������Ǹ�������������£�ȷʵӦ��ֱ��д��ͨ��metaclass�޸Ĵ�����̬��

���ǣ��ܻ�������Ҫͨ��metaclass�޸��ඨ��ġ�ORM����һ�����͵����ӡ�

ORMȫ�ơ�Object Relational Mapping����������-��ϵӳ�䣬���ǰѹ�ϵ���ݿ��һ��ӳ��Ϊһ������Ҳ����һ�����Ӧһ����������д������򵥣�����ֱ�Ӳ���SQL��䡣

Ҫ��дһ��ORM��ܣ����е��඼ֻ�ܶ�̬���壬��Ϊֻ��ʹ���߲��ܸ��ݱ�Ľṹ�������Ӧ��������

�����������Ա�дһ��ORM��ܡ�

#https://blog.csdn.net/johnsonguo/article/details/585193
#https://www.zhihu.com/question/20040039

class Base(object):
    def __init__(self):
        print('Base.__init__')

class A(Base):
    def __init__(self):
        print('A.__init__ begin')
        super(A, self).__init__()
        print('A.__init__ end')

class B(Base):
    def __init__(self):
        print('B.__init__ begin')
        super(B, self).__init__()
        print('B.__init__ end')

class C(A,B):
    def __init__(self):
        print('C.__init__ begin')
        super(C, self).__init__()  # Only one call to super() here
        print('C.__init__ end')

c = C()
print C.__mro__
���н����C.__init__ begin
A.__init__ begin
B.__init__ begin
Base.__init__
B.__init__ end
A.__init__ end
C.__init__ end
(<class '__main__.C'>, <class '__main__.A'>, <class '__main__.B'>, <class '__main__.Base'>, <type 'object'>)
����ʹ�� super() ����ʱ��Python����MRO�б��ϼ���������һ���ࡣ ֻҪÿ���ض���ķ���ͳһʹ�� super() ��ֻ������һ�Σ� ��ô���������ջ����������MRO�б�ÿ������Ҳֻ�ᱻ����һ�Ρ� ��Ҳ��Ϊʲô�ڵڶ����������㲻��������� Base.__init__() ��ԭ��

1. super������һ����������һ������������super(B, self)��ʵ�ϵ�����super��ĳ�ʼ��������
������һ��super����
2. super��ĳ�ʼ��������û����ʲô����Ĳ�����ֻ�Ǽ򵥼�¼�������ͺ;���ʵ����
3. super(B, self).func�ĵ��ò��������ڵ��õ�ǰ��ĸ����func������
4. Python�Ķ�̳�����ͨ��mro�ķ�ʽ����֤��������ĺ�������һ���ã����ұ�֤ÿ�����ຯ��
ֻ����һ�Σ����ÿ���඼ʹ��super����
5. ����super��ͷǰ󶨵ĺ�����һ��Σ����Ϊ������ܵ���Ӧ�õ��õĸ��ຯ��û�е��û���һ
�����ຯ�������ö�Ρ�



#!/usr/bin/env python3
# -*- coding: utf-8 -*-

' XXX module '

__author__ = 'DingBen'

if __name__=='__main__':
    pass


class Field(object):
    def __init__(self, name, column_type):
        self.name = name
        self.column_type = column_type

    def __str__(self):
        return '<%s:%s>' % (self.__class__.__name__, self.name)


class StringField(Field):
    def __init__(self, name):
        super(StringField, self).__init__(name, 'varchar(100')


class IntegerField(Field):
    def __init__(self, name):
        super(IntegerField, self).__init__(name, 'bigint')
        
        
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

' XXX module '
from src.Field import Field

__author__ = 'DingBen'


class ModelMetaClass(type):
    def __new__(cls, name, bases, attrs):
        if name == 'Model':
            return type.__new__(cls, name, bases, attrs)
        print("Found model: %s" % name)
        mappings = dict()
        for k, v in attrs.items():
            if isinstance(v, Field):
                print("Found mapping: %s ==> %s" % (k, v))
                mappings[k] = v
        for k in mappings.keys():
            attrs.pop(k)
        attrs['__mappings__'] = mappings
        attrs['__table__'] = name
        return type.__new__(cls, name, bases, attrs)


class Model(dict, metaclass=ModelMetaClass):
    def __init__(self, **kw):
        super(Model, self).__init__(**kw)

    def __getattr__(self, key):
        try:
            return self[key]
        except KeyError:
            raise AttributeError(r"'Model' object has no attribute '%s'" % key)

    def __setattr__(self, key, value):
        self[key] = value

    def save(self):
        fields = []
        params = []
        args = []
        for k, v in self.__mappings__.items():
            fields.append(v.name)
            params.append('?')
            args.append(getattr(self, k, None))
        sql = 'insert into %s (%s) values (%s)' % (self.__table__, ','.join(fields), ','.join(params))
        print('SQL: %s' % sql)
        print('ARGS: %s' % str(args))

        
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from src.Field import IntegerField, StringField
from src.Model import Model


class User(Model):
    # ����������Ե��е�ӳ�䣺
    id = IntegerField('id')
    name = StringField('username')
    email = StringField('email')
    password = StringField('password')


if __name__ == "__main__":
    u = User(id=12345, name="Michael", email="test@orm.org", password="my--pwd")
    u.save()

���û�����һ��class User(Model)ʱ��Python�����������ڵ�ǰ��User�Ķ����в���metaclass�����û���ҵ����ͼ����ڸ���Model�в���metaclass���ҵ��ˣ���ʹ��Model�ж����metaclass��ModelMetaclass������User�࣬Ҳ����˵��metaclass������ʽ�ؼ̳е����࣬�������Լ�ȴ�о�������

��ModelMetaclass�У�һ�����˼������飺

�ų�����Model����޸ģ�

�ڵ�ǰ�ࣨ����User���в��Ҷ��������������ԣ�����ҵ�һ��Field���ԣ��Ͱ������浽һ��__mappings__��dict�У�ͬʱ����������ɾ����Field���ԣ����������������ʱ����ʵ�������Ի��ڸ����ͬ�����ԣ���

�ѱ������浽__table__�У������Ϊ����Ĭ��Ϊ������

��Model���У��Ϳ��Զ�����ֲ������ݿ�ķ���������save()��delete()��find()��update�ȵȡ�

����ʵ����save()��������һ��ʵ�����浽���ݿ��С���Ϊ�б��������Ե��ֶε�ӳ�������ֵ�ļ��ϣ��Ϳ��Թ����INSERT��䡣

��д�������ԣ�

u = User(id=12345, name='Michael', email='test@orm.org', password='my-pwd')
u.save()
������£�

Found model: User
Found mapping: email ==> <StringField:email>
Found mapping: password ==> <StringField:password>
Found mapping: id ==> <IntegerField:uid>
Found mapping: name ==> <StringField:username>
SQL: insert into User (password,email,username,id) values (?,?,?,?)
ARGS: ['my-pwd', 'test@orm.org', 'Michael', 12345]
���Կ�����save()�����Ѿ���ӡ���˿�ִ�е�SQL��䣬�Լ������б�ֻ��Ҫ�������ӵ����ݿ⣬ִ�и�SQL��䣬�Ϳ�����������Ĺ��ܡ�

����100�д��룬���Ǿ�ͨ��metaclassʵ����һ�������ORM��ܣ��ǲ��Ƿǳ��򵥣�

С��

metaclass��Python�зǳ�����ħ���ԵĶ��������Ըı��ഴ��ʱ����Ϊ������ǿ��Ĺ���ʹ���������С�ġ�
#-----------------------------------------------------------------------------------------
#������
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143191375461417a222c54b7e4d65b258f491c093a515000
try

��������һ������������try�Ļ��ƣ�

try:
    print('try...')
    r = 10 / 0
    print('result:', r)
except ZeroDivisionError as e:
    print('except:', e)
finally:
    print('finally...')
print('END')

int()�������ܻ��׳�ValueError������������һ��except����ValueError������һ��except����ZeroDivisionError��
���⣬���û�д�������������except��������һ��else����û�д�����ʱ�����Զ�ִ��else��䣺
try:
    print('try...')
    r = 10 / int('2')
    print('result:', r)
except ValueError as e:
    print('ValueError:', e)
except ZeroDivisionError as e:
    print('ZeroDivisionError:', e)
else:
    print('no error!')
finally:
    print('finally...')
print('END')

Python���еĴ����Ǵ�BaseException�������ģ������Ĵ������ͺͼ̳й�ϵ�����
#https://docs.python.org/3/library/exceptions.html#exception-hierarchy
The class hierarchy for built-in exceptions is:

BaseException
 +-- SystemExit
 +-- KeyboardInterrupt
 +-- GeneratorExit
 +-- Exception
      +-- StopIteration
      +-- StopAsyncIteration
      +-- ArithmeticError
      |    +-- FloatingPointError
      |    +-- OverflowError
      |    +-- ZeroDivisionError
      +-- AssertionError
      +-- AttributeError
      +-- BufferError
      +-- EOFError
      +-- ImportError
      |    +-- ModuleNotFoundError
      +-- LookupError
      |    +-- IndexError
      |    +-- KeyError
      +-- MemoryError
      +-- NameError
      |    +-- UnboundLocalError
      +-- OSError
      |    +-- BlockingIOError
      |    +-- ChildProcessError
      |    +-- ConnectionError
      |    |    +-- BrokenPipeError
      |    |    +-- ConnectionAbortedError
      |    |    +-- ConnectionRefusedError
      |    |    +-- ConnectionResetError
      |    +-- FileExistsError
      |    +-- FileNotFoundError
      |    +-- InterruptedError
      |    +-- IsADirectoryError
      |    +-- NotADirectoryError
      |    +-- PermissionError
      |    +-- ProcessLookupError
      |    +-- TimeoutError
      +-- ReferenceError
      +-- RuntimeError
      |    +-- NotImplementedError
      |    +-- RecursionError
      +-- SyntaxError
      |    +-- IndentationError
      |         +-- TabError
      +-- SystemError
      +-- TypeError
      +-- ValueError
      |    +-- UnicodeError
      |         +-- UnicodeDecodeError
      |         +-- UnicodeEncodeError
      |         +-- UnicodeTranslateError
      +-- Warning
           +-- DeprecationWarning
           +-- PendingDeprecationWarning
           +-- RuntimeWarning
           +-- SyntaxWarning
           +-- UserWarning
           +-- FutureWarning
           +-- ImportWarning
           +-- UnicodeWarning
           +-- BytesWarning
           +-- ResourceWarning

           
����ջ
�������û�б��������ͻ�һֱ�����ף����Python���������񣬴�ӡһ��������Ϣ��Ȼ������˳���������err.py��
# err.py:
def foo(s):
    return 10 / int(s)

def bar(s):
    return foo(s) * 2

def main():
    bar('0')

main()
ִ�У�������£�

$ python3 err.py
Traceback (most recent call last):
  File "err.py", line 11, in <module>
    main()
  File "err.py", line 9, in main
    bar('0')
  File "err.py", line 6, in bar
    return foo(s) * 2
  File "err.py", line 3, in foo
    return 10 / int(s)
ZeroDivisionError: division by zero
���������£����µ��ǲ�֪����������ˡ����������Ϣ�Ƕ�λ����Ĺؼ������Ǵ������¿��Կ�����������ĵ��ú�������

������Ϣ��1�У�

Traceback (most recent call last):
�����������Ǵ���ĸ�����Ϣ��

��2~3�У�

  File "err.py", line 11, in <module>
    main()
����main()�����ˣ��ڴ����ļ�err.py�ĵ�11�д��룬��ԭ���ǵ�9�У�

  File "err.py", line 9, in main
    bar('0')
����bar('0')�����ˣ��ڴ����ļ�err.py�ĵ�9�д��룬��ԭ���ǵ�6�У�

  File "err.py", line 6, in bar
    return foo(s) * 2
ԭ����return foo(s) * 2����������ˣ����⻹��������ԭ�򣬼������¿���

  File "err.py", line 3, in foo
    return 10 / int(s)
ԭ����return 10 / int(s)����������ˣ����Ǵ��������Դͷ����Ϊ�����ӡ�ˣ�

ZeroDivisionError: integer division or modulo by zero
���ݴ�������ZeroDivisionError�������жϣ�int(s)����û�г�������int(s)����0���ڼ���10 / 0ʱ�������ˣ��ҵ�����Դͷ��

 �����ʱ��һ��Ҫ��������ĵ���ջ��Ϣ�����ܶ�λ�����λ�á�

��¼����
��������������Ȼ������Python����������ӡ�������ջ��������Ҳ�������ˡ���Ȼ�����ܲ�����󣬾Ϳ��԰Ѵ����ջ��ӡ������Ȼ���������ԭ��ͬʱ���ó������ִ����ȥ��
Python���õ�loggingģ����Էǳ����׵ؼ�¼������Ϣ��
# err_logging.py

import logging

def foo(s):
    return 10 / int(s)

def bar(s):
    return foo(s) * 2

def main():
    try:
        bar('0')
    except Exception as e:
        logging.exception(e)

main()
print('END')
ͬ���ǳ����������ӡ�������Ϣ������ִ�У��������˳���

$ python3 err_logging.py
ERROR:root:division by zero
Traceback (most recent call last):
  File "err_logging.py", line 13, in main
    bar('0')
  File "err_logging.py", line 9, in bar
    return foo(s) * 2
  File "err_logging.py", line 6, in foo
    return 10 / int(s)
ZeroDivisionError: division by zero
END
ͨ�����ã�logging�����԰Ѵ����¼����־�ļ�������º��Ų顣


�׳�����
��Ϊ������class������һ��������ǲ��񵽸�class��һ��ʵ������ˣ����󲢲���ƾ�ղ����ģ��������ⴴ�����׳��ġ�Python�����ú������׳��ܶ����͵Ĵ��������Լ���д�ĺ���Ҳ�����׳�����
���Ҫ�׳��������ȸ�����Ҫ�����Զ���һ�������class��ѡ��ü̳й�ϵ��Ȼ����raise����׳�һ�������ʵ����
# err_raise.py
class FooError(ValueError):
    pass

def foo(s):
    n = int(s)
    if n==0:
        raise FooError('invalid value: %s' % s)
    return 10 / n

foo('0')
ִ�У����������ٵ������Լ�����Ĵ���

$ python3 err_raise.py 
Traceback (most recent call last):
  File "err_throw.py", line 11, in <module>
    foo('0')
  File "err_throw.py", line 8, in foo
    raise FooError('invalid value: %s' % s)
__main__.FooError: invalid value: 0
ֻ���ڱ�Ҫ��ʱ��Ŷ��������Լ��Ĵ������͡��������ѡ��Python���е����õĴ������ͣ�����ValueError��TypeError��������ʹ��Python���õĴ������͡�

�������������һ�ִ�����ķ�ʽ��

# err_reraise.py

def foo(s):
    n = int(s)
    if n==0:
        raise ValueError('invalid value: %s' % s)
    return 10 / n

def bar():
    try:
        foo('0')
    except ValueError as e:
        print('ValueError!')
        raise

bar()
��bar()�����У����������Ѿ������˴��󣬵��ǣ���ӡһ��ValueError!���ְѴ���ͨ��raise����׳�ȥ�ˣ��ⲻ�в�ô��

��ʵ���ִ�����ʽ����û���������൱�������������Ŀ��ֻ�Ǽ�¼һ�£����ں���׷�١����ǣ����ڵ�ǰ������֪��Ӧ����ô����ô������ԣ���ǡ���ķ�ʽ�Ǽ��������ף��ö��������ȥ�����ñ�һ��Ա��������һ������ʱ���Ͱ������׸������ϰ壬��������ϰ�Ҳ�����ˣ���һֱ�����ף����ջ��׸�CEOȥ����

raise�����������������ͻ�ѵ�ǰ����ԭ���׳������⣬��except��raiseһ��Error�������԰�һ�����͵Ĵ���ת������һ�����ͣ�

try:
    10 / 0
except ZeroDivisionError:
    raise ValueError('input error!')
ֻҪ�Ǻ����ת���߼��Ϳ��ԣ����ǣ�����Ӧ�ð�һ��IOErrorת���ɺ�����ɵ�ValueError��
#-----------------------------------------------------------------------------------------
#����
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431915578556ad30ab3933ae4e82a03ee2e9a4f70871000
��print()���Ļ����ǽ�������ɾ��������������ﵽ������print()�����н��Ҳ������ܶ�������Ϣ�����ԣ��������еڶ��ַ�����

����
������print()�������鿴�ĵط����������ö��ԣ�assert���������

def foo(s):
    n = int(s)
    assert n != 0, 'n is zero!'
    return 10 / n

def main():
    foo('0')
assert����˼�ǣ����ʽn != 0Ӧ����True�����򣬸��ݳ������е��߼�������Ĵ���϶������

�������ʧ�ܣ�assert��䱾��ͻ��׳�AssertionError��
$ python err.py
Traceback (most recent call last):
  ...
AssertionError: n is zero!
������������������assert����print()���Ҳ�ò�����ȥ������������Python������ʱ������-O�������ر�assert��

$ python -O err.py
Traceback (most recent call last):
  ...
ZeroDivisionError: division by zero
�رպ�����԰����е�assert��䵱��pass������

logging
��print()�滻Ϊlogging�ǵ�3�ַ�ʽ����assert�ȣ�logging�����׳����󣬶��ҿ���������ļ���
import logging

s = '0'
n = int(s)
logging.info('n = %d' % n)
print(10 / n)
logging.info()�Ϳ������һ���ı������У����ֳ���ZeroDivisionError��û���κ���Ϣ����ô���£�

�𼱣���import logging֮�����һ�����������ԣ�

import logging
logging.basicConfig(level=logging.INFO)
��������ˣ�

$ python err.py
INFO:root:n = 0
Traceback (most recent call last):
  File "err.py", line 8, in <module>
    print(10 / n)
ZeroDivisionError: division by zero
�����logging�ĺô�����������ָ����¼��Ϣ�ļ�����debug��info��warning��error�ȼ������𣬵�����ָ��level=INFOʱ��logging.debug�Ͳ��������ˡ�ͬ��ָ��level=WARNING��debug��info�Ͳ��������ˡ�����һ��������Է��ĵ������ͬ�������Ϣ��Ҳ����ɾ�������ͳһ��������ĸ��������Ϣ��

logging����һ���ô���ͨ���򵥵����ã�һ��������ͬʱ�������ͬ�ĵط�������console���ļ���

#https://blog.csdn.net/z_johnny/article/details/50740528 logging����
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import logging
#���ڻ����level��������Ż������Ĭ��ֻ��ʾ�˴��ڵ���WARNING����
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s %(filename)s [line:%(lineno)d] %(levelname)s %(message)s',
                    datefmt='%a %d %b %Y %H:%M:%S',
                    filename='test.log',
                    filemode='w')
                    
if __name__ == "__main__":
    logging.debug("debug") #��debug/info/warning/error/critical�����������Խ��Խ��
    logging.info("info")
    logging.warning("warning")
    logging.error("error")
    logging.critical("critical")
�����
Fri 30 Mar 2018 23:27:20 main.py [line:12] DEBUG debug
Fri 30 Mar 2018 23:27:20 main.py [line:13] INFO info
Fri 30 Mar 2018 23:27:20 main.py [line:14] WARNING warning
Fri 30 Mar 2018 23:27:20 main.py [line:15] ERROR error
Fri 30 Mar 2018 23:27:20 main.py [line:16] CRITICAL critical

�ɼ���logging.basicConfig()�����п�ͨ���������������loggingģ��Ĭ����Ϊ�����ò�����
filename��   ��ָ�����ļ�������FiledHandler����߻���彲��handler�ĸ����������־�ᱻ�洢��ָ�����ļ��С�
filemode��   �ļ��򿪷�ʽ����ָ����filenameʱʹ�����������Ĭ��ֵΪ��a������ָ��Ϊ��w����
format��      ָ��handlerʹ�õ���־��ʾ��ʽ��
datefmt��    ָ������ʱ���ʽ��
level��        ����rootlogger����߻ὲ�����������־����
stream��     ��ָ����stream����StreamHandler������ָ�������sys.stderr,sys.stdout�����ļ���Ĭ��Ϊsys.stderr��
                  ��ͬʱ�г���filename��stream������������stream�����ᱻ���ԡ�

format�����п����õ��ĸ�ʽ������
%(name)s             Logger������
%(levelno)s          ������ʽ����־����
%(levelname)s     �ı���ʽ����־����
%(pathname)s     ������־���������ģ�������·����������û��
%(filename)s        ������־���������ģ����ļ���
%(module)s          ������־���������ģ����
%(funcName)s     ������־��������ĺ�����
%(lineno)d           ������־���������������ڵĴ�����
%(created)f          ��ǰʱ�䣬��UNIX��׼�ı�ʾʱ��ĸ� ������ʾ
%(relativeCreated)d    �����־��Ϣʱ�ģ���Logger������ ���ĺ�����
%(asctime)s                �ַ�����ʽ�ĵ�ǰʱ�䡣Ĭ�ϸ�ʽ�� ��2003-07-08 16:49:45,896�������ź�����Ǻ���
%(thread)d                 �߳�ID������û��
%(threadName)s        �߳���������û��
%(process)d              ����ID������û��
%(message)s            �û��������Ϣ
#-----------------------------------------------------------------------------------------
#��Ԫ����
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143191629979802b566644aa84656b50cd484ec4a7838000
��д��Ԫ����ʱ��������Ҫ��дһ�������࣬��unittest.TestCase�̳С�

��test��ͷ�ķ������ǲ��Է���������test��ͷ�ķ���������Ϊ�ǲ��Է��������Ե�ʱ�򲻻ᱻִ�С�

��ÿһ����Զ���Ҫ��дһ��test_xxx()����������unittest.TestCase�ṩ�˺ܶ����õ������жϣ�����ֻ��Ҫ������Щ�����Ϳ��Զ�������Ƿ��������������ġ���õĶ��Ծ���assertEqual()��

self.assertEqual(abs(-1), 1) # ���Ժ������صĽ����1���
��һ����Ҫ�Ķ��Ծ����ڴ��׳�ָ�����͵�Error������ͨ��d['empty']���ʲ����ڵ�keyʱ�����Ի��׳�KeyError��

with self.assertRaises(KeyError):
    value = d['empty']
��ͨ��d.empty���ʲ����ڵ�keyʱ�������ڴ��׳�AttributeError��

with self.assertRaises(AttributeError):
    value = d.empty
    
    
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import unittest
from src.mydict import Dict


class TestDict(unittest.TestCase):
    def test_init(self):
        d = Dict(a=1, b='test')
        self.assertEqual(d.a, 1)
        self.assertEqual(d.b, 'test')
        self.assertTrue(isinstance(d, dict))

    def test_key(self):
        d = Dict()
        d['key'] = 'value'
        self.assertEqual(d.key, 'value')

    def test_attr(self):
        d = Dict()
        d.key = 'value'
        self.assertTrue('key' in d)
        self.assertEqual(d['key'], 'value')

    def test_keyerror(self):
        d = Dict()
        with self.assertRaises(KeyError):
            value = d['empty']

    def test_attrerror(self):
        d = Dict()
        with self.assertRaises(AttributeError):
            value = d.empty

if __name__ == "__main__":
    unittest.main()
    
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

' XXX module '
class Dict(dict):
    def __init__(self, **kw):
        super().__init__(**kw)

    def __getattr__(self, key):
        try:
            return self[key]
        except KeyError:
            raise AttributeError(r"'Dict' object has no attribute '%s'" % key)

    def __setattr__(self, key, value):
        self[key] = value

���е�Ԫ����
һ����д�õ�Ԫ���ԣ����ǾͿ������е�Ԫ���ԡ���򵥵����з�ʽ����mydict_test.py�����������д��룺
if __name__ == '__main__':
    unittest.main()
�����Ϳ��԰�mydict_test.py����������python�ű����У�

��һ�ַ�������������ͨ������-m unittestֱ�����е�Ԫ���ԣ�
E:\Program Files\JetBrains\PythonProject\Py3TestProject>python -m unittest src\main.py
.....
----------------------------------------------------------------------
Ran 5 tests in 0.001s

OK
�����Ƽ�����������Ϊ��������һ���������кܶ൥Ԫ���ԣ����ң��кܶ๤�߿����Զ���������Щ��Ԫ���ԡ�

setUp��tearDown
�����ڵ�Ԫ�����б�д���������setUp()��tearDown()������������������ֱ���ÿ����һ�����Է�����ǰ��ֱ�ִ�С�
setUp()��tearDown()������ʲô���أ�������Ĳ�����Ҫ����һ�����ݿ⣬��ʱ���Ϳ�����setUp()�������������ݿ⣬��tearDown()�����йر����ݿ⣬������������ÿ�����Է������ظ���ͬ�Ĵ��룺
class TestDict(unittest.TestCase):

    def setUp(self):
        print('setUp...')

    def tearDown(self):
        print('tearDown...')
�����ٴ����в��Կ���ÿ�����Է�������ǰ���Ƿ���ӡ��setUp...��tearDown...��
#-----------------------------------------------------------------------------------------
#�ļ���д
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431917715991ef1ebc19d15a4afdace1169a464eecc2000

file-like Object

��open()�������ص������и�read()�����Ķ�����Python��ͳ��Ϊfile-like Object������file�⣬���������ڴ���ֽ��������������Զ������ȵȡ�file-like Object��Ҫ����ض���̳У�ֻҪд��read()�������С�

StringIO�������ڴ��д�����file-like Object����������ʱ���塣

�������ļ�

ǰ�潲��Ĭ�϶��Ƕ�ȡ�ı��ļ���������UTF-8������ı��ļ���Ҫ��ȡ�������ļ�������ͼƬ����Ƶ�ȵȣ���'rb'ģʽ���ļ����ɣ�

>>> f = open('/Users/michael/test.jpg', 'rb')
>>> f.read()
b'\xff\xd8\xff\xe1\x00\x18Exif\x00\x00...' # ʮ�����Ʊ�ʾ���ֽ�


�ַ�����

Ҫ��ȡ��UTF-8������ı��ļ�����Ҫ��open()��������encoding���������磬��ȡGBK������ļ���

>>> f = open('/Users/michael/gbk.txt', 'r', encoding='gbk')
>>> f.read()
'����'
������Щ���벻�淶���ļ�������ܻ�����UnicodeDecodeError����Ϊ���ı��ļ��п��ܼ�����һЩ�Ƿ�������ַ����������������open()����������һ��errors��������ʾ�����������������δ�����򵥵ķ�ʽ��ֱ�Ӻ��ԣ�

>>> f = open('/Users/michael/gbk.txt', 'r', encoding='gbk', errors='ignore')

С��
��Python�У��ļ���д��ͨ��open()�����򿪵��ļ�������ɵġ�ʹ��with�������ļ�IO�Ǹ���ϰ�ߡ�
#-----------------------------------------------------------------------------------------
#StringIO
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431918785710e86a1a120ce04925bae155012c7fc71e000

�ܶ�ʱ�����ݶ�д��һ�����ļ���Ҳ�������ڴ��ж�д��

StringIO����˼��������ڴ��ж�дstr��

Ҫ��strд��StringIO��������Ҫ�ȴ���һ��StringIO��Ȼ�����ļ�һ��д�뼴�ɣ�

>>> from io import StringIO
>>> f = StringIO()
>>> f.write('hello')
5
>>> f.write(' ')
1
>>> f.write('world!')
6
>>> print(f.getvalue())
hello world!
getvalue()�������ڻ��д����str��

Ҫ��ȡStringIO��������һ��str��ʼ��StringIO��Ȼ������ļ�һ����ȡ��

>>> from io import StringIO
>>> f = StringIO('Hello!\nHi!\nGoodbye!')
>>> while True:
...     s = f.readline()
...     if s == '':
...         break
...     print(s.strip())
...
Hello!
Hi!
Goodbye!
BytesIO

StringIO������ֻ����str�����Ҫ�������������ݣ�����Ҫʹ��BytesIO��

BytesIOʵ�������ڴ��ж�дbytes�����Ǵ���һ��BytesIO��Ȼ��д��һЩbytes��

>>> from io import BytesIO
>>> f = BytesIO()
>>> f.write('����'.encode('utf-8'))
6
>>> print(f.getvalue())
b'\xe4\xb8\xad\xe6\x96\x87'
��ע�⣬д��Ĳ���str�����Ǿ���UTF-8�����bytes��

��StringIO���ƣ�������һ��bytes��ʼ��BytesIO��Ȼ������ļ�һ����ȡ��

>>> from io import BytesIO
>>> f = BytesIO(b'\xe4\xb8\xad\xe6\x96\x87')
>>> f.read()
b'\xe4\xb8\xad\xe6\x96\x87'
С��

StringIO��BytesIO�����ڴ��в���str��bytes�ķ�����ʹ�úͶ�д�ļ�����һ�µĽӿڡ�
#-----------------------------------------------------------------------------------------
#�����ļ���Ŀ¼
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431925324119bac1bc7979664b4fa9843c0e5fcdcf1e000

import os

if __name__ == "__main__":
    print(os.name)
    #print(os.uname())#AttributeError: module 'os' has no attribute 'uname'
    print(os.environ)
    print(os.environ.get("JAVA_HOME"))
�����
nt
environ({'ALLUSERSPROFILE': 'C:\\ProgramData', 'APPDATA': 'C:\\Users\\Ben\\AppData\\Roaming', 'CATALINA_HOME': 'e:\\Program Files\\Apache Software Foundation\\Tomcat 9.0', 'CLASSPATH': '.;e:\\Program Files\\Java\\jdk1.8.0_112\\lib;e:\\Program Files\\Java\\jdk1.8.0_112\\lib\\tools.jar', 'COMMANDER_DRIVE': 'E:', 'COMMANDER_EXE': 'E:\\Program Files\\TotalCMD64\\Totalcmd64.exe', 'COMMANDER_INI': 'E:\\Program Files\\TotalCMD64\\wincmd.ini', 'COMMANDER_PATH': 'E:\\Program Files\\TotalCMD64', 'COMMONPROGRAMFILES': 'C:\\Program Files\\Common Files', 'COMMONPROGRAMFILES(X86)': 'C:\\Program Files (x86)\\Common Files', 'COMMONPROGRAMW6432': 'C:\\Program Files\\Common Files', 'COMPUTERNAME': 'DESKTOP-18O9S0P', 'COMSPEC': 'C:\\WINDOWS\\system32\\cmd.exe', 'FPS_BROWSER_APP_PROFILE_STRING': 'Internet Explorer', 'FPS_BROWSER_USER_PROFILE_STRING': 'Default', 'HOMEDRIVE': 'C:', 'HOMEPATH': '\\Users\\Ben', 'JAVA_HOME': 'e:\\Program Files\\Java\\jdk1.8.0_112', 'LNKENV': 'C:\\Program Files (x86)\\Internet Explorer\\IEXPLORE.EXE', 'LOCALAPPDATA': 'C:\\Users\\Ben\\AppData\\Local', 'LOGONSERVER': '\\\\DESKTOP-18O9S0P', 'MOZ_PLUGIN_PATH': 'C:\\Program Files (x86)\\Foxit Software\\Foxit Reader Plus\\plugins\\', 'NUMBER_OF_PROCESSORS': '8', 'ONEDRIVE': 'C:\\Users\\Ben\\OneDrive', 'OS': 'Windows_NT', 'PATH': 'D:\\Anaconda3\\Library\\bin;C:\\Program Files (x86)\\Intel\\iCLS Client\\;C:\\Program Files\\Intel\\iCLS Client\\;C:\\Windows\\system32;C:\\Windows;C:\\Windows\\System32\\Wbem;C:\\Windows\\System32\\WindowsPowerShell\\v1.0\\;C:\\Program Files (x86)\\Intel\\Intel(R) Management Engine Components\\DAL;C:\\Program Files\\Intel\\Intel(R) Management Engine Components\\DAL;C:\\Program Files (x86)\\Intel\\Intel(R) Management Engine Components\\IPT;C:\\Program Files\\Intel\\Intel(R) Management Engine Components\\IPT;C:\\Program Files (x86)\\NVIDIA Corporation\\PhysX\\Common;e:\\Program Files\\QuickStart;C:\\Program Files\\Microsoft SQL Server\\110\\Tools\\Binn\\;C:\\WINDOWS\\system32;C:\\WINDOWS;C:\\WINDOWS\\System32\\Wbem;C:\\WINDOWS\\System32\\WindowsPowerShell\\v1.0\\;e:\\Program Files\\Java\\jdk1.8.0_112\\bin;e:\\Program Files\\Java\\jdk1.8.0_112\\jre\\bin;C:\\Program Files (x86)\\Windows Kits\\8.1\\Windows Performance Toolkit\\;C:\\Program Files\\Microsoft SQL Server\\120\\Tools\\Binn\\;d:\\Anaconda3\\;C:\\Users\\Ben\\AppData\\Local\\Microsoft\\WindowsApps;', 'PATHEXT': '.COM;.EXE;.BAT;.CMD;.VBS;.VBE;.JS;.JSE;.WSF;.WSH;.MSC', 'PROCESSOR_ARCHITECTURE': 'AMD64', 'PROCESSOR_IDENTIFIER': 'Intel64 Family 6 Model 94 Stepping 3, GenuineIntel', 'PROCESSOR_LEVEL': '6', 'PROCESSOR_REVISION': '5e03', 'PROGRAMDATA': 'C:\\ProgramData', 'PROGRAMFILES': 'C:\\Program Files', 'PROGRAMFILES(X86)': 'C:\\Program Files (x86)', 'PROGRAMW6432': 'C:\\Program Files', 'PSMODULEPATH': 'C:\\WINDOWS\\system32\\WindowsPowerShell\\v1.0\\Modules\\;C:\\Program Files\\Intel\\', 'PUBLIC': 'C:\\Users\\Public', 'PYCHARM_HOSTED': '1', 'PYTHONIOENCODING': 'UTF-8', 'PYTHONPATH': 'E:\\Program Files\\JetBrains\\PythonProject\\Py3TestProject', 'PYTHONUNBUFFERED': '1', 'SESSIONNAME': 'Console', 'SYSTEMDRIVE': 'C:', 'SYSTEMROOT': 'C:\\WINDOWS', 'TEMP': 'C:\\Users\\Ben\\AppData\\Local\\Temp', 'TMP': 'C:\\Users\\Ben\\AppData\\Local\\Temp', 'USERDOMAIN': 'DESKTOP-18O9S0P', 'USERDOMAIN_ROAMINGPROFILE': 'DESKTOP-18O9S0P', 'USERNAME': 'Ben', 'USERPROFILE': 'C:\\Users\\Ben', 'VS120COMNTOOLS': 'D:\\Program Files\\Microsoft Visual Studio 12.0\\Common7\\Tools\\', 'VS140COMNTOOLS': 'D:\\Program Files\\Microsoft Visual Studio 14.0\\Common7\\Tools\\', 'WINDIR': 'C:\\WINDOWS'})
e:\Program Files\Java\jdk1.8.0_112

�����ļ���Ŀ¼
�����ļ���Ŀ¼�ĺ���һ���ַ���osģ���У�һ���ַ���os.pathģ���У���һ��Ҫע��һ�¡��鿴��������ɾ��Ŀ¼������ô���ã�

# �鿴��ǰĿ¼�ľ���·��:
>>> os.path.abspath('.')
'/Users/michael'
# ��ĳ��Ŀ¼�´���һ����Ŀ¼�����Ȱ���Ŀ¼������·����ʾ����:
>>> os.path.join('/Users/michael', 'testdir')
'/Users/michael/testdir'
# Ȼ�󴴽�һ��Ŀ¼:
>>> os.mkdir('/Users/michael/testdir')
# ɾ��һ��Ŀ¼:
>>> os.rmdir('/Users/michael/testdir')
������·���ϳ�һ��ʱ����Ҫֱ��ƴ�ַ�������Ҫͨ��os.path.join()����������������ȷ����ͬ����ϵͳ��·���ָ�������Linux/Unix/Mac�£�os.path.join()�����������ַ�����

part-1/part-2
��Windows�»᷵���������ַ�����

part-1\part-2
ͬ���ĵ���Ҫ���·��ʱ��Ҳ��Ҫֱ��ȥ���ַ�������Ҫͨ��os.path.split()�������������԰�һ��·�����Ϊ�����֣���һ����������󼶱��Ŀ¼���ļ�����

>>> os.path.split('/Users/michael/testdir/file.txt')
('/Users/michael/testdir', 'file.txt')
os.path.splitext()����ֱ������õ��ļ���չ�����ܶ�ʱ��ǳ����㣺

>>> os.path.splitext('/path/to/file.txt')
('/path/to/file', '.txt')

import os

if __name__ == "__main__":
    txt = "abc/def/opq.txt"
    print(os.path.split(txt))
    print(os.path.splitext(txt))
    print(os.path.basename(txt))
    print(os.path.dirname(txt))
�����
('abc/def', 'opq.txt')
('abc/def/opq', '.txt')
opq.txt
abc/def

��Щ�ϲ������·���ĺ�������Ҫ��Ŀ¼���ļ�Ҫ��ʵ���ڣ�����ֻ���ַ������в�����
�ļ�����ʹ������ĺ������ٶ���ǰĿ¼����һ��test.txt�ļ���
# ���ļ�������:
>>> os.rename('test.txt', 'test.py')
# ɾ���ļ�:
>>> os.remove('test.py')
���Ǹ����ļ��ĺ�����Ȼ��osģ���в����ڣ�ԭ���Ǹ����ļ������ɲ���ϵͳ�ṩ��ϵͳ���á������Ͻ�������ͨ����һ�ڵĶ�д�ļ���������ļ����ƣ�ֻ����Ҫ��д�ܶ���롣

���˵���shutilģ���ṩ��copyfile()�ĺ������㻹������shutilģ�����ҵ��ܶ�ʵ�ú��������ǿ��Կ�����osģ��Ĳ��䡣

��󿴿��������Python�������������ļ�����������Ҫ�г���ǰĿ¼�µ�����Ŀ¼��ֻ��Ҫһ�д��룺

>>> [x for x in os.listdir('.') if os.path.isdir(x)]
['.lein', '.local', '.m2', '.npm', '.ssh', '.Trash', '.vim', 'Applications', 'Desktop', ...]

Ҫ�г����е�.py�ļ���Ҳֻ��һ�д��룺
>>> [x for x in os.listdir('.') if os.path.isfile(x) and os.path.splitext(x)[1]=='.py']
['apis.py', 'config.py', 'models.py', 'pymonitor.py', 'test_db.py', 'urls.py', 'wsgiapp.py']
#-----------------------------------------------------------------------------------------
#���л�
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143192607210600a668b5112e4a979dd20e4661cc9c97000
�ڳ������еĹ����У����еı����������ڴ��У����磬����һ��dict��

d = dict(name='Bob', age=20, score=88)
������ʱ�޸ı����������name�ĳ�'Bill'������һ�����������������ռ�õ��ڴ�ͱ�����ϵͳȫ�����ա����û�а��޸ĺ��'Bill'�洢�������ϣ��´��������г��򣬱����ֱ���ʼ��Ϊ'Bob'��

���ǰѱ������ڴ��б�ɿɴ洢����Ĺ��̳�֮Ϊ���л�����Python�н�pickling��������������Ҳ����֮Ϊserialization��marshalling��flattening�ȵȣ�����һ����˼��

���л�֮�󣬾Ϳ��԰����л��������д����̣�����ͨ�����紫�䵽��Ļ����ϡ�

���������ѱ������ݴ����л��Ķ������¶����ڴ����֮Ϊ�����л�����unpickling��

Python�ṩ��pickleģ����ʵ�����л���

���ȣ����ǳ��԰�һ���������л���д���ļ���

>>> import pickle
>>> d = dict(name='Bob', age=20, score=88)
>>> pickle.dumps(d)
b'\x80\x03}q\x00(X\x03\x00\x00\x00ageq\x01K\x14X\x05\x00\x00\x00scoreq\x02KXX\x04\x00\x00\x00nameq\x03X\x03\x00\x00\x00Bobq\x04u.'
pickle.dumps()����������������л���һ��bytes��Ȼ�󣬾Ϳ��԰����bytesд���ļ�����������һ������pickle.dump()ֱ�ӰѶ������л���д��һ��file-like Object��

>>> f = open('dump.txt', 'wb')
>>> pickle.dump(d, f)
>>> f.close()
����д���dump.txt�ļ���һ�����߰�������ݣ���Щ����Python����Ķ����ڲ���Ϣ��

������Ҫ�Ѷ���Ӵ��̶����ڴ�ʱ�������Ȱ����ݶ���һ��bytes��Ȼ����pickle.loads()���������л�������Ҳ����ֱ����pickle.load()������һ��file-like Object��ֱ�ӷ����л����������Ǵ���һ��Python�������������л��ղű���Ķ���

>>> f = open('dump.txt', 'rb')
>>> d = pickle.load(f)
>>> f.close()
>>> d
{'age': 20, 'score': 88, 'name': 'Bob'}
�����������ֻ����ˣ�

��Ȼ�����������ԭ���ı�������ȫ����ɵĶ�������ֻ��������ͬ���ѡ�

Pickle�������������������������е����л�����һ����������ֻ������Python�����ҿ��ܲ�ͬ�汾��Python�˴˶������ݣ���ˣ�ֻ����Pickle������Щ����Ҫ�����ݣ����ܳɹ��ط����л�Ҳû��ϵ��

JSON

�������Ҫ�ڲ�ͬ�ı������֮�䴫�ݶ��󣬾ͱ���Ѷ������л�Ϊ��׼��ʽ������XML�������õķ��������л�ΪJSON����ΪJSON��ʾ��������һ���ַ��������Ա��������Զ�ȡ��Ҳ���Է���ش洢�����̻���ͨ�����紫�䡣JSON�����Ǳ�׼��ʽ�����ұ�XML���죬���ҿ���ֱ����Webҳ���ж�ȡ���ǳ����㡣
JSON��ʾ�Ķ�����Ǳ�׼��JavaScript���ԵĶ���JSON��Python���õ��������Ͷ�Ӧ���£�
JSON����	Python����
{}	dict
[]	list
"string"	str
1234.56	int��float
true/false	True/False
null	None
Python���õ�jsonģ���ṩ�˷ǳ����Ƶ�Python����JSON��ʽ��ת���������ȿ�����ΰ�Python������һ��JSON��

>>> import json
>>> d = dict(name='Bob', age=20, score=88)
>>> json.dumps(d)
'{"age": 20, "score": 88, "name": "Bob"}'
dumps()��������һ��str�����ݾ��Ǳ�׼��JSON�����Ƶģ�dump()��������ֱ�Ӱ�JSONд��һ��file-like Object��

Ҫ��JSON�����л�ΪPython������loads()���߶�Ӧ��load()������ǰ�߰�JSON���ַ��������л������ߴ�file-like Object�ж�ȡ�ַ����������л���

>>> json_str = '{"age": 20, "score": 88, "name": "Bob"}'
>>> json.loads(json_str)
{'age': 20, 'score': 88, 'name': 'Bob'}
����JSON��׼�涨JSON������UTF-8������������������ȷ����Python��str��JSON���ַ���֮��ת����

JSON����

Python��dict�������ֱ�����л�ΪJSON��{}���������ܶ�ʱ�����Ǹ�ϲ����class��ʾ���󣬱��綨��Student�࣬Ȼ�����л���

import json

class Student(object):
    def __init__(self, name, age, score):
        self.name = name
        self.age = age
        self.score = score

s = Student('Bob', 20, 88)
print(json.dumps(s))
���д��룬��������صõ�һ��TypeError��

Traceback (most recent call last):
  ...
TypeError: <__main__.Student object at 0x10603cc50> is not JSON serializable
�����ԭ����Student������һ�������л�ΪJSON�Ķ���

�����class��ʵ�������޷����л�ΪJSON����϶�������

�𼱣�������ϸ����dumps()�����Ĳ����б����Է��֣����˵�һ�������obj�����⣬dumps()�������ṩ��һ��ѵĿ�ѡ������

https://docs.python.org/3/library/json.html#json.dumps

��Щ��ѡ��������������������JSON���л���ǰ��Ĵ���֮�����޷���Student��ʵ�����л�ΪJSON������ΪĬ������£�dumps()������֪����ν�Studentʵ����Ϊһ��JSON��{}����

��ѡ����default���ǰ�����һ��������һ��������ΪJSON�Ķ�������ֻ��ҪΪStudentר��дһ��ת���������ٰѺ�������ȥ���ɣ�

def student2dict(std):
    return {
        'name': std.name,
        'age': std.age,
        'score': std.score
    }
������Studentʵ�����ȱ�student2dict()����ת����dict��Ȼ���ٱ�˳�����л�ΪJSON��

>>> print(json.dumps(s, default=student2dict))
{"age": 20, "name": "Bob", "score": 88}
�������´��������һ��Teacher���ʵ���������޷����л�ΪJSON�����ǿ���͵������������class��ʵ����Ϊdict��

print(json.dumps(s, default=lambda obj: obj.__dict__))
��Ϊͨ��class��ʵ������һ��__dict__���ԣ�������һ��dict�������洢ʵ��������Ҳ���������⣬���綨����__slots__��class��

ͬ���ĵ����������Ҫ��JSON�����л�Ϊһ��Student����ʵ����loads()��������ת����һ��dict����Ȼ�����Ǵ����object_hook���������dictת��ΪStudentʵ����

def dict2student(d):
    return Student(d['name'], d['age'], d['score'])
���н�����£�

>>> json_str = '{"age": 20, "score": 88, "name": "Bob"}'
>>> print(json.loads(json_str, object_hook=dict2student))
<__main__.Student object at 0x10cd3c190>
��ӡ�����Ƿ����л���Studentʵ������


if __name__ == "__main__":
    obj = dict(name='С��', age=20)
    s = json.dumps(obj, ensure_ascii=True)
    print(s)
#-----------------------------------------------------------------------------------------
#�����
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431927781401bb47ccf187b24c3b955157bb12c5882d000

Ҫ��Python����ʵ�ֶ���̣�multiprocessing�����������˽����ϵͳ�����֪ʶ��

Unix/Linux����ϵͳ�ṩ��һ��fork()ϵͳ���ã����ǳ����⡣��ͨ�ĺ������ã�����һ�Σ�����һ�Σ�����fork()����һ�Σ��������Σ���Ϊ����ϵͳ�Զ��ѵ�ǰ���̣���Ϊ�����̣�������һ�ݣ���Ϊ�ӽ��̣���Ȼ�󣬷ֱ��ڸ����̺��ӽ����ڷ��ء�

�ӽ�����Զ����0���������̷����ӽ��̵�ID���������������ǣ�һ�������̿���fork���ܶ��ӽ��̣����ԣ�������Ҫ����ÿ���ӽ��̵�ID�����ӽ���ֻ��Ҫ����getppid()�Ϳ����õ������̵�ID��

Python��osģ���װ�˳�����ϵͳ���ã����оͰ���fork��������Python���������ɴ����ӽ��̣�

import os

print('Process (%s) start...' % os.getpid())
# Only works on Unix/Linux/Mac:
pid = os.fork()
if pid == 0:
    print('I am child process (%s) and my parent is %s.' % (os.getpid(), os.getppid()))
else:
    print('I (%s) just created a child process (%s).' % (os.getpid(), pid))
���н�����£�

Process (876) start...
I (876) just created a child process (877).
I am child process (877) and my parent is 876.
����Windowsû��fork���ã�����Ĵ�����Windows���޷����С�����Macϵͳ�ǻ���BSD��Unix��һ�֣��ںˣ����ԣ���Mac��������û������ģ��Ƽ������MacѧPython��

����fork���ã�һ�������ڽӵ�������ʱ�Ϳ��Ը��Ƴ�һ���ӽ��������������񣬳�����Apache�����������ɸ����̼����˿ڣ�ÿ�����µ�http����ʱ����fork���ӽ����������µ�http����

multiprocessing
���������д����̵ķ������Unix/Linux��������ȷ��ѡ������Windowsû��fork���ã��ѵ���Windows���޷���Python��д����̵ĳ���
����Python�ǿ�ƽ̨�ģ���ȻҲӦ���ṩһ����ƽ̨�Ķ����֧�֡�multiprocessingģ����ǿ�ƽ̨�汾�Ķ����ģ�顣

multiprocessingģ���ṩ��һ��Process��������һ�����̶��������������ʾ������һ���ӽ��̲��ȴ��������
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
from multiprocessing import Process

class RunProc(object):
    def __call__(self, name):
        print('Run child process %s (%s)...' % (name, os.getpid()))

def run_proc(name):
    print('Run child process %s (%s)...' % (name, os.getpid()))

if __name__=='__main__':
    print('Parent process %s.' % os.getpid())
    p = Process(target=run_proc, args=('test',)) #p = Process(target=RunProc(), args=('test',)) дһ����������
    print('Child process will start.')
    p.start()
    p.join() #�������������䣬���ӽ��̺������end��䲻һ��˭�����
    print('Child process end.')
�����
Parent process 8196.
Child process will start.
Run child process test (7296)...
Child process end.

�����ӽ���ʱ��ֻ��Ҫ����һ��ִ�к����ͺ����Ĳ���������һ��Processʵ������start()���������������������̱�fork()��Ҫ�򵥡�
join()�������Եȴ��ӽ��̽������ټ����������У�ͨ�����ڽ��̼��ͬ����

Pool
���Ҫ�����������ӽ��̣������ý��̳صķ�ʽ���������ӽ��̣�
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from multiprocessing import Pool
import os, time, random

def long_time_task(name):
    print('Run task %s (%s)...' % (name, os.getpid()))
    start = time.time()
    time.sleep(4) #random.random() * 3
    end = time.time()
    print('Task %s runs %0.2f seconds.' % (name, (end - start)))

if __name__=='__main__':
    print('Parent process %s.' % os.getpid())
    poolNum = 20
    p = Pool(poolNum) #���̳���������ͬʱ������ж��ٸ�����
    for i in range(poolNum):
        p.apply_async(long_time_task, args=(i,))
    print('Waiting for all subprocesses done...')
    p.close()
    p.join()
    print('All subprocesses done.')
�����
Parent process 8120.
Waiting for all subprocesses done...
Run task 0 (9104)...
Run task 1 (4744)...
Run task 2 (5036)...
Run task 3 (11012)...
Task 0 runs 4.00 seconds.
Task 2 runs 4.00 seconds.
Task 3 runs 4.00 seconds.
Task 1 runs 4.00 seconds.
All subprocesses done.

��������

��Pool�������join()������ȴ������ӽ���ִ����ϣ�����join()֮ǰ�����ȵ���close()������close()֮��Ͳ��ܼ�������µ�Process�ˡ�

��ע������Ľ����task 0��1��2��3������ִ�еģ���task 4Ҫ�ȴ�ǰ��ĳ��task��ɺ��ִ�У�������ΪPool��Ĭ�ϴ�С���ҵĵ�������4����ˣ����ͬʱִ��4�����̡�����Pool������Ƶ����ƣ������ǲ���ϵͳ�����ơ�����ĳɣ�

p = Pool(5)
�Ϳ���ͬʱ��5�����̡�

����Pool��Ĭ�ϴ�С��CPU�ĺ���������㲻��ӵ��8��CPU����Ҫ�ύ����9���ӽ��̲��ܿ�������ĵȴ�Ч���������ﲻ���ر𶮣�ֻҪ�����еĽ���������Pool�����������ܿ����ȴ���Ч����


�ӽ���
�ܶ�ʱ���ӽ��̲�������������һ���ⲿ���̡����Ǵ������ӽ��̺󣬻���Ҫ�����ӽ��̵�����������
subprocessģ����������Ƿǳ����������һ���ӽ��̣�Ȼ�����������������
�����������ʾ�������Python��������������nslookup www.python.org�����������ֱ�����е�Ч����һ���ģ�
import subprocess
if __name__=='__main__':
    print('$ nslookup www.python.org')
    r = subprocess.call(['nslookup', 'www.python.org'])
    print('Exit code:', r)
�����
$ nslookup www.python.org
Server:        192.168.19.4
Address:    192.168.19.4#53

Non-authoritative answer:
www.python.org    canonical name = python.map.fastly.net.
Name:    python.map.fastly.net
Address: 199.27.79.223

Exit code: 0

����ӽ��̻���Ҫ���룬�����ͨ��communicate()�������룺
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import subprocess
if __name__=='__main__':
    print('$ nslookup')
    p = subprocess.Popen(['nslookup'], stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    output, err = p.communicate(b'set q=mx\npython.org\nexit\n')
    print(output.decode('gbk')) #utf-8
    print('Exit code:', p.returncode)
����Ĵ����൱����������ִ������nslookup��Ȼ���ֶ����룺
set q=mx
python.org
exit

�����
$ nslookup
Ĭ�Ϸ�����:  dnspai-public-dns.dnspai.com
Address:  140.207.198.6

> > ������:  dnspai-public-dns.dnspai.com
Address:  140.207.198.6

python.org	MX preference = 50, mail exchanger = mail.python.org
> 
Exit code: 0

���̼�ͨ��
Process֮��϶�����Ҫͨ�ŵģ�����ϵͳ�ṩ�˺ܶ������ʵ�ֽ��̼��ͨ�š�Python��multiprocessingģ���װ�˵ײ�Ļ��ƣ��ṩ��Queue��Pipes�ȶ��ַ�ʽ���������ݡ�
������QueueΪ�����ڸ������д��������ӽ��̣�һ����Queue��д���ݣ�һ����Queue������ݣ�
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from multiprocessing import Process, Queue
import os, time, random

# д���ݽ���ִ�еĴ���:
def write(q):
    print('Process to write: %s' % os.getpid())
    for value in ['A', 'B', 'C']:
        print('Put %s to queue...' % value)
        q.put(value)
        time.sleep(3)#time.sleep(random.random())

# �����ݽ���ִ�еĴ���:
def read(q):
    print('Process to read: %s' % os.getpid())
    while True:
        value = q.get(True)
        print('Get %s from queue.' % value)

if __name__=='__main__':
    # �����̴���Queue�������������ӽ��̣�
    q = Queue()
    pw = Process(target=write, args=(q,))
    pr = Process(target=read, args=(q,))
    # �����ӽ���pw��д��:
    pw.start()
    # �����ӽ���pr����ȡ:
    pr.start()
    # �ȴ�pw����:
    pw.join()
    # pr����������ѭ�����޷��ȴ��������ֻ��ǿ����ֹ:
    pr.terminate()
�����
Process to write: 12176
Put A to queue...
Process to read: 5216
Get A from queue.
Put B to queue...
Get B from queue.
Put C to queue...
Get C from queue.

��Unix/Linux�£�multiprocessingģ���װ��fork()���ã�ʹ���ǲ���Ҫ��עfork()��ϸ�ڡ�����Windowsû��fork���ã���ˣ�multiprocessing��Ҫ��ģ�⡱��fork��Ч��������������Python���󶼱���ͨ��pickle���л��ٴ����ӽ���ȥ�����У����multiprocessing��Windows�µ���ʧ���ˣ�Ҫ�ȿ����ǲ���pickleʧ���ˡ�

С��
��Unix/Linux�£�����ʹ��fork()����ʵ�ֶ���̡�
Ҫʵ�ֿ�ƽ̨�Ķ���̣�����ʹ��multiprocessingģ�顣
���̼�ͨ����ͨ��Queue��Pipes��ʵ�ֵġ�
#-----------------------------------------------------------------------------------------
#���߳�
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143192823818768cd506abbc94eb5916192364506fa5d000
����������ɶ������ɣ�Ҳ������һ�������ڵĶ��߳���ɡ�

����ǰ���ᵽ�˽������������߳���ɵģ�һ������������һ���̡߳�

�����߳��ǲ���ϵͳֱ��֧�ֵ�ִ�е�Ԫ����ˣ��߼�����ͨ�������ö��̵߳�֧�֣�PythonҲ�����⣬���ң�Python���߳���������Posix Thread��������ģ��������̡߳�

Python�ı�׼���ṩ������ģ�飺_thread��threading��_thread�ǵͼ�ģ�飬threading�Ǹ߼�ģ�飬��_thread�����˷�װ�������������£�����ֻ��Ҫʹ��threading����߼�ģ�顣

����һ���߳̾��ǰ�һ���������벢����Threadʵ����Ȼ�����start()��ʼִ�У�

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import time, threading

# ���߳�ִ�еĴ���:
def loop():
    print('thread %s is running...' % threading.current_thread().name)
    n = 0
    while n < 5:
        n = n + 1
        print('thread %s >>> %s' % (threading.current_thread().name, n))
        time.sleep(1)
    print('thread %s ended.' % threading.current_thread().name)

if __name__=='__main__':
    print('thread %s is running...' % threading.current_thread().name)
    t = threading.Thread(target=loop, name='LoopThread')
    t.start()
    t.join()
    print('thread %s ended.' % threading.current_thread().name)
�����
thread MainThread is running...
thread LoopThread is running...
thread LoopThread >>> 1
thread LoopThread >>> 2
thread LoopThread >>> 3
thread LoopThread >>> 4
thread LoopThread >>> 5
thread LoopThread ended.
thread MainThread ended.

#�������name���������߳̾Ͱ�˳������������֮ǰ���߳��Ƿ����
t = threading.Thread(target=loop)
t2 = threading.Thread(target=loop)
t.start()
time.sleep(10)
t2.start()
t.join()
t2.join()
�����
thread MainThread is running...
thread Thread-1 is running...
thread Thread-1 >>> 1
thread Thread-1 >>> 2
thread Thread-1 >>> 3
thread Thread-1 >>> 4
thread Thread-1 >>> 5
thread Thread-1 ended.
thread Thread-2 is running...
thread Thread-2 >>> 1
thread Thread-2 >>> 2
thread Thread-2 >>> 3
thread Thread-2 >>> 4
thread Thread-2 >>> 5
thread Thread-2 ended.
thread MainThread ended.

�����κν���Ĭ�Ͼͻ�����һ���̣߳����ǰѸ��̳߳�Ϊ���̣߳����߳��ֿ��������µ��̣߳�Python��threadingģ���и�current_thread()����������Զ���ص�ǰ�̵߳�ʵ�������߳�ʵ�������ֽ�MainThread�����̵߳������ڴ���ʱָ����������LoopThread�������̡߳����ֽ����ڴ�ӡʱ������ʾ����ȫû���������壬�����������Python���Զ����߳�����ΪThread-1��Thread-2����

Lock
���̺߳Ͷ�������Ĳ�ͬ���ڣ�������У�ͬһ��������������һ�ݿ���������ÿ�������У�����Ӱ�죬�����߳��У����б������������̹߳������ԣ��κ�һ�����������Ա��κ�һ���߳��޸ģ���ˣ��߳�֮�乲����������Σ�����ڶ���߳�ͬʱ��һ�������������ݸ������ˡ�

����������߳�ͬʱ����һ��������ô�����ݸ������ˣ�
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import threading

balance = 0

def change_it(n):
    # �ȴ��ȡ�����Ӧ��Ϊ0:
    global balance
    balance = balance + n
    balance = balance - n

def run_thread(n):
    for i in range(100000):
        change_it(n)

if __name__=='__main__':
    t1 = threading.Thread(target=run_thread, args=(5,))
    t2 = threading.Thread(target=run_thread, args=(8,))
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    print(balance)
�����
��ʱ��Ϊ0����ʱΪ5

����ԭ������Ϊ�޸�balance��Ҫ������䣬��ִ���⼸�����ʱ���߳̿����жϣ��Ӷ����¶���̰߳�ͬһ����������ݸ����ˡ�

�����߳�ͬʱһ��һȡ���Ϳ��ܵ������ԣ���϶���ϣ��������д��Ī������ر���˸��������ԣ����Ǳ���ȷ��һ���߳����޸�balance��ʱ�򣬱���߳�һ�����ܸġ�

�������Ҫȷ��balance������ȷ����Ҫ��change_it()��һ��������ĳ���߳̿�ʼִ��change_it()ʱ������˵�����߳���Ϊ�����������������̲߳���ͬʱִ��change_it()��ֻ�ܵȴ���ֱ�������ͷź󣬻�ø����Ժ���ܸġ�������ֻ��һ�������۶����̣߳�ͬһʱ�����ֻ��һ���̳߳��и��������ԣ���������޸ĵĳ�ͻ������һ��������ͨ��threading.Lock()��ʵ�֣�
balance = 0
lock = threading.Lock()

def run_thread(n):
    for i in range(100000):
        # ��Ҫ��ȡ��:
        lock.acquire()
        try:
            # ���ĵظİ�:
            change_it(n)
        finally:
            # ������һ��Ҫ�ͷ���:
            lock.release()
������߳�ͬʱִ��lock.acquire()ʱ��ֻ��һ���߳��ܳɹ��ػ�ȡ����Ȼ�����ִ�д��룬�����߳̾ͼ����ȴ�ֱ�������Ϊֹ��
��������߳������һ��Ҫ�ͷ�����������Щ���ȴ������߳̽���Զ�ȴ���ȥ����Ϊ���̡߳�����������try...finally��ȷ����һ���ᱻ�ͷš�
���ĺô�����ȷ����ĳ�ιؼ�����ֻ����һ���̴߳�ͷ��β������ִ�У�������ȻҲ�ܶ࣬��������ֹ�˶��̲߳���ִ�У���������ĳ�δ���ʵ����ֻ���Ե��߳�ģʽִ�У�Ч�ʾʹ����½��ˡ���Σ����ڿ��Դ��ڶ��������ͬ���̳߳��в�ͬ����������ͼ��ȡ�Է����е���ʱ�����ܻ�������������¶���߳�ȫ�����𣬼Ȳ���ִ�У�Ҳ�޷�������ֻ�ܿ�����ϵͳǿ����ֹ��
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import threading, multiprocessing

def loop():
    x = 0
    while True:
        x = x ^ 1

if __name__=='__main__':
    cpu_count = multiprocessing.cpu_count()
    for i in range(cpu_count):
        t = threading.Thread(target=loop)
        t.start()
������4��8�̣߳�ֻ������������ռ��������

������CPU����������ͬ��N���̣߳���4��CPU�Ͽ��Լ�ص�CPUռ���ʽ���102%��Ҳ���ǽ�ʹ����һ�ˡ�
������C��C++��Java����д��ͬ����ѭ����ֱ�ӿ��԰�ȫ������������4�˾��ܵ�400%��8�˾��ܵ�800%��ΪʲôPython�����أ�
��ΪPython���߳���Ȼ���������̣߳���������ִ�д���ʱ����һ��GIL����Global Interpreter Lock���κ�Python�߳�ִ��ǰ�������Ȼ��GIL����Ȼ��ÿִ��100���ֽ��룬���������Զ��ͷ�GIL�����ñ���߳��л���ִ�С����GILȫ����ʵ���ϰ������̵߳�ִ�д��붼�������������ԣ����߳���Python��ֻ�ܽ���ִ�У���ʹ100���߳�����100��CPU�ϣ�Ҳֻ���õ�1���ˡ�

GIL��Python��������Ƶ���ʷ�������⣬ͨ�������õĽ������ǹٷ�ʵ�ֵ�CPython��Ҫ�������ö�ˣ�������дһ������GIL�Ľ�������

���ԣ���Python�У�����ʹ�ö��̣߳�����Ҫָ������Ч���ö�ˡ����һ��Ҫͨ�����߳����ö�ˣ���ֻ��ͨ��C��չ��ʵ�֣�����������ʧȥ��Python�����õ��ص㡣

������Ҳ���ù��ڵ��ģ�Python��Ȼ�������ö��߳�ʵ�ֶ�����񣬵�����ͨ�������ʵ�ֶ�����񡣶��Python�����и��Զ�����GIL��������Ӱ�졣
#-----------------------------------------------------------------------------------------
#ThreadLocal
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431928972981094a382e5584413fa040b46d46cce48e000

�ڶ��̻߳����£�ÿ���̶߳����Լ������ݡ�һ���߳�ʹ���Լ��ľֲ�������ʹ��ȫ�ֱ����ã���Ϊ�ֲ�����ֻ���߳��Լ��ܿ���������Ӱ�������̣߳���ȫ�ֱ������޸ı��������

���Ǿֲ�����Ҳ�����⣬�����ں������õ�ʱ�򣬴����������鷳��

def process_student(name):
    std = Student(name)
    # std�Ǿֲ�����������ÿ��������Ҫ��������˱��봫��ȥ��
    do_task_1(std)
    do_task_2(std)

def do_task_1(std):
    do_subtask_1(std)
    do_subtask_2(std)

def do_task_2(std):
    do_subtask_2(std)
    do_subtask_2(std)
ÿ������һ��һ����ö���ô�������ǻ����ˣ���ȫ�ֱ�����Ҳ���У���Ϊÿ���̴߳���ͬ��Student���󣬲��ܹ���

�����һ��ȫ��dict������е�Student����Ȼ����thread������Ϊkey����̶߳�Ӧ��Student������Σ�

global_dict = {}

def std_thread(name):
    std = Student(name)
    # ��std�ŵ�ȫ�ֱ���global_dict�У�
    global_dict[threading.current_thread()] = std
    do_task_1()
    do_task_2()

def do_task_1():
    # ������std�����Ǹ��ݵ�ǰ�̲߳��ң�
    std = global_dict[threading.current_thread()]
    ...

def do_task_2():
    # �κκ��������Բ��ҳ���ǰ�̵߳�std������
    std = global_dict[threading.current_thread()]
    ...
���ַ�ʽ�������ǿ��еģ��������ŵ���������std������ÿ�㺯���еĴ������⣬���ǣ�ÿ��������ȡstd�Ĵ����е��

��û�и��򵥵ķ�ʽ��

ThreadLocalӦ�˶��������ò���dict��ThreadLocal�����Զ�������£�
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import threading

# ����ȫ��ThreadLocal����:
local_school = threading.local()

def process_student():
    # ��ȡ��ǰ�̹߳�����student:
    std = local_school.student
    print('Hello, %s (in %s)' % (std, threading.current_thread().name))


def process_thread(name):
    # ��ThreadLocal��student:
    local_school.student = name
    process_student()

if __name__=='__main__':
    t1 = threading.Thread(target=process_thread, args=('Alice',), name='Thread-A')
    t2 = threading.Thread(target=process_thread, args=('Bob',), name='Thread-B')
    t1.start()
    t2.start()
    t1.join()
    t2.join()
�����
Hello, Alice (in Thread-A)
Hello, Bob (in Thread-B)

ȫ�ֱ���local_school����һ��ThreadLocal����ÿ��Thread���������Զ�дstudent���ԣ�������Ӱ�졣����԰�local_school����ȫ�ֱ�������ÿ��������local_school.student�����̵߳ľֲ����������������д���������ţ�Ҳ���ù����������⣬ThreadLocal�ڲ��ᴦ��

�������Ϊȫ�ֱ���local_school��һ��dict������������local_school.student�������԰�������������local_school.teacher�ȵȡ�

ThreadLocal��õĵط�����Ϊÿ���̰߳�һ�����ݿ����ӣ�HTTP�����û������Ϣ�ȣ�����һ���̵߳����е��õ��Ĵ����������Էǳ�����ط�����Щ��Դ��

С��

һ��ThreadLocal������Ȼ��ȫ�ֱ�������ÿ���̶߳�ֻ�ܶ�д�Լ��̵߳Ķ����������������š�ThreadLocal����˲�����һ���߳��и�������֮�以�ഫ�ݵ����⡣
#-----------------------------------------------------------------------------------------
#���� vs. �߳�
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014319292979766bd3285c9d6b4942a8ea9b4e9cfb48d8000
���ǽ����˶���̺Ͷ��̣߳�����ʵ�ֶ�������õ����ַ�ʽ�����ڣ�����������һ�������ַ�ʽ����ȱ�㡣

���ȣ�Ҫʵ�ֶ�����ͨ�����ǻ����Master-Workerģʽ��Master�����������Worker����ִ��������ˣ������񻷾��£�ͨ����һ��Master�����Worker��

����ö����ʵ��Master-Worker�������̾���Master���������̾���Worker��

����ö��߳�ʵ��Master-Worker�����߳̾���Master�������߳̾���Worker��

�����ģʽ�����ŵ�����ȶ��Ըߣ���Ϊһ���ӽ��̱����ˣ�����Ӱ�������̺������ӽ��̡�����Ȼ�����̹������н��̾�ȫ���ˣ�����Master����ֻ����������񣬹ҵ��ĸ��ʵͣ�������Apache������ǲ��ö����ģʽ��

�����ģʽ��ȱ���Ǵ������̵Ĵ��۴���Unix/Linuxϵͳ�£���fork���û��У���Windows�´������̿����޴����⣬����ϵͳ��ͬʱ���еĽ�����Ҳ�����޵ģ����ڴ��CPU�������£�����м�ǧ������ͬʱ���У�����ϵͳ�����ȶ�������⡣

���߳�ģʽͨ���ȶ���̿�һ�㣬����Ҳ�첻����ȥ�����ң����߳�ģʽ������ȱ������κ�һ���̹߳ҵ�������ֱ������������̱�������Ϊ�����̹߳�����̵��ڴ档��Windows�ϣ����һ���߳�ִ�еĴ���������⣬�㾭�����Կ�����������ʾ�����ó���ִ���˷Ƿ������������رա�����ʵ������ĳ���̳߳������⣬���ǲ���ϵͳ��ǿ�ƽ����������̡�

��Windows�£����̵߳�Ч�ʱȶ����Ҫ�ߣ�����΢���IIS������Ĭ�ϲ��ö��߳�ģʽ�����ڶ��̴߳����ȶ��Ե����⣬IIS���ȶ��ԾͲ���Apache��Ϊ�˻���������⣬IIS��Apache�������ж����+���̵߳Ļ��ģʽ�����ǰ�����Խ��Խ���ӡ�

�߳��л�

�����Ƕ���̻��Ƕ��̣߳�ֻҪ����һ�࣬Ч�ʿ϶��ϲ�ȥ��Ϊʲô�أ�

���Ǵ���ȷ��������㲻������׼���п���ÿ��������Ҫ�����ġ���ѧ��Ӣ�������ѧ��5�Ƶ���ҵ��ÿ����ҵ��ʱ1Сʱ��

������Ȼ�1Сʱ��������ҵ�������ˣ��ٻ�1Сʱ����ѧ��ҵ������������ȫ�����꣬һ����5Сʱ�����ַ�ʽ��Ϊ������ģ�ͣ���������������ģ�͡�

����������л���������ģ�ͣ���������1�������ģ����л�����ѧ��ҵ����1���ӣ����л���Ӣ��Դ����ƣ�ֻҪ�л��ٶ��㹻�죬���ַ�ʽ�ͺ͵���CPUִ�ж�������һ�����ˣ����׶�԰С���ѵ��۹��������������ͬʱд5����ҵ��

���ǣ��л���ҵ���д��۵ģ�����������е���ѧ��Ҫ����ʰ�����ϵ������鱾���ֱʣ���б����ֳ�����Ȼ�󣬴���ѧ�α����ҳ�Բ��ֱ�ߣ����׼���»����������ܿ�ʼ����ѧ��ҵ������ϵͳ���л����̻����߳�ʱҲ��һ���ģ�����Ҫ�ȱ��浱ǰִ�е��ֳ�������CPU�Ĵ���״̬���ڴ�ҳ�ȣ���Ȼ�󣬰��������ִ�л���׼���ã��ָ��ϴεļĴ���״̬���л��ڴ�ҳ�ȣ������ܿ�ʼִ�С�����л�������Ȼ�ܿ죬����Ҳ��Ҫ�ķ�ʱ�䡣����м�ǧ������ͬʱ���У�����ϵͳ���ܾ���Ҫæ���л����񣬸���û�ж���ʱ��ȥִ�������ˣ������������ľ���Ӳ�̿��죬�㴰���޷�Ӧ��ϵͳ���ڼ���״̬��

���ԣ�������һ���ൽһ���޶ȣ��ͻ����ĵ�ϵͳ���е���Դ�����Ч�ʼ����½����������������á�

�����ܼ��� vs. IO�ܼ���

�Ƿ���ö�����ĵڶ�����������������͡����ǿ��԰������Ϊ�����ܼ��ͺ�IO�ܼ��͡�

�����ܼ���������ص���Ҫ���д����ļ��㣬����CPU��Դ���������Բ���ʡ�����Ƶ���и������ȵȣ�ȫ��CPU���������������ּ����ܼ���������ȻҲ�����ö�������ɣ���������Խ�࣬���������л���ʱ���Խ�࣬CPUִ�������Ч�ʾ�Խ�ͣ����ԣ�Ҫ���Ч������CPU�������ܼ�������ͬʱ���е�����Ӧ������CPU�ĺ�������

�����ܼ�������������Ҫ����CPU��Դ����ˣ���������Ч��������Ҫ��Python�����Ľű���������Ч�ʺܵͣ���ȫ���ʺϼ����ܼ������񡣶��ڼ����ܼ������������C���Ա�д��

�ڶ��������������IO�ܼ��ͣ��漰�����硢����IO��������IO�ܼ�����������������ص���CPU���ĺ��٣�����Ĵ󲿷�ʱ�䶼�ڵȴ�IO������ɣ���ΪIO���ٶ�ԶԶ����CPU���ڴ���ٶȣ�������IO�ܼ�����������Խ�࣬CPUЧ��Խ�ߣ���Ҳ��һ���޶ȡ������Ĵ󲿷�������IO�ܼ������񣬱���WebӦ�á�

IO�ܼ�������ִ���ڼ䣬99%��ʱ�䶼����IO�ϣ�����CPU�ϵ�ʱ����٣���ˣ��������ٶȼ����C�����滻��Python���������ٶȼ��͵Ľű����ԣ���ȫ�޷���������Ч�ʡ�����IO�ܼ�����������ʵ����Ծ��ǿ���Ч����ߣ����������٣������ԣ��ű���������ѡ��C������

�첽IO

���ǵ�CPU��IO֮��޴���ٶȲ��죬һ��������ִ�еĹ����д󲿷�ʱ�䶼�ڵȴ�IO�����������̵��߳�ģ�ͻᵼ�±�������޷�����ִ�У���ˣ����ǲ���Ҫ�����ģ�ͻ��߶��߳�ģ����֧�ֶ����񲢷�ִ�С�

�ִ�����ϵͳ��IO�����Ѿ����˾޴�ĸĽ��������ص����֧���첽IO�����������ò���ϵͳ�ṩ���첽IO֧�֣��Ϳ����õ����̵��߳�ģ����ִ�ж���������ȫ�µ�ģ�ͳ�Ϊ�¼�����ģ�ͣ�Nginx����֧���첽IO��Web�����������ڵ���CPU�ϲ��õ�����ģ�;Ϳ��Ը�Ч��֧�ֶ������ڶ��CPU�ϣ��������ж�����̣�������CPU��������ͬ����������ö��CPU������ϵͳ�ܵĽ�������ʮ�����ޣ���˲���ϵͳ���ȷǳ���Ч�����첽IO���ģ����ʵ�ֶ�������һ����Ҫ�����ơ�

��Ӧ��Python���ԣ����̵߳��첽���ģ�ͳ�ΪЭ�̣�����Э�̵�֧�֣��Ϳ��Ի����¼�������д��Ч�Ķ�����������ǻ��ں���������α�дЭ�̡�

#-----------------------------------------------------------------------------------------
#�ֲ�ʽ����
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431929340191970154d52b9d484b88a7b343708fcc60000

��Thread��Process�У�Ӧ����ѡProcess����ΪProcess���ȶ������ң�Process���Էֲ�����̨�����ϣ���Thread���ֻ�ֲܷ���ͬһ̨�����Ķ��CPU�ϡ�

Python��multiprocessingģ�鲻��֧�ֶ���̣�����managers��ģ�黹֧�ְѶ���̷ֲ�����̨�����ϡ�һ��������̿�����Ϊ�����ߣ�������ֲ���������������У���������ͨ�š�����managersģ���װ�ܺã������˽�����ͨ�ŵ�ϸ�ڣ��Ϳ��Ժ����׵ر�д�ֲ�ʽ����̳���

�ٸ����ӣ���������Ѿ���һ��ͨ��Queueͨ�ŵĶ���̳�����ͬһ̨���������У����ڣ����ڴ�������Ľ��������أ�ϣ���ѷ�������Ľ��̺ʹ�������Ľ��̷ֲ�����̨�����ϡ���ô�÷ֲ�ʽ����ʵ�֣�

ԭ�е�Queue���Լ���ʹ�ã����ǣ�ͨ��managersģ���Queueͨ�����籩¶��ȥ���Ϳ��������������Ľ��̷���Queue�ˡ�

�����ȿ�������̣�������̸�������Queue����Queueע�ᵽ�����ϣ�Ȼ����Queue����д������
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# task_master.py

import random, time, queue
from multiprocessing.managers import BaseManager

# ��������Ķ���:
task_queue = queue.Queue()
# ���ս���Ķ���:
result_queue = queue.Queue()

def return_task_queue():
    global task_queue
    return task_queue


def return_result_queue():
    global result_queue
    return result_queue

# ��BaseManager�̳е�QueueManager:
class QueueManager(BaseManager):
    pass

if __name__ == '__main__':
    # ������Queue��ע�ᵽ������, callable����������Queue����:
    QueueManager.register('get_task_queue', callable=return_task_queue) #������ԭ�Ĳ�һ�� �ο�https://blog.csdn.net/lilong117194/article/details/76051843
    QueueManager.register('get_result_queue', callable=return_result_queue) #������ԭ�Ĳ�һ��
    # �󶨶˿�5000, ������֤��'abc':
    manager = QueueManager(address=('127.0.0.1', 5000), authkey=b'abc')
    # ����Queue:
    manager.start()
    # ���ͨ��������ʵ�Queue����:
    task = manager.get_task_queue()
    result = manager.get_result_queue()
    # �ż��������ȥ:
    for i in range(10):
        n = random.randint(0, 10000)
        print('Put task %d...' % n)
        task.put(n)
    # ��result���ж�ȡ���:
    print('Try get results...')
    for i in range(10):
        r = result.get(timeout=10) #�������10s�������ȡ���������쳣queue.Empty
        print('Result: %s' % r)
    # �ر�:
    manager.shutdown()
    print('master exit.')

��ע�⣬��������һ̨������д����̳���ʱ��������Queue����ֱ�������ã����ǣ��ڷֲ�ʽ����̻����£��������Queue������ֱ�Ӷ�ԭʼ��task_queue���в������������ƹ���QueueManager�ķ�װ������ͨ��manager.get_task_queue()��õ�Queue�ӿ���ӡ�

Ȼ������һ̨����������������̣�����������Ҳ���ԣ���
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# task_worker.py

import time, sys, queue
from multiprocessing.managers import BaseManager

# �������Ƶ�QueueManager:
class QueueManager(BaseManager):
    pass

if __name__ == '__main__':
    # �������QueueManagerֻ�������ϻ�ȡQueue������ע��ʱֻ�ṩ����:
    QueueManager.register('get_task_queue')
    QueueManager.register('get_result_queue')

    # ���ӵ���������Ҳ��������task_master.py�Ļ���:
    server_addr = '127.0.0.1'
    print('Connect to server %s...' % server_addr)
    # �˿ں���֤��ע�Ᵽ����task_master.py���õ���ȫһ��:
    m = QueueManager(address=(server_addr, 5000), authkey=b'abc')
    # ����������:
    m.connect()
    # ��ȡQueue�Ķ���:
    task = m.get_task_queue()
    result = m.get_result_queue()
    # ��task����ȡ����,���ѽ��д��result����:
    for i in range(10):
        try:
            n = task.get(timeout=1)
            print('run task %d * %d...' % (n, n))
            r = '%d * %d = %d' % (n, n, n*n)
            time.sleep(1)
            result.put(r)
        except queue.Empty: #�����master��put��ѭ��ɾ������worker��get���ｫ��10���쳣
            print('task queue is empty.')
    # �������:
    print('worker exit.')

�����Ҫͨ���������ӵ�������̣�����Ҫָ��������̵�IP��

���ڣ��������Էֲ�ʽ���̵Ĺ���Ч���ˡ�������task_master.py������̣�

$ python3 task_master.py 
Put task 3411...
Put task 1605...
Put task 1398...
Put task 4729...
Put task 5300...
Put task 7471...
Put task 68...
Put task 4219...
Put task 339...
Put task 7866...
Try get results...
task_master.py���̷���������󣬿�ʼ�ȴ�result���еĽ������������task_worker.py���̣�

$ python3 task_worker.py
Connect to server 127.0.0.1...
run task 3411 * 3411...
run task 1605 * 1605...
run task 1398 * 1398...
run task 4729 * 4729...
run task 5300 * 5300...
run task 7471 * 7471...
run task 68 * 68...
run task 4219 * 4219...
run task 339 * 339...
run task 7866 * 7866...
worker exit.

task_worker.py���̽�������task_master.py�����л������ӡ�������
Result: 3411 * 3411 = 11634921
Result: 1605 * 1605 = 2576025
Result: 1398 * 1398 = 1954404
Result: 4729 * 4729 = 22363441
Result: 5300 * 5300 = 28090000
Result: 7471 * 7471 = 55815841
Result: 68 * 68 = 4624
Result: 4219 * 4219 = 17799961
Result: 339 * 339 = 114921
Result: 7866 * 7866 = 61873956

����򵥵�Master/Workerģ����ʲô�ã���ʵ�����һ���򵥵������ķֲ�ʽ���㣬�Ѵ����ԼӸ��죬�������worker���Ϳ��԰�����ֲ�����̨������ʮ̨�����ϣ�����Ѽ���n*n�Ĵ��뻻�ɷ����ʼ�����ʵ�����ʼ����е��첽���͡�

Queue����洢���ģ�ע�⵽task_worker.py�и���û�д���Queue�Ĵ��룬���ԣ�Queue����洢��task_master.py�����У�

                                             ��
��������������������������������������������������������������������������������������     ��������������������������������������������������������������������������������
��task_master.py                           ��  ��  ��task_worker.py                        ��
��                                         ��     ��                                      ��
��  task = manager.get_task_queue()        ��  ��  ��  task = manager.get_task_queue()     ��
��  result = manager.get_result_queue()    ��     ��  result = manager.get_result_queue() ��
��              ��                          ��  ��  ��              ��                       ��
��              ��                          ��     ��              ��                       ��
��              ��                          ��  ��  ��              ��                       ��
��  ����������������������������������������������������������������������    ��     ��              ��                       ��
��  ��QueueManager                     ��    ��  ��  ��              ��                       ��
��  �� ���������������������������� �������������������������������� ��    ��     ��              ��                       ��
��  �� �� task_queue �� �� result_queue �� ��<�������੤���੤���੤����������������������������                       ��
��  �� ���������������������������� �������������������������������� ��    ��     ��                                      ��
��  ����������������������������������������������������������������������    ��  ��  ��                                      ��
��������������������������������������������������������������������������������������     ��������������������������������������������������������������������������������
                                             ��

                                          Network
��Queue֮������ͨ��������ʣ�����ͨ��QueueManagerʵ�ֵġ�����QueueManager����Ĳ�ֹһ��Queue�����ԣ�Ҫ��ÿ��Queue��������ýӿ�������֣�����get_task_queue��

authkey��ʲô�ã�����Ϊ�˱�֤��̨��������ͨ�ţ�������������������š����task_worker.py��authkey��task_master.py��authkey��һ�£��϶����Ӳ��ϡ�

С��

Python�ķֲ�ʽ���̽ӿڼ򵥣���װ���ã��ʺ���Ҫ�ѷ�������ֲ�����̨�����Ļ����¡�

ע��Queue��������������������ͽ��ս����ÿ�����������������Ҫ����С�����緢��һ��������־�ļ������񣬾Ͳ�Ҫ���ͼ����׵���־�ļ��������Ƿ�����־�ļ���ŵ�����·������Worker������ȥ����Ĵ����϶�ȡ�ļ���
#-----------------------------------------------------------------------------------------
#������ʽ
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143193331387014ccd1040c814dee8b2164bb4f064cff000

#-----------------------------------------------------------------------------------------
#datetime
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431937554888869fb52b812243dda6103214cd61d0c2000
datetime��Python�������ں�ʱ��ı�׼�⡣

��ȡ��ǰ���ں�ʱ��

�����ȿ���λ�ȡ��ǰ���ں�ʱ�䣺

>>> from datetime import datetime
>>> now = datetime.now() # ��ȡ��ǰdatetime
>>> print(now)
2015-05-18 16:28:07.198690
>>> print(type(now))
<class 'datetime.datetime'>
ע�⵽datetime��ģ�飬datetimeģ�黹����һ��datetime�࣬ͨ��from datetime import datetime����Ĳ���datetime����ࡣ

���������import datetime�����������ȫ��datetime.datetime��

datetime.now()���ص�ǰ���ں�ʱ�䣬��������datetime��

��ȡָ�����ں�ʱ��

Ҫָ��ĳ�����ں�ʱ�䣬����ֱ���ò�������һ��datetime��

>>> from datetime import datetime
>>> dt = datetime(2015, 4, 19, 12, 20) # ��ָ������ʱ�䴴��datetime
>>> print(dt)
2015-04-19 12:20:00
datetimeת��Ϊtimestamp

�ڼ�����У�ʱ��ʵ�����������ֱ�ʾ�ġ����ǰ�1970��1��1�� 00:00:00 UTC+00:00ʱ����ʱ�̳�Ϊepoch time����Ϊ0��1970����ǰ��ʱ��timestampΪ����������ǰʱ����������epoch time����������Ϊtimestamp��

�������Ϊ��

timestamp = 0 = 1970-1-1 00:00:00 UTC+0:00
��Ӧ�ı���ʱ���ǣ�

timestamp = 0 = 1970-1-1 08:00:00 UTC+8:00
�ɼ�timestamp��ֵ��ʱ�����޹�ϵ����Ϊtimestampһ��ȷ������UTCʱ���ȷ���ˣ�ת��������ʱ����ʱ��Ҳ����ȫȷ���ģ������Ϊʲô������洢�ĵ�ǰʱ������timestamp��ʾ�ģ���Ϊȫ����صļ����������ʱ�̵�timestamp������ȫ��ͬ�ģ��ٶ�ʱ����У׼����

��һ��datetime����ת��Ϊtimestampֻ��Ҫ�򵥵���timestamp()������

>>> from datetime import datetime
>>> dt = datetime(2015, 4, 19, 12, 20) # ��ָ������ʱ�䴴��datetime
>>> dt.timestamp() # ��datetimeת��Ϊtimestamp
1429417200.0
ע��Python��timestamp��һ���������������С��λ��С��λ��ʾ��������

ĳЩ������ԣ���Java��JavaScript����timestampʹ��������ʾ�����������������ֻ��Ҫ��timestamp����1000�͵õ�Python�ĸ����ʾ������

timestampת��Ϊdatetime

Ҫ��timestampת��Ϊdatetime��ʹ��datetime�ṩ��fromtimestamp()������

>>> from datetime import datetime
>>> t = 1429417200.0
>>> print(datetime.fromtimestamp(t))
2015-04-19 12:20:00
ע�⵽timestamp��һ������������û��ʱ���ĸ����datetime����ʱ���ġ�����ת������timestamp�ͱ���ʱ����ת����

����ʱ����ָ��ǰ����ϵͳ�趨��ʱ�������籱��ʱ���Ƕ�8�����򱾵�ʱ�䣺

2015-04-19 12:20:00
ʵ���Ͼ���UTC+8:00ʱ����ʱ�䣺

2015-04-19 12:20:00 UTC+8:00
���˿̵ĸ������α�׼ʱ���뱱��ʱ�����8Сʱ��Ҳ����UTC+0:00ʱ����ʱ��Ӧ���ǣ�

2015-04-19 04:20:00 UTC+0:00
timestampҲ����ֱ�ӱ�ת����UTC��׼ʱ����ʱ�䣺

>>> from datetime import datetime
>>> t = 1429417200.0
>>> print(datetime.fromtimestamp(t)) # ����ʱ��
2015-04-19 12:20:00
>>> print(datetime.utcfromtimestamp(t)) # UTCʱ��
2015-04-19 04:20:00
strת��Ϊdatetime

�ܶ�ʱ���û���������ں�ʱ�����ַ�����Ҫ�������ں�ʱ�䣬���ȱ����strת��Ϊdatetime��ת��������ͨ��datetime.strptime()ʵ�֣���Ҫһ�����ں�ʱ��ĸ�ʽ���ַ�����

>>> from datetime import datetime
>>> cday = datetime.strptime('2015-6-1 18:19:59', '%Y-%m-%d %H:%M:%S')
>>> print(cday)
2015-06-01 18:19:59
�ַ���'%Y-%m-%d %H:%M:%S'�涨�����ں�ʱ�䲿�ֵĸ�ʽ����ϸ��˵����ο�Python�ĵ���

ע��ת�����datetime��û��ʱ����Ϣ�ġ�

datetimeת��Ϊstr

����Ѿ�����datetime����Ҫ������ʽ��Ϊ�ַ�����ʾ���û�������Ҫת��Ϊstr��ת��������ͨ��strftime()ʵ�ֵģ�ͬ����Ҫһ�����ں�ʱ��ĸ�ʽ���ַ�����

>>> from datetime import datetime
>>> now = datetime.now()
>>> print(now.strftime('%a, %b %d %H:%M'))
Mon, May 05 16:28
datetime�Ӽ�

�����ں�ʱ����мӼ�ʵ���Ͼ��ǰ�datetime�������ǰ���㣬�õ��µ�datetime���Ӽ�����ֱ����+��-�������������Ҫ����timedelta����ࣺ

>>> from datetime import datetime, timedelta
>>> now = datetime.now()
>>> now
datetime.datetime(2015, 5, 18, 16, 57, 3, 540997)
>>> now + timedelta(hours=10)
datetime.datetime(2015, 5, 19, 2, 57, 3, 540997)
>>> now - timedelta(days=1)
datetime.datetime(2015, 5, 17, 16, 57, 3, 540997)
>>> now + timedelta(days=2, hours=12)
datetime.datetime(2015, 5, 21, 4, 57, 3, 540997)
�ɼ���ʹ��timedelta����Ժ����׵����ǰ����ͺ����ʱ�̡�

����ʱ��ת��ΪUTCʱ��

����ʱ����ָϵͳ�趨ʱ����ʱ�䣬���籱��ʱ����UTC+8:00ʱ����ʱ�䣬��UTCʱ��ָUTC+0:00ʱ����ʱ�䡣

һ��datetime������һ��ʱ������tzinfo������Ĭ��ΪNone�������޷��������datetime�������ĸ�ʱ��������ǿ�и�datetime����һ��ʱ����

>>> from datetime import datetime, timedelta, timezone
>>> tz_utc_8 = timezone(timedelta(hours=8)) # ����ʱ��UTC+8:00
>>> now = datetime.now()
>>> now
datetime.datetime(2015, 5, 18, 17, 2, 10, 871012)
>>> dt = now.replace(tzinfo=tz_utc_8) # ǿ������ΪUTC+8:00
>>> dt
datetime.datetime(2015, 5, 18, 17, 2, 10, 871012, tzinfo=datetime.timezone(datetime.timedelta(0, 28800)))
���ϵͳʱ��ǡ����UTC+8:00����ô�������������ȷ�ģ����򣬲���ǿ������ΪUTC+8:00ʱ����

ʱ��ת��

���ǿ�����ͨ��utcnow()�õ���ǰ��UTCʱ�䣬��ת��Ϊ����ʱ����ʱ�䣺

# �õ�UTCʱ�䣬��ǿ������ʱ��ΪUTC+0:00:
>>> utc_dt = datetime.utcnow().replace(tzinfo=timezone.utc)
>>> print(utc_dt)
2015-05-18 09:05:12.377316+00:00
# astimezone()��ת��ʱ��Ϊ����ʱ��:
>>> bj_dt = utc_dt.astimezone(timezone(timedelta(hours=8)))
>>> print(bj_dt)
2015-05-18 17:05:12.377316+08:00
# astimezone()��ת��ʱ��Ϊ����ʱ��:
>>> tokyo_dt = utc_dt.astimezone(timezone(timedelta(hours=9)))
>>> print(tokyo_dt)
2015-05-18 18:05:12.377316+09:00
# astimezone()��bj_dtת��ʱ��Ϊ����ʱ��:
>>> tokyo_dt2 = bj_dt.astimezone(timezone(timedelta(hours=9)))
>>> print(tokyo_dt2)
2015-05-18 18:05:12.377316+09:00
ʱ��ת���Ĺؼ����ڣ��õ�һ��datetimeʱ��Ҫ��֪����ȷ��ʱ����Ȼ��ǿ������ʱ������Ϊ��׼ʱ�䡣

���ô�ʱ����datetime��ͨ��astimezone()����������ת��������ʱ����

ע�����Ǳ����UTC+0:00ʱ��ת��������ʱ�����κδ�ʱ����datetime��������ȷת������������bj_dt��tokyo_dt��ת����

С��

datetime��ʾ��ʱ����Ҫʱ����Ϣ����ȷ��һ���ض���ʱ�䣬����ֻ����Ϊ����ʱ�䡣

���Ҫ�洢datetime����ѷ����ǽ���ת��Ϊtimestamp�ٴ洢����Ϊtimestamp��ֵ��ʱ����ȫ�޹ء�
#-----------------------------------------------------------------------------------------
#collections
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001431953239820157155d21c494e5786fce303f3018c86000
collections��Python�ڽ���һ������ģ�飬�ṩ��������õļ����ࡣ

namedtuple

����֪��tuple���Ա�ʾ���伯�ϣ����磬һ����Ķ�ά����Ϳ��Ա�ʾ�ɣ�

>>> p = (1, 2)
���ǣ�����(1, 2)�����ѿ������tuple��������ʾһ������ġ�

����һ��class��С������ˣ���ʱ��namedtuple���������ó���

>>> from collections import namedtuple
>>> Point = namedtuple('Point', ['x', 'y'])
>>> p = Point(1, 2)
>>> p.x
1
>>> p.y
2
namedtuple��һ������������������һ���Զ����tuple���󣬲��ҹ涨��tupleԪ�صĸ����������������Զ���������������tuple��ĳ��Ԫ�ء�

����һ����������namedtuple���Ժܷ���ض���һ���������ͣ����߱�tuple�Ĳ����ԣ��ֿ��Ը������������ã�ʹ��ʮ�ַ��㡣

������֤������Point������tuple��һ�����ࣺ

>>> isinstance(p, Point)
True
>>> isinstance(p, tuple)
True
���Ƶģ����Ҫ������Ͱ뾶��ʾһ��Բ��Ҳ������namedtuple���壺

# namedtuple('����', [����list]):
Circle = namedtuple('Circle', ['x', 'y', 'r'])
deque

ʹ��list�洢����ʱ������������Ԫ�غܿ죬���ǲ����ɾ��Ԫ�ؾͺ����ˣ���Ϊlist�����Դ洢�����������ʱ�򣬲����ɾ��Ч�ʺܵ͡�

deque��Ϊ�˸�Чʵ�ֲ����ɾ��������˫���б��ʺ����ڶ��к�ջ��

>>> from collections import deque
>>> q = deque(['a', 'b', 'c'])
>>> q.append('x')
>>> q.appendleft('y')
>>> q
deque(['y', 'a', 'b', 'c', 'x'])
deque����ʵ��list��append()��pop()�⣬��֧��appendleft()��popleft()�������Ϳ��Էǳ���Ч����ͷ����ӻ�ɾ��Ԫ�ء�

defaultdict

ʹ��dictʱ��������õ�Key�����ڣ��ͻ��׳�KeyError�����ϣ��key������ʱ������һ��Ĭ��ֵ���Ϳ�����defaultdict��

>>> from collections import defaultdict
>>> dd = defaultdict(lambda: 'N/A')
>>> dd['key1'] = 'abc'
>>> dd['key1'] # key1����
'abc'
>>> dd['key2'] # key2�����ڣ�����Ĭ��ֵ
'N/A'
ע��Ĭ��ֵ�ǵ��ú������صģ��������ڴ���defaultdict����ʱ���롣

������Key������ʱ����Ĭ��ֵ��defaultdict��������Ϊ��dict����ȫһ���ġ�

OrderedDict

ʹ��dictʱ��Key������ġ��ڶ�dict������ʱ�������޷�ȷ��Key��˳��

���Ҫ����Key��˳�򣬿�����OrderedDict��

>>> from collections import OrderedDict
>>> d = dict([('a', 1), ('b', 2), ('c', 3)])
>>> d # dict��Key�������
{'a': 1, 'c': 3, 'b': 2}
>>> od = OrderedDict([('a', 1), ('b', 2), ('c', 3)])
>>> od # OrderedDict��Key�������
OrderedDict([('a', 1), ('b', 2), ('c', 3)])
ע�⣬OrderedDict��Key�ᰴ�ղ����˳�����У�����Key��������

>>> od = OrderedDict()
>>> od['z'] = 1
>>> od['y'] = 2
>>> od['x'] = 3
>>> list(od.keys()) # ���ղ����Key��˳�򷵻�
['z', 'y', 'x']
OrderedDict����ʵ��һ��FIFO���Ƚ��ȳ�����dict����������������ʱ����ɾ��������ӵ�Key��

from collections import OrderedDict

class LastUpdatedOrderedDict(OrderedDict):

    def __init__(self, capacity):
        super(LastUpdatedOrderedDict, self).__init__()
        self._capacity = capacity

    def __setitem__(self, key, value):
        containsKey = 1 if key in self else 0
        if len(self) - containsKey >= self._capacity:
            last = self.popitem(last=False)
            print('remove:', last)
        if containsKey:
            del self[key]
            print('set:', (key, value))
        else:
            print('add:', (key, value))
        OrderedDict.__setitem__(self, key, value)
Counter

Counter��һ���򵥵ļ����������磬ͳ���ַ����ֵĸ�����

>>> from collections import Counter
>>> c = Counter()
>>> for ch in 'programming':
...     c[ch] = c[ch] + 1
...
>>> c
Counter({'g': 2, 'm': 2, 'r': 2, 'a': 1, 'i': 1, 'o': 1, 'n': 1, 'p': 1})
Counterʵ����Ҳ��dict��һ�����࣬����Ľ�����Կ������ַ�'g'��'m'��'r'�����������Σ������ַ���������һ�Ρ�

С��

collectionsģ���ṩ��һЩ���õļ����࣬���Ը�����Ҫѡ�á�
#-----------------------------------------------------------------------------------------
#contextlib
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001478651770626de401ff1c0d94f379774cabd842222ff000
��Python�У���д�ļ���������ԴҪ�ر�ע�⣬������ʹ����Ϻ���ȷ�ر����ǡ���ȷ�ر��ļ���Դ��һ��������ʹ��try...finally��
try:
try:
    f = open('test.txt', 'r')
    print(f.read())
finally:
    if f:
        f.close()

дtry...finally�ǳ�������Python��with����������Ƿǳ������ʹ����Դ�������ص�����Դû�йرգ���������Ĵ�����Լ�Ϊ��
with open('test.txt', 'r') as f:
    print(f.read())

�������ǾͿ��԰��Լ�д����Դ��������with��䣺
with Query('Bob') as q:
    q.query()
    
#example.py
class Query(object):
    def __init__(self, name):
        self.name = name

    def __enter__(self):
        print('Begin')
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        if exc_type:
            print('Error')
        else:
            print('End')

    def query(self):
        print('Query info about %s...' % self.name)

if __name__=='__main__':
    with Query('Bob') as q:
        q.query()
        #raise RuntimeError("error")
�����
Begin
Query info about Bob...
End

@contextmanager
��д__enter__��__exit__��Ȼ�ܷ��������Python�ı�׼��contextlib�ṩ�˸��򵥵�д��������Ĵ�����Ը�д���£�
from contextlib import contextmanager
class Query(object):
    def __init__(self, name):
        self.name = name

    def query(self):
        print('Query info about %s...' % self.name)

@contextmanager
def create_query(name):
    print('Begin')
    q = Query(name)
    yield q
    print('End')

if __name__=='__main__':
    with create_query('Bob') as q:
        print("Before q.query()")
        q.query()
        print("After q.query()")
�����
Begin
Before q.query()
Query info about Bob...
After q.query()
End

�ܶ�ʱ������ϣ����ĳ�δ���ִ��ǰ���Զ�ִ���ض����룬Ҳ������@contextmanagerʵ�֡����磺
from contextlib import contextmanager

@contextmanager
def tag(name):
    print("<%s>" % name)
    yield
    print("</%s>" % name)

if __name__=='__main__':
    with tag("h1"):
        print("hello")
        print("world")
�����
<h1>
hello
world
</h1>

�����ִ��˳���ǣ�
with�������ִ��yield֮ǰ����䣬��˴�ӡ��<h1>��
yield���û�ִ��with����ڲ���������䣬��˴�ӡ��hello��world��
���ִ��yield֮�����䣬��ӡ��</h1>��
��ˣ�@contextmanager������ͨ����дgenerator���������Ĺ���

@closing
���һ������û��ʵ�������ģ����ǾͲ��ܰ�������with��䡣���ʱ�򣬿�����closing()���Ѹö����Ϊ�����Ķ������磬��with���ʹ��urlopen()��
from contextlib import closing
from urllib.request import urlopen

with closing(urlopen('https://www.python.org')) as page:#������Ӳ���ǡ������ʹ��closing��Ҳ����ʹ��with���with urlopen('https://www.python.org') as page:
    for line in page:
        print(line)

�������Ҫ��with as��ʹ�ã���ö�����Ҫʵ��__enter__��__exit__���������û��ʵ��������������������with as��ʹ�ö�����ֻҪ�ṩ��close�������Ϳ�����closing������
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from contextlib import closing


class Query(object):
    def __init__(self, name):
        self.name = name

    def query(self):
        print('Query info about %s...' % self.name)

    def close(self):
        print('Query object close')

if __name__=='__main__':
    with closing(Query('Bob')) as q:
        q.query()

#https://www.cnblogs.com/DswCnblog/p/6126588.html
with����ǿ��֮���������Դ����쳣���������Ѿ�ע�⵽Sample���__exit__��������������- val, type �� trace�� ��Щ�������쳣�������൱���á���������һ�´��룬����������ι����ġ�
#!/usr/bin/env python
# with_example02.py
 
class Sample:
    def __enter__(self):
        return self
 
    def __exit__(self, type, value, trace):
        print "type:", type
        print "value:", value
        print "trace:", trace
 
    def do_something(self):
        bar = 1/0
        return bar + 10
 
with Sample() as sample:
    sample.do_something()

��������У�with�����get_sample()�����Sample()����û���κι�ϵ��ֻҪ����with�������������صĶ�����__enter__()��__exit__()�������ɡ������У�Sample()��__enter__()���������´�����Sample���󣬲���ֵ������sample��
����ִ�к�

bash-3.2$ ./with_example02.py
type: <type 'exceptions.ZeroDivisionError'>
value: integer division or modulo by zero
trace: <traceback object at 0x1004a8128>
Traceback (most recent call last):
  File "./with_example02.py", line 19, in <module>
    sample.do_something()
  File "./with_example02.py", line 15, in do_something
    bar = 1/0
ZeroDivisionError: integer division or modulo by zero

ʵ���ϣ���with����Ĵ�����׳��κ��쳣ʱ��__exit__()������ִ�С�����������ʾ���쳣�׳�ʱ����֮������type��value��stack trace����__exit__()����������׳���ZeroDivisionError�쳣����ӡ�����ˡ�������ʱ��������Դ���ر��ļ��ȵȲ����������Է���__exit__�������С�
��ˣ�Python��with������ṩһ����Ч�Ļ��ƣ��ô����������ͬʱ���쳣����ʱ�����������򵥡�

#https://blog.csdn.net/zly9923218/article/details/53404849
���⣬__exit__��������tear things down�������Խ����쳣�ļ�غʹ���ע��󼸸�������Ҫ����һ���쳣��ֻ��Ҫ���ظú���True���ɡ�����������������������е�TypeError�����������쳣�����׳���
def __exit__(self, type, value, traceback):
    return isinstance(value, TypeError)

��python2.5���Ժ�file�����Ѿ�д����enter��exit���������ǿ����������ԣ�

>>> f = open("x.txt")
>>> f
<open file 'x.txt', mode 'r' at 0x00AE82F0>
>>> f.__enter__()
<open file 'x.txt', mode 'r' at 0x00AE82F0>
>>> f.read(1)
'X'
>>> f.__exit__(None, None, None)
>>> f.read(1)
Traceback (most recent call last):
    File "<stdin>", line 1, in <module>
ValueError: I/O operation on closed file
֮���������Ҫ���ļ�����֤���ر�����ֻ��Ҫ��ô����
with open("x.txt") as f:
    data = f.read()
    do something with data
����ж������ǿ�����ôд��

with open("x.txt") as f1, open('xxx.txt') as f2:
    do something with f1,f2
#-----------------------------------------------------------------------------------------
# chardet https://www.liaoxuefeng.com/wiki/1016959663602400/1183255880134144
�ַ�������һֱ�����˷ǳ�ͷ�۵����⣬�����������ڴ���һЩ���淶�ĵ�������ҳ��ʱ����ȻPython�ṩ��Unicode��ʾ��str��bytes�����������ͣ����ҿ���ͨ��encode()��decode()����ת�������ǣ��ڲ�֪�����������£���bytes��decode()��������

����δ֪�����bytes��Ҫ����ת����str����Ҫ�ȡ��²⡱���롣�²�ķ�ʽ�����ռ����ֱ���������ַ������������ַ��жϣ������кܴ���ʡ��¶ԡ���

��Ȼ�����ǿ϶����ܴ�ͷ�Լ�д���������Ĺ��ܣ���������ʱ������chardet��������������þ��������ó��������������룬�����á�

��װchardet
�����װ��Anaconda��chardet���Ѿ������ˡ�������Ҫ����������ͨ��pip��װ��
$ pip install chardet

�������Permission denied��װʧ�ܣ������sudo���ԡ�

ʹ��chardet
�������õ�һ��bytesʱ���Ϳ��Զ�������롣��chardet�����룬ֻ��Ҫһ�д��룺
>>> chardet.detect(b'Hello, world!')
{'encoding': 'ascii', 'confidence': 1.0, 'language': ''}
�����ı�����ascii��ע�⵽���и�confidence�ֶΣ���ʾ���ĸ�����1.0����100%����

���������Լ��GBK��������ģ�
>>> data = '����ԭ�ϲݣ�һ��һ����'.encode('gbk')
>>> chardet.detect(data)
{'encoding': 'GB2312', 'confidence': 0.7407407407407407, 'language': 'Chinese'}
���ı�����GB2312��ע�⵽GBK��GB2312�ĳ�����������ͬһ�ֱ��룬�����ȷ�ĸ�����74%��language�ֶ�ָ����������'Chinese'��

��UTF-8������м�⣺
>>> data = '����ԭ�ϲݣ�һ��һ����'.encode('utf-8')
>>> chardet.detect(data)
{'encoding': 'utf-8', 'confidence': 0.99, 'language': ''}

���������Զ����Ľ��м�⣺
>>> data = '���¤���Ҫ�˥�`��'.encode('euc-jp')
>>> chardet.detect(data)
{'encoding': 'EUC-JP', 'confidence': 0.99, 'language': 'Japanese'}

�ɼ�����chardet�����룬ʹ�ü򵥡���ȡ���������ת��Ϊstr���Ϳ��Է����������

chardet֧�ּ��ı����б���ο��ٷ��ĵ�Supported encodings��

С��
ʹ��chardet������ǳ����ף�chardet֧�ּ�����ġ����ġ����ĵȶ������ԡ�
#-----------------------------------------------------------------------------------------
#TCP���
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432004374523e495f640612f4b08975398796939ec3c000
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ����socket��:
import socket

if __name__ == '__main__':
    # ����һ��socket:
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    # ��������:
    s.connect(('www.sina.com.cn', 80))

    # ��������:
    s.send(b'GET / HTTP/1.1\r\nHost: www.sina.com.cn\r\nConnection: close\r\n\r\n')

    buffer = []
    while True:
        # ÿ��������1k�ֽ�:
        d = s.recv(1024)
        if d:
            buffer.append(d)
        else:
            break
    data = b''.join(buffer)

    # �ر�����:
    s.close()

    header, html = data.split(b'\r\n\r\n', 1)
    print(header.decode('utf-8'))
    # �ѽ��յ�����д���ļ�:
    with open('sina.html', 'wb') as f:
        f.write(html)
        
        
#server.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ����socket��:
import socket
import threading
import time


def tcplink(sock, addr):
    print('Accept new connection from %s:%s...' % addr)
    sock.send(b'Welcome!')
    while True:
        data = sock.recv(1024)
        time.sleep(1)
        if not data or data.decode('utf-8') == 'exit':
            break
        sock.send(('Hello, %s!' % data.decode('utf-8')).encode('utf-8'))
    sock.close()
    print('Connection from %s:%s closed.' % addr)


if __name__ == '__main__':
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    # �����˿�:
    s.bind(('127.0.0.1', 9999))
    s.listen(5)
    print('Waiting for connection...')
    while True:
        # ����һ��������:
        sock, addr = s.accept()
        # �������߳�������TCP����:
        t = threading.Thread(target=tcplink, args=(sock, addr))
        t.start()
�����
Waiting for connection...
Accept new connection from 127.0.0.1:9327...
Connection from 127.0.0.1:9327 closed.


#client.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# client.py
import socket

if __name__ == '__main__':
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    # ��������:
    s.connect(('127.0.0.1', 9999))
    # ���ջ�ӭ��Ϣ:
    print(s.recv(1024).decode('utf-8'))
    for data in [b'Michael', b'Tracy', b'Sarah']:
        # ��������:
        s.send(data)
        print(s.recv(1024).decode('utf-8'))
    s.send(b'exit')
    s.close()
�����
Welcome!
Hello, Michael!
Hello, Tracy!
Hello, Sarah!

С��
��TCPЭ�����Socket�����Python��ʮ�ּ򵥣����ڿͻ��ˣ�Ҫ�������ӷ�������IP��ָ���˿ڣ����ڷ�������Ҫ���ȼ���ָ���˿ڣ�Ȼ�󣬶�ÿһ���µ����ӣ�����һ���̻߳����������ͨ�������������������������ȥ��

ͬһ���˿ڣ���һ��Socket�����Ժ󣬾Ͳ��ܱ����Socket���ˡ�
#-----------------------------------------------------------------------------------------
#UDP���
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432004977916a212e2168e21449981ad65cd16e71201000

#server.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ����socket��:
import socket

if __name__ == '__main__':
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    # �󶨶˿�:
    s.bind(('127.0.0.1', 9999))
    print('Bind UDP on 9999...')
    while True:
        # ��������:
        data, addr = s.recvfrom(1024)
        print('Received from %s:%s.' % addr)
        s.sendto(b'Hello, %s!' % data, addr)
�����
Bind UDP on 9999...
Received from 127.0.0.1:57897.
Received from 127.0.0.1:57897.
Received from 127.0.0.1:57897.

#client.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# client.py
import socket

if __name__ == '__main__':
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    for data in [b'Michael', b'Tracy', b'Sarah']:
        # ��������:
        s.sendto(data, ('127.0.0.1', 9999))
        # ��������:
        print(s.recv(1024).decode('utf-8'))
    s.close()
�����
Hello, Michael!
Hello, Tracy!
Hello, Sarah!

С��
UDP��ʹ����TCP���ƣ����ǲ���Ҫ�������ӡ����⣬��������UDP�˿ں�TCP�˿ڻ�����ͻ��Ҳ����˵��UDP��9999�˿���TCP��9999�˿ڿ��Ը��԰󶨡�
#-----------------------------------------------------------------------------------------
#�����ʼ�
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432005156604f38836be1707453eb025ce8c3079978d000
Email����ʷ��Web��Ҫ��Զ��ֱ�����ڣ�EmailҲ�ǻ�������Ӧ�÷ǳ��㷺�ķ���

�������еı�����Զ�֧�ַ��ͺͽ��յ����ʼ������ǣ��ȵȵȣ������ǿ�ʼ��д����֮ǰ���б�Ҫ����������ʼ�������ڻ������������ġ�

������������ͳ�ʼ�����������ġ������������ڱ�����Ҫ��һ����۵����ѷ�һ���ţ���ô���أ�

�������д���ţ�װ���ŷ⣬д�ϵ�ַ��������Ʊ��Ȼ��ͽ��Ҹ��ʾ֣������Խ�ȥ��

�ż���Ӿͽ���С�ʾ�ת�˵����ʾ֣��ٴӴ��ʾ�����ĳ��з��������ȷ���������ߺ��˵�����ۣ�Ҳ�����߾����ߵ���ۣ������㲻�ù��ľ���·�ߣ���ֻ��Ҫ֪��һ���£������ż��ߵú���������Ҫ����ʱ�䡣

�ż�������۵�ĳ���ʾ֣�Ҳ����ֱ���͵����ѵļ����Ϊ�ʾֵ������Ǻܴ����ģ�����������Ѳ��ڼң�һ��һ�˵ذ��ܣ����ԣ��ż���Ͷ�ݵ�������ѵ��������������ڹ�Ԣ��һ�㣬���߼��ſڣ�ֱ��������ѻؼҵ�ʱ�������䣬�����ż��󣬾Ϳ���ȡ���ʼ��ˡ�

�����ʼ������̻�����Ҳ�ǰ�����ķ�ʽ�����ģ�ֻ�����ٶȲ��ǰ����㣬���ǰ����㡣

�������ǻص������ʼ������������Լ��ĵ����ʼ���ַ��me@163.com���Է��ĵ����ʼ���ַ��friend@sina.com��ע���ַ�����鹹�Ĺ���������������Outlook����Foxmail֮������д���ʼ������϶Է���Email��ַ���㡰���͡��������ʼ��ͷ���ȥ�ˡ���Щ�����ʼ��������ΪMUA��Mail User Agent�����ʼ��û�����

Email��MUA����ȥ������ֱ�ӵ���Է����ԣ����Ƿ���MTA��Mail Transfer Agent�����ʼ��������������ЩEmail�����ṩ�̣��������ס����˵ȵȡ����������Լ��ĵ����ʼ���163.com�����ԣ�Email���ȱ�Ͷ�ݵ������ṩ��MTA���������׵�MTA�����Է������̣�Ҳ�������˵�MTA����������м���ܻ��ᾭ�����MTA���������ǲ����ľ���·�ߣ�����ֻ�����ٶȡ�

Email�������˵�MTA�����ڶԷ�ʹ�õ���@sina.com�����䣬��ˣ����˵�MTA���EmailͶ�ݵ��ʼ�������Ŀ�ĵ�MDA��Mail Delivery Agent�����ʼ�Ͷ�ݴ���Email����MDA�󣬾;������������˵�ĳ���������ϣ������ĳ���ļ�����������ݿ�����ǽ�������ڱ����ʼ��ĵط���֮Ϊ�������䡣

ͬ��ͨ�ʼ����ƣ�Email����ֱ�ӵ���Է��ĵ��ԣ���Ϊ�Է����Բ�һ������������Ҳ��һ���������Է�Ҫȡ���ʼ�������ͨ��MUA��MDA�ϰ��ʼ�ȡ���Լ��ĵ����ϡ�

���ԣ�һ������ʼ����ó̾��ǣ�

������ -> MUA -> MTA -> MTA -> ���ɸ�MTA -> MDA <- MUA <- �ռ���
���������������Ҫ��д���������ͺͽ����ʼ��������Ͼ��ǣ�

��дMUA���ʼ�����MTA��

��дMUA��MDA�����ʼ���

���ʼ�ʱ��MUA��MTAʹ�õ�Э�����SMTP��Simple Mail Transfer Protocol�������MTA����һ��MTAҲ����SMTPЭ�顣

���ʼ�ʱ��MUA��MDAʹ�õ�Э�������֣�POP��Post Office Protocol��Ŀǰ�汾��3���׳�POP3��IMAP��Internet Message Access Protocol��Ŀǰ�汾��4���ŵ��ǲ�����ȡ�ʼ���������ֱ�Ӳ���MDA�ϴ洢���ʼ���������ռ����Ƶ������䣬�ȵȡ�

�ʼ��ͻ�������ڷ��ʼ�ʱ��������������SMTP��������Ҳ������Ҫ�����ĸ�MTA�ϡ�����������ʹ��163�����䣬��Ͳ���ֱ�ӷ������˵�MTA�ϣ���Ϊ��ֻ�������˵��û������ԣ������163�ṩ��SMTP��������ַ��smtp.163.com��Ϊ��֤������163���û���SMTP��������Ҫ������д�����ַ��������������MUA���������ذ�Emailͨ��SMTPЭ�鷢�͵�MTA��

���Ƶģ���MDA���ʼ�ʱ��MDA������ҲҪ����֤���������ȷ����������ð������ȡ����ʼ������ԣ�Outlook֮����ʼ��ͻ��˻�Ҫ������дPOP3��IMAP��������ַ�������ַ�Ϳ��������MUA����˳����ͨ��POP��IMAPЭ���MDAȡ���ʼ���

��ʹ��Python�շ��ʼ�ǰ������׼�����������������ʼ�����xxx@163.com��xxx@sina.com��xxx@qq.com�ȣ�ע���������䲻Ҫ��ͬһ���ʼ������̡�

����ر�ע�⣬Ŀǰ������ʼ������̶���Ҫ�ֶ���SMTP���ź�POP���ŵĹ��ܣ�����ֻ��������ҳ��¼��

#-----------------------------------------------------------------------------------------
#SMTP�����ʼ�
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432005226355aadb8d4b2f3f42f6b1d6f2c5bd8d5263000

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from email.header import Header
from email.mime.text import MIMEText
import smtplib

def send_email(SMTP_host, from_account, from_passwd, to_account, subject, content):
    email_client = smtplib.SMTP(SMTP_host, 25)
    email_client.login(from_account, from_passwd)
    # create msg
    msg = MIMEText(content, 'plain', 'utf-8')
    msg['Subject'] = Header(subject, 'utf-8')  # subject
    msg['From'] = from_account
    msg['To'] = to_account
    email_client.set_debuglevel(1)
    email_client.sendmail(from_account, to_account, msg.as_string())

    email_client.quit()

if __name__=='__main__':
    send_email('smtp.163.com', 'bigben0204@163.com', 'ilovesjtu8624', '464754973@qq.com', 'abc', 'hello, send by Ben')

������set_debuglevel(1)�Ϳ��Դ�ӡ����SMTP������������������Ϣ��SMTPЭ����Ǽ򵥵��ı��������Ӧ��login()����������¼SMTP��������sendmail()�������Ƿ��ʼ������ڿ���һ�η�������ˣ����Դ���һ��list���ʼ�������һ��str��as_string()��MIMEText������str��

����HTML�ʼ�
�������Ҫ����HTML�ʼ�����������ͨ�Ĵ��ı��ļ���ô�죿�����ܼ򵥣��ڹ���MIMEText����ʱ����HTML�ַ�������ȥ���ٰѵڶ���������plain��Ϊhtml�Ϳ����ˣ�

msg = MIMEText('<html><body><h1>Hello</h1>' +
    '<p>send by <a href="http://www.python.org">Python</a>...</p>' +
    '</body></html>', 'html', 'utf-8')
�ٷ���һ���ʼ����㽫������HTML��ʾ���ʼ���


SMTP�����ʼ�
�Ķ�: 89487
SMTP�Ƿ����ʼ���Э�飬Python���ö�SMTP��֧�֣����Է��ʹ��ı��ʼ���HTML�ʼ��Լ����������ʼ���

Python��SMTP֧����smtplib��email����ģ�飬email�������ʼ���smtplib�������ʼ���

���ȣ�����������һ����򵥵Ĵ��ı��ʼ���

from email.mime.text import MIMEText
msg = MIMEText('hello, send by Python...', 'plain', 'utf-8')
ע�⵽����MIMEText����ʱ����һ�����������ʼ����ģ��ڶ���������MIME��subtype������'plain'��ʾ���ı������յ�MIME����'text/plain'�����һ��Ҫ��utf-8���뱣֤�����Լ����ԡ�

Ȼ��ͨ��SMTP����ȥ��

# ����Email��ַ�Ϳ���:
from_addr = input('From: ')
password = input('Password: ')
# �����ռ��˵�ַ:
to_addr = input('To: ')
# ����SMTP��������ַ:
smtp_server = input('SMTP server: ')

import smtplib
server = smtplib.SMTP(smtp_server, 25) # SMTPЭ��Ĭ�϶˿���25
server.set_debuglevel(1)
server.login(from_addr, password)
server.sendmail(from_addr, [to_addr], msg.as_string())
server.quit()
������set_debuglevel(1)�Ϳ��Դ�ӡ����SMTP������������������Ϣ��SMTPЭ����Ǽ򵥵��ı��������Ӧ��login()����������¼SMTP��������sendmail()�������Ƿ��ʼ������ڿ���һ�η�������ˣ����Դ���һ��list���ʼ�������һ��str��as_string()��MIMEText������str��

���һ��˳�����Ϳ������ռ����������յ����Ǹշ��͵�Email��

send-mail

��ϸ�۲죬�����������⣺

�ʼ�û�����⣻
�ռ��˵�����û����ʾΪ�Ѻõ����֣�����Mr Green <green@example.com>��
�����յ����ʼ���ȴ��ʾ�����ռ����С�
������Ϊ�ʼ����⡢�����ʾ�����ˡ��ռ��˵���Ϣ������ͨ��SMTPЭ�鷢��MTA�����ǰ����ڷ���MTA���ı��еģ����ԣ����Ǳ����From��To��Subject��ӵ�MIMEText�У�����һ���������ʼ���

from email import encoders
from email.header import Header
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr

import smtplib

def _format_addr(s):
    name, addr = parseaddr(s)
    return formataddr((Header(name, 'utf-8').encode(), addr))

from_addr = input('From: ')
password = input('Password: ')
to_addr = input('To: ')
smtp_server = input('SMTP server: ')

msg = MIMEText('hello, send by Python...', 'plain', 'utf-8')
msg['From'] = _format_addr('Python������ <%s>' % from_addr)
msg['To'] = _format_addr('����Ա <%s>' % to_addr)
msg['Subject'] = Header('����SMTP���ʺ򡭡�', 'utf-8').encode()

server = smtplib.SMTP(smtp_server, 25)
server.set_debuglevel(1)
server.login(from_addr, password)
server.sendmail(from_addr, [to_addr], msg.as_string())
server.quit()
���Ǳ�д��һ������_format_addr()����ʽ��һ���ʼ���ַ��ע�ⲻ�ܼ򵥵ش���name <addr@example.com>����Ϊ����������ģ���Ҫͨ��Header������б��롣

msg['To']���յ����ַ���������list������ж���ʼ���ַ����,�ָ����ɡ�

�ٷ���һ���ʼ����Ϳ������ռ��������п�����ȷ�ı��⡢�����˺��ռ��ˣ�

mail-with-header

�㿴�����ռ��˵����ֺܿ��ܲ������Ǵ���Ĺ���Ա����Ϊ�ܶ��ʼ�����������ʾ�ʼ�ʱ������ռ��������Զ��滻Ϊ�û�ע������֣����������ռ������ֵ���ʾ����Ӱ�졣

������ǲ鿴Email��ԭʼ���ݣ����Կ������¾���������ʼ�ͷ��

From: =?utf-8?b?UHl0aG9u54ix5aW96ICF?= <xxxxxx@163.com>
To: =?utf-8?b?566h55CG5ZGY?= <xxxxxx@qq.com>
Subject: =?utf-8?b?5p2l6IeqU01UUOeahOmXruWAmeKApuKApg==?=
����Ǿ���Header���������ı�������utf-8������Ϣ��Base64������ı�����������Լ����ֶ����������ı����ı�����Ȼ�Ƚϸ��ӡ�

����HTML�ʼ�
�������Ҫ����HTML�ʼ�����������ͨ�Ĵ��ı��ļ���ô�죿�����ܼ򵥣��ڹ���MIMEText����ʱ����HTML�ַ�������ȥ���ٰѵڶ���������plain��Ϊhtml�Ϳ����ˣ�

msg = MIMEText('<html><body><h1>Hello</h1>' +
    '<p>send by <a href="http://www.python.org">Python</a>...</p>' +
    '</body></html>', 'html', 'utf-8')
�ٷ���һ���ʼ����㽫������HTML��ʾ���ʼ���

html-mail

���͸���
���Email��Ҫ���ϸ�����ô�죿���������ʼ����Կ����������ɲ��ֵ��ʼ����ı��͸��������������ԣ����Թ���һ��MIMEMultipart��������ʼ�����Ȼ�����������һ��MIMEText��Ϊ�ʼ����ģ��ټ�����������ϱ�ʾ������MIMEBase���󼴿ɣ�

# �ʼ�����:
msg = MIMEMultipart()
msg['From'] = _format_addr('Python������ <%s>' % from_addr)
msg['To'] = _format_addr('����Ա <%s>' % to_addr)
msg['Subject'] = Header('����SMTP���ʺ򡭡�', 'utf-8').encode()

# �ʼ�������MIMEText:
msg.attach(MIMEText('send with file...', 'plain', 'utf-8'))

# ��Ӹ������Ǽ���һ��MIMEBase���ӱ��ض�ȡһ��ͼƬ:
with open('/Users/michael/Downloads/test.png', 'rb') as f:
    # ���ø�����MIME���ļ�����������png����:
    mime = MIMEBase('image', 'png', filename='test.png')
    # ���ϱ�Ҫ��ͷ��Ϣ:
    mime.add_header('Content-Disposition', 'attachment', filename='test.png')
    mime.add_header('Content-ID', '<0>')
    mime.add_header('X-Attachment-Id', '0')
    # �Ѹ��������ݶ�����:
    mime.set_payload(f.read())
    # ��Base64����:
    encoders.encode_base64(mime)
    # ��ӵ�MIMEMultipart:
    msg.attach(mime)
    
����ͼƬ
���Ҫ��һ��ͼƬǶ�뵽�ʼ���������ô����ֱ����HTML�ʼ�������ͼƬ��ַ�в��У����ǣ��󲿷��ʼ������̶����Զ����δ���������ͼƬ����Ϊ��֪����Щ�����Ƿ�ָ�������վ��

Ҫ��ͼƬǶ�뵽�ʼ������У�����ֻ�谴�շ��͸����ķ�ʽ���Ȱ��ʼ���Ϊ������ӽ�ȥ��Ȼ����HTML��ͨ������src="cid:0"�Ϳ��԰Ѹ�����ΪͼƬǶ���ˡ�����ж��ͼƬ�����������α�ţ�Ȼ�����ò�ͬ��cid:x���ɡ�

������������MIMEMultipart��MIMEText��plain��Ϊhtml��Ȼ�����ʵ���λ������ͼƬ��

msg.attach(MIMEText('<html><body><h1>Hello</h1>' +
    '<p><img src="cid:0"></p>' +
    '</body></html>', 'html', 'utf-8'))
�ٴη��ͣ��Ϳ��Կ���ͼƬֱ��Ƕ�뵽�ʼ����ĵ�Ч����


ͬʱ֧��HTML��Plain��ʽ
������Ƿ���HTML�ʼ����ռ���ͨ�����������Outlook֮�������ǿ�����������ʼ����ݵģ����ǣ�����ռ���ʹ�õ��豸̫���ϣ��鿴����HTML�ʼ���ô�죿

�취���ڷ���HTML��ͬʱ�ٸ���һ�����ı�������ռ����޷��鿴HTML��ʽ���ʼ����Ϳ����Զ������鿴���ı��ʼ���

����MIMEMultipart�Ϳ������һ��HTML��Plain��Ҫע��ָ��subtype��alternative��

msg = MIMEMultipart('alternative')
msg['From'] = ...
msg['To'] = ...
msg['Subject'] = ...

msg.attach(MIMEText('hello', 'plain', 'utf-8'))
msg.attach(MIMEText('<html><body><h1>Hello</h1></body></html>', 'html', 'utf-8'))
# ��������msg����...
����SMTP
ʹ�ñ�׼��25�˿�����SMTP������ʱ��ʹ�õ������Ĵ��䣬�����ʼ����������̿��ܻᱻ������Ҫ����ȫ�ط����ʼ������Լ���SMTP�Ự��ʵ���Ͼ����ȴ���SSL��ȫ���ӣ�Ȼ����ʹ��SMTPЭ�鷢���ʼ���

ĳЩ�ʼ������̣�����Gmail���ṩ��SMTP�������Ҫ���ܴ��䡣�������������ͨ��Gmail�ṩ�İ�ȫSMTP�����ʼ���

����֪����Gmail��SMTP�˿���587����ˣ��޸Ĵ������£�

smtp_server = 'smtp.gmail.com'
smtp_port = 587
server = smtplib.SMTP(smtp_server, smtp_port)
server.starttls()
# ʣ�µĴ����ǰ���һģһ��:
server.set_debuglevel(1)
...
ֻ��Ҫ�ڴ���SMTP��������̵���starttls()�������ʹ����˰�ȫ���ӡ�����Ĵ����ǰ��ķ����ʼ�������ȫһ����

�����Ϊ���������޷�����Gmail��SMTP�����������������ǵĴ�����û������ģ�����Ҫ�����������������Ҫ�ĵ�����

С��
ʹ��Python��smtplib�����ʼ�ʮ�ּ򵥣�ֻҪ�����˸����ʼ����͵Ĺ��췽������ȷ���ú��ʼ�ͷ���Ϳ���˳��������

����һ���ʼ��������һ��Messag�����������һ��MIMEText���󣬾ͱ�ʾһ���ı��ʼ������������һ��MIMEImage���󣬾ͱ�ʾһ����Ϊ������ͼƬ��Ҫ�Ѷ�������������������MIMEMultipart���󣬶�MIMEBase���Ա�ʾ�κζ������ǵļ̳й�ϵ���£�

Message
+- MIMEBase
   +- MIMEMultipart
   +- MIMENonMultipart
      +- MIMEMessage
      +- MIMEText
      +- MIMEImage
����Ƕ�׹�ϵ�Ϳ��Թ�������⸴�ӵ��ʼ��������ͨ��email.mime�ĵ��鿴�������ڵİ��Լ���ϸ���÷���
#-----------------------------------------------------------------------------------------
#POP3��ȡ�ʼ�
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014320098721191b70a2cf7b5441deb01595edd8147196000
#-----------------------------------------------------------------------------------------
#HTTPЭ����
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432011939547478fd5482deb47b08716557cc99764e0000
#-----------------------------------------------------------------------------------------
#WSGI�ӿ�
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432012393132788f71e0edad4676a3f76ac7776f3a16000

����WSGI����
�����ȱ�дhello.py��ʵ��WebӦ�ó����WSGI��������

# hello.py

def application(environ, start_response):
    start_response('200 OK', [('Content-Type', 'text/html')])
    return [b'<h1>Hello, web!</h1>']
Ȼ���ٱ�дһ��server.py����������WSGI������������application()������

# server.py
# ��wsgirefģ�鵼��:
from wsgiref.simple_server import make_server
# ���������Լ���д��application����:
from hello import application

# ����һ����������IP��ַΪ�գ��˿���8000����������application:
httpd = make_server('', 8000, application)
print('Serving HTTP on port 8000...')
# ��ʼ����HTTP����:
httpd.serve_forever()
ȷ�����������ļ���ͬһ��Ŀ¼�£�Ȼ��������������python server.py������WSGI��������

wsgiref-start

ע�⣺���8000�˿��ѱ���������ռ�ã�������ʧ�ܣ����޸ĳ������˿ڡ�

�����ɹ��󣬴������������http://localhost:8000/���Ϳ��Կ�������ˣ�

hello-web

�������п��Կ���wsgiref��ӡ��log��Ϣ��

Serving HTTP on port 8000...
127.0.0.1 - - [14/Apr/2018 23:16:26] "GET / HTTP/1.1" 200 20
127.0.0.1 - - [14/Apr/2018 23:16:26] "GET /favicon.ico HTTP/1.1" 200 20


��Ctrl+C��ֹ��������

�����������WebӦ��̫���ˣ�������΢����һ�£���environ���ȡPATH_INFO������������ʾ���Ӷ�̬�����ݣ�
# hello.py
def application(environ, start_response):
    start_response('200 OK', [('Content-Type', 'text/html')])
    body = '<h1>Hello, %s!</h1>' % (environ['PATH_INFO'][1:] or 'web')
    return [body.encode('utf-8')]
������ڵ�ַ�������û�����ΪURL��һ���֣�������Hello, xxx!��

hello-michael

�ǲ����е�Web App�ĸо��ˣ�

С��
���۶�ô���ӵ�WebӦ�ó�����ڶ���һ��WSGI��������HTTP���������������Ϣ������ͨ��environ��ã�HTTP��Ӧ�����������ͨ��start_response()���Ϻ�������ֵ��ΪBody��

���ӵ�WebӦ�ó��򣬹⿿һ��WSGI������������̫�ײ��ˣ�������Ҫ��WSGI֮���ٳ����Web��ܣ���һ����Web������
#-----------------------------------------------------------------------------------------
#ʹ��Web���
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432012745805707cb9f00a484d968c72dbb7cfc90b91000

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ��wsgirefģ�鵼��:
from flask import Flask
from flask import request

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def home():
    return '<h1>Home</h1>'

@app.route('/signin', methods=['GET']) #�����û��������
def signin_form():
    return '''<form action="/signin" method="post">
              <p><input name="username"></p>
              <p><input name="password" type="password"></p>
              <p><button type="submit">Sign In</button></p>
              </form>'''

@app.route('/signin', methods=['POST']) #���û��������
def signin():
    # ��Ҫ��request�����ȡ�����ݣ�
    if request.form['username']=='admin' and request.form['password']=='password':
        return '<h3>Hello, admin!</h3>'
    return '<h3>Bad username or password.</h3>'

if __name__ == '__main__':
    app.run()

    
ʵ�ʵ�Web AppӦ���õ��û����Ϳ����ȥ���ݿ��ѯ�ٱȶԣ����ж��û��Ƿ��ܵ�¼�ɹ���

����Flask��������Python Web��ܻ��У�

Django��ȫ����Web��ܣ�

web.py��һ��С�ɵ�Web��ܣ�

Bottle����Flask���Ƶ�Web��ܣ�

Tornado��Facebook�Ŀ�Դ�첽Web��ܡ�

��Ȼ�ˣ���Ϊ����Python��Web���Ҳ����ʲô���£����Ǻ���Ҳ�ὲ������Web��ܵ����ݡ�

С��
����Web��ܣ������ڱ�дWebӦ��ʱ��ע�����ʹ�WSGI������ת�Ƶ�URL+��Ӧ�Ĵ���������������дWeb App�͸��Ӽ��ˡ�

�ڱ�дURL������ʱ����������URL�⣬��HTTP�����õ��û�����Ҳ�Ƿǳ���Ҫ�ġ�Web��ܶ��ṩ���Լ���API��ʵ����Щ���ܡ�Flaskͨ��request.form['name']����ȡ�������ݡ�
#-----------------------------------------------------------------------------------------
#ʹ��ģ��
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014320129740415df73bf8f81e478982bf4d5c8aa3817a000

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from flask import Flask, request, render_template

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def home():
    return render_template('home.html')

@app.route('/signin', methods=['GET'])
def signin_form():
    return render_template('form.html')

@app.route('/signin', methods=['POST'])
def signin():
    username = request.form['username']
    password = request.form['password']
    if username=='admin' and password=='password':
        return render_template('signin-ok.html', username=username)
    return render_template('form.html', message='Bad username or password', username=username)

if __name__ == '__main__':
    app.run()

#templates/home.html
<html>
<head>
  <title>Home</title>
</head>
<body>
  <h1 style="font-style:italic">Home</h1>
</body>
</html>

#templates/form.html
<html>
<head>
  <title>Please Sign In</title>
</head>
<body>
  {% if message %}
  <p style="color:red">{{ message }}</p>
  {% endif %}
  <form action="/signin" method="post">
    <legend>Please sign in:</legend>
    <p><input name="username" placeholder="Username" value="{{ username }}"></p>
    <p><input name="password" placeholder="Password" type="password"></p>
    <p><button type="submit">Sign In</button></p>
  </form>
</body>
</html>

#templates/signin-ok.html
<html>
<head>
  <title>Welcome, {{ username }}</title>
</head>
<body>
  <p>Welcome, {{ username }}!</p>
</body>
</html>

ͨ��MVC��������Python�����д���M��Model��C��Controller����V��View��ͨ��ģ�崦��ģ����������Ǿͳɹ��ذ�Python�����HTML��������޶ȵط����ˡ�
ʹ��ģ�����һ��ô��ǣ�ģ��������ܷ��㣬���ң����걣���ˢ����������ܿ������µ�Ч��������ڵ���HTML��CSS��JavaScript��ǰ�˹���ʦ��˵ʵ����̫��Ҫ�ˡ�

��Jinja2ģ���У�������{{ name }}��ʾһ����Ҫ�滻�ı������ܶ�ʱ�򣬻���Ҫѭ���������жϵ�ָ����䣬��Jinja2�У���{% ... %}��ʾָ�

����ѭ�����ҳ�룺

{% for i in page_list %}
    <a href="/page/{{ i }}">{{ i }}</a>
{% endfor %}
���page_list��һ��list��[1, 2, 3, 4, 5]�������ģ�彫���5�������ӡ�

����Jinja2��������ģ�廹�У�

Mako����<% ... %>��${xxx}��һ��ģ�壻

Cheetah��Ҳ����<% ... %>��${xxx}��һ��ģ�壻

Django��Django��һվʽ��ܣ�����һ����{% ... %}��{{ xxx }}��ģ�塣

С��
����MVC�����Ǿͷ�����Python�����HTML���롣HTML����ȫ���ŵ�ģ���д��������Ч�ʡ�
#-----------------------------------------------------------------------------------------
#�첽IO
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00143208573480558080fa77514407cb23834c78c6c7309000
#-----------------------------------------------------------------------------------------
#Э��
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432090171191d05dae6e129940518d1d6cf6eeaaa969000
Python��Э�̵�֧����ͨ��generatorʵ�ֵġ�

��generator�У����ǲ�������ͨ��forѭ���������������Բ��ϵ���next()������ȡ��yield��䷵�ص���һ��ֵ��

����Python��yield�������Է���һ��ֵ���������Խ��յ����߷����Ĳ�����

�������ӣ�

��ͳ��������-������ģ����һ���߳�д��Ϣ��һ���߳�ȡ��Ϣ��ͨ�������ƿ��ƶ��к͵ȴ�����һ��С�ľͿ���������

�������Э�̣�������������Ϣ��ֱ��ͨ��yield��ת�������߿�ʼִ�У���������ִ����Ϻ��л��������߼���������Ч�ʼ��ߣ�

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
def consumer():
    r = ''
    while True:
        nFromProducer = yield r
        if not nFromProducer:
            return
        print('[CONSUMER] Consuming %s...' % nFromProducer)
        r = '200 OK'

def produce(c):
    c.send(None)
    n = 0
    while n < 5:
        n = n + 1
        print('[PRODUCER] Producing %s...' % n)
        r = c.send(n)
        print('[PRODUCER] Consumer return: %s' % r)
    c.close()

if __name__ == '__main__':
    c = consumer()
    produce(c)
�����
[PRODUCER] Producing 1...
[CONSUMER] Consuming 1...
[PRODUCER] Consumer return: 200 OK
[PRODUCER] Producing 2...
[CONSUMER] Consuming 2...
[PRODUCER] Consumer return: 200 OK
[PRODUCER] Producing 3...
[CONSUMER] Consuming 3...
[PRODUCER] Consumer return: 200 OK
[PRODUCER] Producing 4...
[CONSUMER] Consuming 4...
[PRODUCER] Consumer return: 200 OK
[PRODUCER] Producing 5...
[CONSUMER] Consuming 5...
[PRODUCER] Consumer return: 200 OK

ע�⵽consumer������һ��generator����һ��consumer����produce��

���ȵ���c.send(None)������������

Ȼ��һ�������˶�����ͨ��c.send(n)�л���consumerִ�У�

consumerͨ��yield�õ���Ϣ��������ͨ��yield�ѽ�����أ�

produce�õ�consumer����Ľ��������������һ����Ϣ��

produce�����������ˣ�ͨ��c.close()�ر�consumer���������̽�����

����������������һ���߳�ִ�У�produce��consumerЭ������������Գ�Ϊ��Э�̡��������̵߳���ռʽ������

�������Donald Knuth��һ�仰�ܽ�Э�̵��ص㣺

���ӳ������Э�̵�һ����������

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
def generatorTest():
    for i in range(5):
        n = yield i
        print("receiving %s" % n)

if __name__ == '__main__':
    t = generatorTest() #����һ��Generator
    t.send(None) #�����ȶ�һ��Generator����send(None)��������Generator
    m = "abc"
    print(t.send(m)) #�Ժ�ÿ��sendֵʱ�����Ǵ�yield��ʼִ�У�����θ�ֵ��yield�Ⱥ���ߵı���
    t.close()
#-----------------------------------------------------------------------------------------
#asyncio
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/001432090954004980bd351f2cd4cc18c9e6c06d855c498000

asyncio��Python 3.4�汾����ı�׼�⣬ֱ�������˶��첽IO��֧�֡�

asyncio�ı��ģ�;���һ����Ϣѭ�������Ǵ�asyncioģ����ֱ�ӻ�ȡһ��EventLoop�����ã�Ȼ�����Ҫִ�е�Э���ӵ�EventLoop��ִ�У���ʵ�����첽IO��

��asyncioʵ��Hello world�������£�

@asyncio.coroutine��һ��generator���Ϊcoroutine���ͣ�Ȼ�����ǾͰ����coroutine�ӵ�EventLoop��ִ�С�

hello()�����ȴ�ӡ��Hello world!��Ȼ��yield from�﷨���������Ƿ���ص�����һ��generator������asyncio.sleep()Ҳ��һ��coroutine�������̲߳���ȴ�asyncio.sleep()������ֱ���жϲ�ִ����һ����Ϣѭ������asyncio.sleep()����ʱ���߳̾Ϳ��Դ�yield from�õ�����ֵ���˴���None����Ȼ�����ִ����һ����䡣

��asyncio.sleep(1)������һ����ʱ1���IO�������ڴ��ڼ䣬���̲߳�δ�ȴ�������ȥִ��EventLoop����������ִ�е�coroutine�ˣ���˿���ʵ�ֲ���ִ�С�

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import asyncio

@asyncio.coroutine
def hello():
    print("Hello world!")
    # �첽����asyncio.sleep(1):
    r = yield from asyncio.sleep(1)
    print("Hello again!")

if __name__ == '__main__':
    # ��ȡEventLoop:
    loop = asyncio.get_event_loop()
    # ִ��coroutine
    loop.run_until_complete(hello())
    loop.close()

������Task��װ����coroutine���ԣ�
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import threading
import asyncio

@asyncio.coroutine
def hello(s):
    print('Hello world! (%s)' % threading.currentThread())
    print("Before %s" % s)
    yield from asyncio.sleep(3)
    print("After %s" % s)
    print('Hello again! (%s)' % threading.currentThread())

if __name__ == '__main__':
    loop = asyncio.get_event_loop()
    tasks = [hello(1), hello(2), hello(3), hello(4)]
    loop.run_until_complete(asyncio.wait(tasks))
    loop.close()

�۲�ִ�й��̣�
Hello world! (<_MainThread(MainThread, started 2640)>)
Before 4
Hello world! (<_MainThread(MainThread, started 2640)>)
Before 2
Hello world! (<_MainThread(MainThread, started 2640)>)
Before 1
Hello world! (<_MainThread(MainThread, started 2640)>)
Before 3
Լ��3��
After 4
Hello again! (<_MainThread(MainThread, started 2640)>)
After 1
Hello again! (<_MainThread(MainThread, started 2640)>)
After 2
Hello again! (<_MainThread(MainThread, started 2640)>)
After 3
Hello again! (<_MainThread(MainThread, started 2640)>)
�ɴ�ӡ�ĵ�ǰ�߳����ƿ��Կ���������coroutine����ͬһ���̲߳���ִ�еġ�

�����asyncio.sleep()����������IO����������coroutine�Ϳ�����һ���̲߳���ִ�С�

������asyncio���첽������������ȡsina��sohu��163����վ��ҳ��
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import asyncio

@asyncio.coroutine
def wget(host):
    print('wget %s...' % host)
    connect = asyncio.open_connection(host, 80)

    print("Before yield from connect processing host [%s]" % host)
    reader, writer = yield from connect
    print("After yield from connect processing host [%s]" % host)

    header = 'GET / HTTP/1.0\r\nHost: %s\r\n\r\n' % host
    writer.write(header.encode('utf-8'))
    yield from writer.drain()
    while True:
        line = yield from reader.readline()
        if line == b'\r\n':
            break
        print('%s header > %s' % (host, line.decode('utf-8').rstrip()))
    # Ignore the body, close the socket
    writer.close()

if __name__ == '__main__':
    loop = asyncio.get_event_loop()
    tasks = [wget(host) for host in ['www.sina.com.cn', 'www.sohu.com', 'www.163.com']]
    loop.run_until_complete(asyncio.wait(tasks))
    loop.close()
�����
wget www.sohu.com...
Before yield from connect processing host [www.sohu.com]
wget www.sina.com.cn...
Before yield from connect processing host [www.sina.com.cn]
wget www.163.com...
Before yield from connect processing host [www.163.com]
After yield from connect processing host [www.sohu.com]
www.sohu.com header > HTTP/1.1 200 OK
www.sohu.com header > Content-Type: text/html;charset=UTF-8
www.sohu.com header > Connection: close
www.sohu.com header > Server: nginx
www.sohu.com header > Date: Sun, 15 Apr 2018 07:02:14 GMT
www.sohu.com header > Cache-Control: max-age=60
www.sohu.com header > X-From-Sohu: X-SRC-Cached
www.sohu.com header > Content-Encoding: gzip
www.sohu.com header > FSS-Cache: HIT from 13539701.18454911.21477824
www.sohu.com header > FSS-Proxy: Powered by 10131777.11639115.18069848
After yield from connect processing host [www.sina.com.cn]
After yield from connect processing host [www.163.com]
www.sina.com.cn header > HTTP/1.1 200 OK
www.sina.com.cn header > Server: nginx
www.sina.com.cn header > Date: Sun, 15 Apr 2018 07:02:36 GMT
www.sina.com.cn header > Content-Type: text/html
www.sina.com.cn header > Content-Length: 603236
www.sina.com.cn header > Connection: close
www.sina.com.cn header > Last-Modified: Sun, 15 Apr 2018 07:00:39 GMT
www.sina.com.cn header > Vary: Accept-Encoding
www.sina.com.cn header > X-Powered-By: shci_v1.03
www.sina.com.cn header > Expires: Sun, 15 Apr 2018 07:03:28 GMT
www.sina.com.cn header > Cache-Control: max-age=60
www.sina.com.cn header > Age: 2
www.sina.com.cn header > Via: http/1.1 ctc.ningbo.ha2ts4.97 (ApacheTrafficServer/6.2.1 [cHs f ]), http/1.1 ctc.ningbo.ha2ts4.106 (ApacheTrafficServer/6.2.1 [cHs f ])
www.sina.com.cn header > X-Via-Edge: 152377575604889f65d65eebeee734a4224ac
www.sina.com.cn header > X-Cache: HIT.106
www.sina.com.cn header > X-Via-CDN: f=edge,s=ctc.ningbo.ha2ts4.102.nb.sinaedge.com,c=101.93.246.137;f=Edge,s=ctc.ningbo.ha2ts4.106,c=115.238.190.102
www.163.com header > HTTP/1.1 200 OK
www.163.com header > Expires: Sun, 15 Apr 2018 07:03:56 GMT
www.163.com header > Date: Sun, 15 Apr 2018 07:02:36 GMT
www.163.com header > Server: nginx
www.163.com header > Content-Type: text/html; charset=GBK
www.163.com header > Vary: Accept-Encoding,User-Agent,Accept
www.163.com header > Cache-Control: max-age=80
www.163.com header > X-Via: 1.1 qzh148:2 (Cdn Cache Server V2.0), 1.1 PSshqzdx2mj185:2 (Cdn Cache Server V2.0)
www.163.com header > Connection: close

С��
asyncio�ṩ�����Ƶ��첽IO֧�֣�

�첽������Ҫ��coroutine��ͨ��yield from��ɣ�

���coroutine���Է�װ��һ��TaskȻ�󲢷�ִ�С�

�ο�����ѧϰЭ��
#http://python.jobbole.com/87310/
#-----------------------------------------------------------------------------------------
#async/await
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/00144661533005329786387b5684be385062a121e834ac7000

��asyncio�ṩ��@asyncio.coroutine���԰�һ��generator���Ϊcoroutine���ͣ�Ȼ����coroutine�ڲ���yield from������һ��coroutineʵ���첽������

Ϊ�˼򻯲����õر�ʶ�첽IO����Python 3.5��ʼ�������µ��﷨async��await��������coroutine�Ĵ��������׶���

��ע�⣬async��await�����coroutine�����﷨��Ҫʹ���µ��﷨��ֻ��Ҫ�������򵥵��滻��

��@asyncio.coroutine�滻Ϊasync��
��yield from�滻Ϊawait��
�����ǶԱ�һ����һ�ڵĴ��룺

@asyncio.coroutine
def hello():
    print("Hello world!")
    r = yield from asyncio.sleep(1)
    print("Hello again!")
�����﷨���±�д���£�

async def hello():
    print("Hello world!")
    r = await asyncio.sleep(1)
    print("Hello again!")
#-----------------------------------------------------------------------------------------
#aiohttp
#https://www.liaoxuefeng.com/wiki/0014316089557264a6b348958f449949df42a6d3a2e542c000/0014320981492785ba33cc96c524223b2ea4e444077708d000

#-----------------------------------------------------------------------------------------
#sklearn ѧϰ 
https://www.bilibili.com/video/av17003173/?p=8

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import print_function
from sklearn import datasets
from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier

if __name__ == '__main__':
    iris = datasets.load_iris()
    iris_X = iris.data
    iris_y = iris.target

    # print(iris_X[:2, :])
    # print(iris_y)

    X_train, X_test, y_train, y_test = train_test_split(iris_X, iris_y, test_size=0.3)
    # print(y_train)

    knn = KNeighborsClassifier() #����KNN������
    knn.fit(X_train, y_train) #ѵ�����ݼ�
    y_predict = knn.predict(X_test) #���ݲ��Լ�����Ԥ����
    print(y_predict)
    print(y_test)

    # �����Լ�д����������
    correctNum = sum(map(lambda x, y: x == y, y_predict, y_test))
    correctRatio = correctNum / len(y_test)
    print(correctRatio)

    # Sklearn�Դ�����������
    score = knn.score(X_test, y_test)
    print(score)
#-----------------------------------------------------------------------------------------
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from sklearn import datasets
from sklearn.linear_model import LinearRegression

if __name__ == '__main__':
    load_data = datasets.load_boston()
    data_X = load_data.data
    data_y = load_data.target
    #     print(data_X)
    #     print(data_y)
    model = LinearRegression()
    model.fit(data_X, data_y)
    
#     print(model.predict(data_X[:4]))
#     print(data_y[:4])

#     print(model.coef_)
#     print(model.intercept_)
#     print(model.get_params())
    print(model.score(data_X, data_y)) #R^2 coefficient of determination
#-----------------------------------------------------------------------------------------
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from sklearn import datasets
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt

if __name__ == '__main__':
    X, y = datasets.make_regression(n_samples=200, n_features=1, n_targets=1, noise=1)
    plt.scatter(X, y)
    plt.show()
    
    model = LinearRegression()
    model.fit(X, y)
    
    score = model.score(X, y)
    print(model.predict(X[:4]))
    print(y[:4])
    print(score)
#-----------------------------------------------------------------------------------------
7. normalization ��׼������

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from sklearn import preprocessing
import numpy as np
from sklearn.datasets import make_classification
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.svm import SVC

if __name__ == '__main__':
    # a = np.array([[10, 2.7, 3.6],
    #              [-100, 5, -2],
    #              [120, 20, 40]], dtype=np.float64)
    # print(a)
    # print(preprocessing.scale(a))

    X, y = make_classification(n_samples=300, n_features=2, n_redundant=0, n_informative=2, random_state=22,
                               n_clusters_per_class=1, scale=100)
    # plt.scatter(X[:, 0], X[:, 1], c=y)
    # plt.show()
    # X = preprocessing.scale(X) #ȥ�����б�׼��������׼ȷ�ʴ���½�
    # X = preprocessing.minmax_scale(X, feature_range=(-1, 1))
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=.3)
    clf = SVC()
    clf.fit(X_train, y_train)
    print(clf.score(X_test, y_test))
#-----------------------------------------------------------------------------------------
8. cross_validation 1

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from sklearn.model_selection import cross_val_score
from sklearn.datasets import load_iris # iris���ݼ�
from sklearn.model_selection import train_test_split # �ָ�����ģ��
from sklearn.neighbors import KNeighborsClassifier # K�����(kNN��k-NearestNeighbor)�����㷨
import matplotlib.pyplot as plt

if __name__ == '__main__':
    # ����iris���ݼ�
    iris = load_iris()
    X = iris.data
    y = iris.target

    # �ָ����ݲ�
    # X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=4)
    #
    # # ����ģ��
    # knn = KNeighborsClassifier(n_neighbors=5)
    #
    # # ѵ��ģ��
    # knn.fit(X_train, y_train)
    #
    # # ��׼ȷ�ʴ�ӡ��
    # print(knn.score(X_test, y_test))

    # �ֳɶ�����Լ���ѵ����
    # knn = KNeighborsClassifier(n_neighbors=5)
    # scores = cross_val_score(knn, X, y, cv=5, scoring='accuracy')
    # print(scores.mean()) # print(scores)

    k_range = range(1, 31)
    k_scores = []
    for k in k_range:
        knn = KNeighborsClassifier(n_neighbors=k)
        # scores = cross_val_score(knn, X, y, cv=10, scoring='accuracy') # for classification
        # k_scores.append(scores.mean())
        loss = -cross_val_score(knn, X, y, cv=10, scoring='mean_squared_error')  # for regression
        k_scores.append(loss.mean())

    plt.plot(k_range, k_scores)
    plt.xlabel('Value of K for KNN')
    plt.ylabel('Cross-Validated Accuracy')
    plt.show()

#-----------------------------------------------------------------------------------------
9. cross_validation 2

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import matplotlib.pyplot as plt
import numpy as np
from sklearn.datasets import load_digits
from sklearn.model_selection import learning_curve
from sklearn.svm import SVC

if __name__ == '__main__':
    digits = load_digits()
    X = digits.data
    y = digits.target

    train_sizes, train_loss, test_loss = learning_curve(
        SVC(gamma=0.001), X, y, cv=10, scoring='mean_squared_error',
        train_sizes=[0.1, 0.25, 0.5, 0.75, 1])

    # ƽ��ÿһ�����õ���ƽ������(��5�֣��ֱ�Ϊ����10%��25%��50%��75%��100%)
    train_loss_mean = -np.mean(train_loss, axis=1)
    test_loss_mean = -np.mean(test_loss, axis=1)

    plt.plot(train_sizes, train_loss_mean, 'o-', color="r", label="Training")
    plt.plot(train_sizes, test_loss_mean, 'o-', color="g", label="Cross-validation")

    plt.xlabel("Training examples")
    plt.ylabel("Loss")
    plt.legend(loc="best")
    plt.show()

#-----------------------------------------------------------------------------------------
10. cross_validation 3
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import matplotlib.pyplot as plt
import numpy as np
from sklearn.datasets import load_digits
from sklearn.model_selection import validation_curve
from sklearn.svm import SVC

if __name__ == '__main__':
    digits = load_digits()
    X = digits.data
    y = digits.target

    param_range = np.logspace(-6, -2.3, 5)
    train_loss, test_loss = validation_curve(
        SVC(), X, y, param_name='gamma', param_range=param_range, cv=10,
        scoring='mean_squared_error')

    # ƽ��ÿһ�����õ���ƽ������(��5�֣��ֱ�Ϊ����10%��25%��50%��75%��100%)
    train_loss_mean = -np.mean(train_loss, axis=1)
    test_loss_mean = -np.mean(test_loss, axis=1)

    plt.plot(param_range, train_loss_mean, 'o-', color="r", label="Training")
    plt.plot(param_range, test_loss_mean, 'o-', color="g", label="Cross-validation")

    plt.xlabel("Gamma")
    plt.ylabel("Loss")
    plt.legend(loc="best")
    plt.show()

#-----------------------------------------------------------------------------------------
11 save
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from sklearn import svm
from sklearn import datasets
import pickle

from sklearn.externals import joblib

if __name__ == '__main__':
    clf = svm.SVC()
    digits = datasets.load_digits()
    X, y= digits.data, digits.target
    clf.fit(X, y)
    #
    # save
    # # method 1: pickle
    # with open('save/clf.pickle', 'wb') as f:
    #     pickle.dump(clf, f)
    #
    # restore
    # with open('save/clf.pickle', 'rb') as f:
    #     clf2 = pickle.load(f)
    #     print(clf2.predict(X[0:1]))

    # method 2: joblib
    # save
    joblib.dump(clf, 'save/clf.pkl')
    # restore
    clf3 = joblib.load('save/clf.pkl') #�ٶȻ����Щ
    print(clf3.predict(X[0:1]))
#-----------------------------------------------------------------------------------------
# str.format https://www.cnblogs.com/Alexzzzz/p/6832253.html
format�����������ַ����ĸ�ʽ�������

 print('{0}+{1}={2}'.format(1,2,1+2))   #in
1+2=3   #out
�ɼ��ַ����д������ڵ����ֱַ��Ӧ��format�ļ���������

��ʡ�����֣�
print('{}+{}={}'.format(1,2,1+2))   #in
���Եõ�ͬ�����������������滻˳��Ĭ�ϰ���[0],[1],[2]...���С�

���滻{0}��{1}��
print('{1}+{0}={2}'.format(1,2,1+2))   #in
2+1=3   #out

����ַ���:
print('{0} am {1}'.format('i','alex'))  
i am alex   #out

���������ֵ:
length = 4
name = 'alex'
print('the length of {0} is {1}'.format(name,length))
the length of alex is 4

���ȿ��ƣ�
print('{0:.3}'.format(1/3))
0.333

��ȿ��ƣ�
print('{0:7}{1:7}'.format('use','python'))
use    python 

����ȿ���(����ھ���)��
print('{0:<7.3}..'.format(1/3))   
0.333  ..
��ʵ����ȿ��ƺ�������C�е�printf������
ͬ��'>'Ϊ���ң�'^'Ϊ���С����ź�����

��ȫ��
#!/usr/bin/python
#python3.6
print('{0:0>3}'.format(1)) #���ң������0��ȫ
print('{0:{1}>3}'.format(1,0))  #Ҳ������ôд
#���������ʹ�ÿո�ȫ��ʱ��ϵͳ���Զ�����Ӣ�Ŀո�����ܻ���ɲ�����
#for example
blog = {'1':'�й�ʯ�ʹ�ѧ','2':'�㽭��ѧ','3':'�Ͼ����պ����ѧ'}
print('�����룺')
print('{0:^4}\t\t{1:^8}'.format('���','����'))
for no,name in blog.items(): #�ֵ��items()��������һ����ֵ�ԣ��ֱ�ֵ��no��name
    print('{0:^4}\t\t{1:^8}'.format(no,name))
print('\n���룺')
print('{0:^4}\t\t{1:{2}^8}'.format('���','����',chr(12288))) #chr(12288)ΪUTF-8�е����Ŀո�
for no,name in blog.items():
    print('{0:^4}\t\t{1:{2}^8}'.format(no,name,chr(12288)))
#out
001
001
�����룺
 ���              ����   
 1               �й�ʯ�ʹ�ѧ 
 2                �㽭��ѧ  
 3              �Ͼ����պ����ѧ

���룺
 ���           ���������ơ�����
 1              ���й�ʯ�ʹ�ѧ��
 2              �����㽭��ѧ����
 3              �Ͼ����պ����ѧ
#-----------------------------------------------------------------------------------------
# pdb���� https://www.cnblogs.com/xiaohai2003ly/p/8529472.html
#pytest.py
def add(a, b):
    return a + b

s = '5'
n = int(s)
c = add(n, 10)
print(c)

Python ������֮pdb
ʹ��PDB�ķ�ʽ������:
1. ����ִ�д���,ͨ������ python -m pdb xxx.py �����ű������뵥��ִ��ģʽ
 pdb�����У�
1������������Debugģʽ��python -m pdb xxx.py
2��h����help������
3��w����where����ӡ��ǰִ�ж�ջ
4��d����down��ִ����ת���ڵ�ǰ��ջ����һ�㣨���뵽��һ������ջ���������Ϊ��һ��������
5��u����up��ִ����ת����ǰ��ջ����һ�㣨���ص���һ���ջ���������Ϊɾ����ǰ��ջ�壩
6��b����break����Ӷϵ� ��������ӵĶϵ㣬ֻҪ�����˳�����ϵ�һֱ���ڣ����exit������q�������õĶϵ��ʧЧ��
             b �г���ǰ���жϵ㣬�Ͷϵ�ִ�е�ͳ�ƴ���
             b line_no����ǰ�ű���line_no����Ӷϵ�
             b filename:line_no���ű�filename��line_no����Ӷϵ�
             b function���ں���function�ĵ�һ����ִ����䴦��Ӷϵ�
7��tbreak����temporary break����ʱ�ϵ�
             �ڵ�һ��ִ�е�����ϵ�֮�󣬾��Զ�ɾ������ϵ㣬�÷���bһ��
8��cl����clear������ϵ�
            cl ������жϵ�
            cl bpnumber1 bpnumber2... ����ϵ��Ϊbpnumber1,bpnumber2...�Ķϵ�
            cl lineno �����ǰ�ű�lineno�еĶϵ�
            cl filename:line_no ����ű�filename��line_no�еĶϵ�
9��disable��ͣ�öϵ㣬����Ϊbpnumber����cl�������ǣ��ϵ���Ȼ���ڣ�ֻ�ǲ�����
10��enable������ϵ㣬����Ϊbpnumber
11��s����step��ִ����һ������ ��ִ����һ����������һ��Ϊ�����������뵽�����У�
            ��������Ǻ������ã���s��ִ�е������ĵ�һ��
12��n����next��ִ����һ����� ��ִ����һ����������һ��Ϊ���������ִ���꺯����
            ��������Ǻ������ã���ִ�к���������ִ�е�ǰִ��������һ����
13��r����return��ִ�е�ǰ���к��������� ���Ѿ��ߵ���ǰ�����Ľ�����䣬�޷���ʹ��j linenumber���ڵ�ǰ��������ת�ˣ�
14��c����continue������ִ�У�ֱ��������һ���ϵ�
15��l����list���г�Դ�� ������ll��䣩
             l �г���ǰִ�������Χ11������
             l first �г�first����Χ11������
             l first, second �г�first--second��Χ�Ĵ��룬���second<first��second��������Ϊ���� ����l 10,15�����г�10��15�У�
16��a����args���г���ǰִ�к����ĺ���
17��p expression����print�����expression��ֵ
18��pp expression���ÿ�һ���p expression
19��run����������debug���൱��restart
20��q����quit���˳�debug
21��j lineno����jump����������ִ�е���亯�� ����������ת��䣬������ת��ִ�й������������ִ�У�Ҳ������ת��������䣬���м���������䶼���Թ���ִ�У�
            ֻ���ڶ�ջ����ײ���ת���������ִ�У���ǰ��ֱ��ִ�е��к�
22��unt����until��ִ�е���һ�У�����ѭ���������ߵ�ǰ��ջ���� ��ִ�е��ض���䣬����кţ��磺unt 7��ʹ��untҲ����ִ�е���ǰ���������һ�У����ǲ�û���ߵ�����������䣬����ʹ��j linenumber����ת��
23��condition bpnumber conditon�����ϵ�����������������condition����True��ʱ��bpnumber�ϵ���Ч������bpnumber�ϵ���Ч ��ֻ���������õĶϵ����������������ҵ�2�������Ƕϵ�ţ��磺condition 3 a == 4����3�Ŷϵ���������a == 4ʱ��

ע�⣺
1��ֱ������Enter����ִ����һ�����
2������PDB����ʶ�����PDB���������Python����ڵ�ǰ������ִ�У�



2. pdb����ִ��̫�鷳�ˣ����Եڶ��ַ�����import pdb ֮��ֱ���ڴ�������Ҫ���Եĵط���һ��pdb.set_trace()���Ϳ�������һ���ϵ㣬 �������pdb.set_trace()��ͣ������pdb���Ի�����������pdb �������鿴����������c��������
#pdbtest.py
import pdb

def add(a, b):
    print("hello, world")
    print("good day")
    pdb.set_trace()
    return a + b

def loop(n):
    i = 1
    pdb.set_trace()
    while (i < n):
        m = i * i
        print(m)
        i += 1
    print("i = {}".format(i))

s = '2'
n = int(s)
c = add(n, 10)
c += c
loop(c)
print(c)

���ʹ�ò��� -m pdb��������������ģʽ�����ֹͣ�ڵ�һ�У����ҳ������й����е�print()��Ϣ������ʾ������
python -m pdb pdbtest.py

�����ʹ�ò�����������ֱ���������У����һ�ֹͣ�ڴ���pdb.set_trace()�ĺ�һ�䣬Ҳ��ѳ����е�print()��Ϣ�������������̨��
python pdbtest.py


#https://blog.csdn.net/qq_21398167/article/details/52325464
�ڵ��Ե�ʱ��̬�ı�ֵ ���ڵ��Ե�ʱ����Զ�̬�ı������ֵ����������ʵ������Ҫע����������и�����ԭ���� b �Ѿ�����ֵ�ˣ���������¸ı� b �ĸ�ֵ����Ӧ��ʹ�ã� B��
�嵥 9. �ڵ��Ե�ʱ��̬�ı�ֵ
[root@rcc-pok-idg-2255 ~]# python epdb2.py 
 > /root/epdb2.py(10)?() 
 -> b = "bbb"
 (Pdb) var = "1234"
 (Pdb) b = "avfe"
 *** The specified object '= "avfe"' is not a function 
 or was not found along sys.path. 
 (Pdb) !b="afdfd"
 (Pdb)
#-----------------------------------------------------------------------------------------
# python 3.5�����TypeHint ֻ������̬��飬�޷���������������
# https://blog.csdn.net/sunt2018/article/details/83022493
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ���������str������������str
def get(param: str) -> object:
    if param == "hello":
        return 1
    return param * 2


class Test(object):
    @staticmethod
    def hello() -> None:
        print("hello")


if __name__ == '__main__':
    my_str = get("hello")
    print(my_str)

    i: int = get(5)
    print(i)

    Test.hello()
    hello: Test = Test() # ���еı����󶼿���ð����ָ������
    hello.hello()
#-----------------------------------------------------------------------------------------
# Python3��ʹ��urllib����https��վ
import ssl
import urllib.request

if __name__ == '__main__':
    url = "http://baike.baidu.com/item/Python/407313"

    ssl._create_default_https_context = ssl._create_unverified_context
    with urllib.request.urlopen(url) as response1:
        print(response1.getcode())
        print(len(response1.read()))
#-----------------------------------------------------------------------------------------
# Python3 urljoin
#
from urllib.parse import urljoin

if __name__ == '__main__':
    url = r'https://baike.baidu.com/item/Python/407313'
    new_url = r'/item/�붮�ǿ���'
    new_full_url = urljoin(url, new_url)  # ƴ��URL
    print(new_full_url)
�����
https://baike.baidu.com/item/�붮�ǿ���

#
from posixpath import normpath
from urllib.parse import urljoin
from urllib.parse import urlparse
from urllib.parse import urlunparse


def myjoin(base, url):
    url1 = urljoin(base, url)
    arr = urlparse(url1)
    path = normpath(arr[2])
    return urlunparse((arr.scheme, arr.netloc, path, arr.params, arr.query, arr.fragment))


if __name__ == '__main__':
    print(myjoin("http://www.baidu.com/lmn", "def/lmn/abc.html"))
    print(myjoin("http://www.baidu.com", "/../../abc.html"))
    print(myjoin("http://www.baidu.com/xxx", "./../../abc.html"))
    print(myjoin("http://www.baidu.com", "abc.html?key=value&m=x"))
�����
http://www.baidu.com/def/lmn/abc.html
http://www.baidu.com/abc.html
http://www.baidu.com/abc.html
http://www.baidu.com/abc.html?key=value&m=x
#-----------------------------------------------------------------------------------------
# urlopen��������bytes��stringת��
# �ο� https://blog.csdn.net/haoxizh/article/details/44649451
import contextlib
import urllib.request

if __name__ == '__main__':
    url = 'http://www.baidu.com'
    with contextlib.closing(urllib.request.urlopen(url)) as response:  # ʹ��contextlib.closing()Ҳ���ԣ�ֱ��ʹ��with as Ҳ���ԣ�˵��urlopen()���صĶ������__enter__()������__exit__()������close()����������ͨ��help(response)���鿴
        page = response.read()
        the_page = page.decode("UTF-8")
        b_page = the_page.encode("UTF-8")
        print(f"page type is {type(page)}, the_page type is {type(the_page)}, b_page type is {type(b_page)}")
�����
page type is <class 'bytes'>, the_page type is <class 'str'>, b_page type is <class 'bytes'>
#-----------------------------------------------------------------------------------------
# unitteset ���� https://blog.csdn.net/xiaoxinyu316/article/details/53170463
# 1��ͨ��unittest.main()��ִ�в��������ķ�ʽ��
#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import unittest


class UCTestCase(unittest.TestCase):
    def setUp(self):
        # ����ǰ��ִ�еĲ���
        print("setUp")
        pass

    def tearDown(self):
        # ��������ִ���������ִ�еĲ���
        print("tearDown")
        pass

    # ��������1
    def testCreateFolder(self):
        # ����Ĳ��Խű�
        print("testCreateFolder")

    # ��������2
    def testDeleteFolder(self): # ֻ����test��ʼ�ĺ������Ż����У�����Ѻ�������Ϊt2estXXX���򲻻�����
        # ����Ĳ��Խű�
        print("testDeleteFolder")


if __name__ == "__main__":
    unittest.main()

    
# 2��ͨ��testsuit��ִ�в��������ķ�ʽ��
if __name__ == "__main__":
    # ������Լ�
    suite = unittest.TestSuite()
    suite.addTest(UCTestCase("testCreateFolder"))
    # suite.addTest(UCTestCase("testDeleteFolder"))
    # ִ�в���
    runner = unittest.TextTestRunner()
    runner.run(suite)

# ��PyCharm������
�����ԣ�ֱ��ʹ��PyCharm Ctrl+Shift+F10���У���ʹֻaddһ��Test���Ի�����ȫ��������������ʱ���õ�Configuration���Python tests���еġ�

�����Configuration���޸�ΪPython���У���ᰴ������addTest�����С�

# ��������������
�����������з�ʽ���У�python src\test\unitest_learn.py ��ʱֻ������һ���������������£�
D:\Program Files\JetBrains\PythonProject\Py3TestProject>python src\test\unitest_learn.py
setUp
testCreateFolder
tearDown
.
----------------------------------------------------------------------
Ran 1 test in 0.002s

OK

����Ӳ��� -m unittest�����Ի�����ȫ��������
D:\Program Files\JetBrains\PythonProject\Py3TestProject>python -m unittest src\test\unitest_learn.py
setUp
testCreateFolder
tearDown
.setUp
testDeleteFolder
tearDown
.
----------------------------------------------------------------------
Ran 2 tests in 0.005s

OK

˵������suite.addTest(UCTestCase("testCreateFolder"))���ַ�ʽ�����в�����������ʹ����������test���أ��Կ����������С�


# 3��ͨ��testLoader��ʽ��
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import unittest


class TestCase1(unittest.TestCase):
    # def setUp(self):
    # def tearDown(self):
    def testCase1(self):
        print('aaa')

    def testCase2(self):
        print('bbb')


class TestCase2(unittest.TestCase):
    # def setUp(self):
    # def tearDown(self):
    def testCase1(self):
        print('aaa1')

    def testCase2(self):
        print('bbb1')


if __name__ == "__main__":
    # ���÷�����ͬʱ���Զ����
    suite1 = unittest.TestLoader().loadTestsFromTestCase(TestCase1)
    suite2 = unittest.TestLoader().loadTestsFromTestCase(TestCase2)
    suite = unittest.TestSuite([suite1, suite2])
    unittest.TextTestRunner(verbosity=2).run(suite)

˵����unittest.TextTestRunner(verbosity=2).run(suite)����������TextTestRunner�ཫ����ִ�еĽ����text��ʽ�����verbosityĬ��ֵΪ1����������������������������ɹ������.��,ʧ�������F��,���������E��;verbosity=2�������������Ϣ��verbosity=2��ָ���Խ�����������ϸ�̶ȣ���0-6�����������ʵ�ֿɿ�Python27\Lib\unittest\runner.pyԴ���롣 ��https://cloud.tencent.com/info/2ab386c4b78655f1f5d277023d702983.html��
#-----------------------------------------------------------------------------------------
# scrapy xpath()�е�/��//������ https://blog.csdn.net/changer_WE/article/details/84553986

from lxml import etree
myhtml = """
    <body>
        <div id="div1">
            <div id="div2">
                <p>hello world</p>
                <div id="div3">
                    <a href="xxxA.com">ת��A</a>
                    <a href="xxxB.com">ת��B</a>
                </div>    
            </div>
            <h1>�������</h1>
        </div>
    </body>
    """
selector = etree.HTML(myhtml)
content1 = selector.xpath('//div[@id="div1"]/p/text()')  # ��/��ȡp��ǩ�����ݣ���ȡ���� 
print(content1)      # /����µ�xpath������
content2 = selector.xpath('//div[@id="div1"]//p/text()')  # ��//��ȡp��ǩ������
print(content2)      # //����µ�xpath������
content3 = selector.xpath('//div[@id="div1"]/div[@id="div2"]/div[@id="div3"]/a/text()')  # ��/��ȡa��ǩ������
print(content3)
content4 = selector.xpath('//div[@id="div1"]//a/text()')  # ��//��ȡa��ǩ������

���ʵ����Կ�����
/ ? ?��������������˵�ľ���·��
// ? ��һ���ݴ��Ը��ߵ�д�������Բ��ܾ��壬���Կ���ܶ��
����˵����������� ?// ? д�ģ��;������� ?/ ?������Լ������һЩ�ݴ�ָ�����ͻ�������Լ�����ʱ��Ĵ���һЩ��ҳdiv����div��
������ģʽ��

//a//b/@abc ָ�����ĵ�������aԪ�ص�����Ϊabc�ĺ��bԪ�أ������Ӵ�Ԫ�أ����༶����
//a/b/@abc ָ�����ĵ�������aԪ�ص�����Ϊabc���Ӵ�bԪ�أ�һ������
/a/b/@abc ָ���Ǹ��ڵ�bԪ�ص�����Ϊabc���Ӵ�bԪ�أ�һ������

ͨ���õ�һ�־ͺ���
#-----------------------------------------------------------------------------------------
# pymysql����mysql http://www.runoob.com/python3/python3-mysql.html
# mysql_test.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pymysql

if __name__ == '__main__':
    connect = pymysql.connect(host='localhost', port=3306, user='root', passwd='root123',
                              db='hibernate', charset='utf8')
    cursor = connect.cursor()
    # ʹ�� execute()  ����ִ�� SQL ��ѯ
    cursor.execute("SELECT VERSION()")

    # ʹ�� fetchone() ������ȡ��������.
    # data = cursor.fetchone()
    #
    # print("Database version : %s " % data)  # Database version : 8.0.11 

    data = {'serial_number': '248', 'movie_name': 'ǹ��', 'introduce': '1999/���/���鶯������', 'star': '8.7',
            'evaluate': '126052������', 'description': 'һȺ�ݼ���տ��Ϸ�ǣ����׳�һ�����µĺڰ�СƷ���ɾͶ�����ȺϷ���۷�֮����'}

    sql_list = """
    insert into douban_movie(serial_number, movie_name, introduce, star, evaluate, description)
        values(%s, %s, %s, %s, %s, %s)
    """

    cursor.execute(sql_list, (data['serial_number'], data['movie_name'], data['introduce'], data['star'], data['evaluate'], data['description']))

    # sql_dict = """
    # insert into douban_movie(%(name)s, %(name)s, %(name)s, %(name)s, %(name)s, %(name)s)
    #     values(%s, %s, %s, %s, %s, %s)
    # """
    # cursor.execute(sql_dict, data)  # û�ҵ�dictʹ������

    connect.commit()
    # �ر����ݿ�����
    connect.close()
#-----------------------------------------------------------------------------------------
# Python����ʹ�ú������� https://www.cnblogs.com/erbaodabao0611/p/7490439.html
def test(a):
    print(a)


def test(a, b):
    print(a, b)
    
test(1)
test(1, 2)
���б���
Traceback (most recent call last):
  File "D:/Program Files/JetBrains/PythonProject/Py3TestProject/src/test/main.py", line 33, in <module>
    test(1)
TypeError: test() missing 1 required positional argument: 'b'
#-----------------------------------------------------------------------------------------
# �ж��ַ������Ƿ�������
# main.py
def has_chinese(word):
    for ch in word:
        if '\u4e00' <= ch <= '\u9fff':
            return True
    return False

# test_has_chinese.py
from unittest import TestCase

from src.test.main import has_chinese


class TestHas_chinese(TestCase):
    def test_str_all_chars(self):
        s = 'abc'
        self.assertFalse(has_chinese(s))

    def test_str_has_chinese(self):
        s = '��ã�world'
        self.assertTrue(has_chinese(s))

#-----------------------------------------------------------------------------------------
# 299. Bulls and Cows https://leetcode.com/problems/bulls-and-cows/
#
import collections


class Solution:
    def getHint(self, secret: str, guess: str) -> str:
        # �����㷨
        A = sum(s == g for s, g in zip(secret, guess))
        B = sum((collections.Counter(secret) & collections.Counter(guess)).values()) - A
        return "{A}A{B}B".format(A=A, B=B)

#
from unittest import TestCase

from src.test.main import Solution


class TestSolution(TestCase):
    def test_getHint(self):
        output = Solution().getHint("1807", "7810")
        self.assertEqual(output, '1A3B')
#-----------------------------------------------------------------------------------------
# �б��Ƶ�ʽ����dict��map��������dict
if __name__ == '__main__':
    d1 = {i: i for i in [1, 2]}
    print(d1)
    d2 = dict(map(lambda i: (i, i), [1, 2]))
    print(d2)

#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------


#-----------------------------------------------------------------------------------------

